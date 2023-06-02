import time
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import collections

import afccp.core.comprehensive_functions

def AFSC_scoring_data_structure(instance):
    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

        AFSC_scores[a] = afsc_scored_cadets
    return AFSC_scores

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit
    # are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the
    # "second matching".


def cadet_scoring_data_structure(instance):
    # My Score Calculations for CFMs
    np.random.seed(2)
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]

    cadet_merit_list = [m for m in instance.parameters['merit']]

    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # If wanted sorted from best to worst. We do this later according to cadet rank.
    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    return cadet_scores, unmatched_ordered_cadets


def build_ranking_data_structures(instance):
    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    AFSC_scores = AFSC_scoring_data_structure(instance)
    cadet_scores, unmatched_ordered_cadets = cadet_scoring_data_structure(instance)

    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # i.e., the cade is not ineligible.
                cadet_scoring[c][a] = score

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this function to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    return cadet_score_from_AFSC, cadet_scoring, cadet_ranking, afsc_scoring, afsc_ranking


def set_afsc_capacities(instance, capacity_type='lower_quotas'):
    afscs = instance.parameters['afsc_vector']
    if capacity_type == 'lower_quotas':
        afscs = instance.parameters['afsc_vector']
        lower_quotas = instance.parameters['quota_min']
        afsc_capacities = dict(zip(afscs, lower_quotas))
        return afsc_capacities
    if capacity_type == 'max':
        max_capacities = instance.parameters["quota_max"]
        afsc_capacities = dict(zip(afscs, max_capacities))
        return afsc_capacities
    if capacity_type == 'uncapacitated':
        uncapacitated_capacities = [500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500,
                                    500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500]
        afsc_capacities = dict(zip(afscs, uncapacitated_capacities))
        return afsc_capacities


def afsc_matches_build(instance):
    """Builds afsc_matches dictionary with each AFSC as a key and its values as an empty list"""
    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i
    return afsc_matches


def highest_percentile_afsc(c, instance):  # This sets afsc_ranking as a default value so it can be left blank
    """ """
    cadet_score_from_AFSC, cadet_scoring, cadet_ranking, afsc_scoring, afsc_ranking = build_ranking_data_structures(
        instance)
    afsc_that_wants_cadet_most = ['99N', 1.1]  # index 0 is AFSC and index 1 is percentile.


    # for interval in interval_list:
    for a in afsc_ranking:
        if c in afsc_ranking[a]:  # [interval]
            percentile = afsc_ranking[a].index(c) / len(afsc_ranking[a])
            if percentile < afsc_that_wants_cadet_most[1]:
                afsc_that_wants_cadet_most = [a, percentile]

    return afsc_that_wants_cadet_most[0]


def next_highest_percentile_afsc(c, instance):  # This sets afsc_ranking as a default value so it can be left blank
    """ Used to finish off matching cadets who couldn't match with the AFSC that ranked them the highest."""
    cadet_score_from_AFSC, cadet_scoring, cadet_ranking, afsc_scoring, afsc_ranking = build_ranking_data_structures(
        instance)
    percentile_dict = {}

    for a in afsc_ranking:
        if c in afsc_ranking[a]:  # [interval]
            percentile = afsc_ranking[a].index(c) / len(afsc_ranking[a])
            if percentile > 0:
                percentile_dict[a] = percentile

    return percentile_dict


def OLEA_Scoring_Build_func(instance):
    """ This is the function that builds the OLEA ranking structure for the AFSC's ranking of cadets."""

    # My Score Calculations for CFMs
    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

        AFSC_scores[a] = afsc_scored_cadets

        # Create a dictionary with AFSC and the scores from my CFM scoring matrix so we can compare each cadet's value to the values provided by OLEA. We can then turn the OLEA rankings into 0s where ineligible.

    afsc_dict = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_dict[a] = i

    # Created new dictionary of AFSCs and cadet scores using CFM scoring matrix using the OLEA 2023 data. We can now compare each cadet in this dictionary for the particular AFSC and assign a 0 score to all the cadets in the afsc_scoring sheet from OLEA based on who holds a 0 in this dictionary.
    list_of_AFSCs = afsc_dict.keys()
    list_of_AFSC_scores = list(AFSC_scores.values())
    AFSCs_scores = dict(zip(list_of_AFSCs, list_of_AFSC_scores))

    # Trying to create AFSC_scoring dict to use with OLEA's output to compare cadet positions, assign 0s, and remove ineligibles.

    ranking = {}
    # for a in range(instance.parameters['M']):
    for a in afsc_dict:
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSCs_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    afsc_scoring = {}
    for a in afsc_dict:
        afsc_scoring[a] = {}
        for i, v in enumerate(AFSCs_scores[a]):
            cadet = ranking[a][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSCs_scores[a][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # Read in the file from OLEA from the ranking's sheet.
    path = os.getcwd()
    data_file = 'Real 2023OLEA.xlsx'
    afsc_ranking_df = pd.read_excel(os.path.join(path, "afccp", "resources", "matching", "instances", data_file),
                                    sheet_name='AFSC Ranking', engine='openpyxl')
    # afsc_ranking_df = pd.read_excel(data_file, sheet_name='AFSC Ranking')
    afsc_ranking_df.set_index('Cadet', inplace=True)
    afsc_scoring_OLEA = afsc_ranking_df.to_dict()

    # Create dictionary of only eligible cadets for each AFSC from OLEA's roster.
    for a in afsc_dict:
        # afsc_scoring[a] = {}
        for c in afsc_scoring[a]:
            if c in afsc_scoring[a] and c in afsc_scoring_OLEA[a]:
                cadet_percentile_from_afsc = afsc_scoring_OLEA[a][c]
                afsc_scoring[a][c] = cadet_percentile_from_afsc

    sorted_percentiles = {}
    for a in afsc_dict:
        # print(a)
        cadets_in_order_list = []
        for k, v in sorted(afsc_scoring[a].items(), key=lambda item: item[1], reverse=True):
            # print('k:',k, 'v', v )
            cadets_in_order_list.append(k)
        sorted_percentiles[a] = cadets_in_order_list

    # Now we can take the index of each cadet
    for a in afsc_dict:
        for i, cadet in enumerate(sorted_percentiles[a], start=1):
            cadet_index = i
            c = cadet
            afsc_scoring[a][c] = cadet_index

    # Assign percentile as the cadet's score
    for a in afsc_dict:
        for c in afsc_scoring[a]:
            percentile = afsc_scoring[a][c] / len(afsc_scoring[a])
            afsc_scoring[a][c] = percentile

    # End OLEA Edits - previous code picks up below.
    # ----------------------------------------------------------------------------------------------------------
    # ----------------------------------------------------------------------------------------------------------

    # Original cadet scoring build items needed to finish OLEA build.
    cadet_scores = {}
    import numpy as np
    np.random.seed(2)
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])),
                                              reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    afsc_ranking = {} # Created as placeholder so I can switch between scoring systems

    OLEA_scoring_files = {
        "AFSC_scores": AFSC_scores,
        "ranking": ranking,
        "afsc_scoring": afsc_scoring,
        "cadet_scores": cadet_scores,
        "cadet_score_from_AFSC": cadet_score_from_AFSC,
        "cadet_scoring": cadet_scoring,
        "cadet_ranking": cadet_ranking,
        "afsc_ranking": afsc_ranking,
        "unmatched_ordered_cadets": unmatched_ordered_cadets}

    # print('OlEA_scoring keys:', OLEA_scoring_files.keys())
    # print('OlEA_scoring items:', OLEA_scoring_files["AFSC_scores"])

    return OLEA_scoring_files
    # ----------------------------------------------------------------------------------------------------------


def Tiered_Degree_Scoring_Build(instance):
    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

        AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit
    # are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the
    # "second matching".

    np.random.seed(2)
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]

    cadet_merit_list = [m for m in instance.parameters['merit']]

    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # If wanted sorted from best to worst. We do this later according to cadet rank.
    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    # AFSC_scores = AFSC_scoring_data_structure(instance)
    # cadet_scores, unmatched_ordered_cadets = cadet_scoring_data_structure(instance)

    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # i.e., the cade is not ineligible.
                cadet_scoring[c][a] = score

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this function to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    tiered_degree_scoring_files = {
        "AFSC_scores": AFSC_scores,
        "ranking": ranking,
        "afsc_scoring": afsc_scoring,
        "cadet_scores": cadet_scores,
        "cadet_score_from_AFSC": cadet_score_from_AFSC,
        "cadet_scoring": cadet_scoring,
        "cadet_ranking": cadet_ranking,
        "afsc_ranking":  afsc_ranking,
        "unmatched_ordered_cadets": unmatched_ordered_cadets}

    return tiered_degree_scoring_files


def HR_Classic_Algorithm(instance, capacity_type, scoring_files):
    """Classic HR algorithm made into a function

    scoring_files keys = AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking,\
    unmatched_ordered_cadets

    """

    print('Algorithm: Classic HR:', 'capacity_type:', capacity_type)

    # Using the two if statements so I can maintain the scoring_files name so that each graph prints accurately based on input.
    # scoring_files_dict calls the function to build the appropriate ranking structure for use in the rest of the algorithm.

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)


    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    unmatched_ordered_cadets_file = scoring_files_dict["unmatched_ordered_cadets"]



    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)

#     New Version
#     initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets_file  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    # ---------------------------------------------- #
    # Testing pre-process of max HR
    # New Version
    # initialize parameters


    cadets_f_32EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '32EXE':
                if cadet_scoring[cadet]['32EXE'] == 8.0:
                    cadets_f_32EXE[cadet] = cadet_scoring[cadet]['32EXE']


    ordered_cadets_f_32EXE = {k: v for k, v in sorted(cadets_f_32EXE.items(), key=lambda item: item[1], reverse=True)}
    print('32EXE cadets',ordered_cadets_f_32EXE)

    import itertools
    ordered_cadets_f_32EXE_3c = dict(
        itertools.islice(ordered_cadets_f_32EXE.items(), instance.parameters['quota_min'][14]))

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0 and cadet not in ordered_cadets_f_32EXE_3c:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                if cadet not in ordered_cadets_f_32EXE_3c:
                    # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                    cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_32EXE_3c = {}

    for c in ordered_cadets_f_32EXE:
        if c not in ordered_cadets_f_62EXC_20c and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_32EXE_3c[c] = cadet_scoring[c]['32EXE']

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_32EXE and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(),
                         instance.parameters['quota_min'][24]))  # taking min from 62EXC

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_32EXE_3c.keys() and c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    thirty_two_EXE_cadets = []
    for c in filtered_cadets_f_32EXE_3c.keys():
        thirty_two_EXE_cadets.append(c)
        afsc_matches['32EXE'] = thirty_two_EXE_cadets
        M[c] = '32EXE'

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    sixty_two_EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        sixty_two_EXE_cadets.append(c)
        afsc_matches['62EXE'] = sixty_two_EXE_cadets
        M[c] = '62EXE'

    # ---------------------------------------------- #

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print("Total iterations:", iter_num)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of cannot_match cadets:', len(cannot_match))

    # Find blocking pairs

    cadet_ranking = scoring_files_dict["cadet_ranking"]

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            if c in afsc_matches[a]:
                cadet_score_list.append(afsc_scoring[a][c])
            lowest_score = min(cadet_score_list)
            lowest_scoring_cadet_in_afsc[a] = lowest_score


    print("Working on finding blocking pairs...")
    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 OLEA Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])

    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Classic HR - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Classic HR - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Classic HR - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Classic HR - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Classic HR - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Classic HR - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")


    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})


    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")


    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()

    return afsc_matches, cannot_match, M, blocking_pairs_dictionary


def Pre_Processing_Algorithm(instance, capacity_type, scoring_files):
    """Pre-Processing algorithm to solve the 32EXE, 62EXC and 62EXE issue with lower quotas. Solves LQs when run at max
    capacity with 1574 cadets. New 1533 length cadet data means not enough cadets for 32EXE."""


    # AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking,\
    # unmatched_ordered_cadets = OLEA_Scoring_Build_func(instance)
    afsc_capacities = set_afsc_capacities(instance, capacity_type) #feeds capacity type to alg from input in func
    afsc_matches = afsc_matches_build(instance)
    print('Algorithm: Alg 2:', 'capacity_type:', capacity_type)

    # Using the two if statements so I can maintain the scoring_files name so that each graph prints accurately based on input.
    # scoring_files_dict calls the function to build the appropriate ranking structure for use in the rest of the algorithm.

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    start_time = time.time()

    # New Version
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    cadets_f_32EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '32EXE':
                if cadet_scoring[cadet]['32EXE'] == 8.0:
                    cadets_f_32EXE[cadet] = cadet_scoring[cadet]['32EXE']


    ordered_cadets_f_32EXE = {k: v for k, v in sorted(cadets_f_32EXE.items(), key=lambda item: item[1], reverse=True)}
    print('32EXE cadets',ordered_cadets_f_32EXE)

    import itertools
    ordered_cadets_f_32EXE_3c = dict(
        itertools.islice(ordered_cadets_f_32EXE.items(), instance.parameters['quota_min'][14]))

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0 and cadet not in ordered_cadets_f_32EXE_3c:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                if cadet not in ordered_cadets_f_32EXE_3c:
                    # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                    cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_32EXE_3c = {}

    for c in ordered_cadets_f_32EXE:
        if c not in ordered_cadets_f_62EXC_20c and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_32EXE_3c[c] = cadet_scoring[c]['32EXE']

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_32EXE and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(),
                         instance.parameters['quota_min'][24]))  # taking min from 62EXC

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_32EXE_3c.keys() and c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    thirty_two_EXE_cadets = []
    for c in filtered_cadets_f_32EXE_3c.keys():
        thirty_two_EXE_cadets.append(c)
        afsc_matches['32EXE'] = thirty_two_EXE_cadets
        M[c] = '32EXE'

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    sixty_two_EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        sixty_two_EXE_cadets.append(c)
        afsc_matches['62EXE'] = sixty_two_EXE_cadets
        M[c] = '62EXE'
# ----------------------------------------------------------------------------------------------------------------------
# Old working version below - excludes 32EXE
    # cadets_f_62EXC = {}
    # for cadet, afsc_w_scores in cadet_scoring.items():
    #     for key in afsc_w_scores:
    #         if key == '62EXC':
    #             if cadet_scoring[cadet]['62EXC'] >= 7.0:
    #                 cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']
    #
    # ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # # print(ordered_cadets_f_62EXC)
    #
    # import itertools
    # ordered_cadets_f_62EXC_20c = dict(
    #     itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))
    #
    # cadets_f_62EXE = {}
    # for cadet, afsc_w_scores in cadet_scoring.items():
    #     for key in afsc_w_scores:
    #         if key == '62EXE':
    #             # print(cadet,':', cadet_scoring[cadet]['62EXE'])
    #             cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']
    #
    # ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # # print(ordered_cadets_f_62EXE)
    # # print(len(ordered_cadets_f_62EXE))
    #
    # # Need to take only 53 because we have to leave 3 cadets for 32EXE
    # ordered_cadets_f_62EXE_53c = dict(
    #     itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))
    #
    # filtered_cadets_f_62EXC_20c = {}
    #
    # for c in ordered_cadets_f_62EXC:
    #     if c not in ordered_cadets_f_62EXE_53c:
    #         filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']
    #
    # # force it to just take the 20 cadets you want.
    # filtered_cadets_f_62EXC_20c = dict(
    #     itertools.islice(filtered_cadets_f_62EXC_20c.items(), instance.parameters['quota_min'][24])) #taking min from 62EXC
    #
    # # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    # pre_processed_unmatched_ordered_cadets = []
    # for c in unmatched_ordered_cadets:
    #     if c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
    #         pre_processed_unmatched_ordered_cadets.append(c)
    # unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # # print(len(unmatched_ordered_cadets))
    #
    # # Match cadets to AFSC 62EXC - update afsc_matches and M
    # EXC_cadets = []
    # for c in filtered_cadets_f_62EXC_20c.keys():
    #     EXC_cadets.append(c)
    #     afsc_matches['62EXC'] = EXC_cadets
    #     M[c] = '62EXC'
    #
    # # Match cadets to AFSC 62EXE - update afsc_matches and M
    # EXE_cadets = []
    # for c in ordered_cadets_f_62EXE_53c.keys():
    #     EXE_cadets.append(c)
    #     afsc_matches['62EXE'] = EXE_cadets
    #     M[c] = '62EXE'

# --------------------------------------------------------------------------------------------------------------------

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print("Total iterations:", iter_num)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of cannot_match cadets:', len(cannot_match))

    # Pull cadets from overclassified AFSCs and put them in AFSCs that are below quota.

    afscs = instance.parameters['afsc_vector']
    lower_quotas = instance.parameters['quota_min']
    afsc_lower_quotas = dict(zip(afscs, lower_quotas))

    moved_cadets = []
    immobile_cadets = []  # cadets that can't be moved
    a_h_q_afscs_w_no_cadets_to_donate = []
    iter_num = 0
    total_iterations = 0
    # while afsc lower quotas are not met or exceeded:
    distance_from_quota = {}
    lower_quota_ratios = [20]
    # next_cadet = False  # Set to look at the next cadet.
    # while distance_from_quota.values() < 100:
    # (v < 100 for v in distance_from_quota.values()) or iter_num <= iter_limit:

    iter_limit = 200
    while (lower_quota_ratios[
               0] < 99.0):  # iter_num < 1000 # Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
        # next_iteration = False # added
        next_cadet = False
        iter_num += 1

        # Select AFSC most under quota [a_l_q]
        distance_from_quota = {}
        for a in afsc_matches:
            if len(afsc_matches[a]) / afsc_lower_quotas[a] * 100 >= 100:  # if they are more than 100% of lower quota
                if (len(afsc_matches[a]) - 1) / afsc_lower_quotas[
                    a] * 100 < 100:  # If taking one cadet will put them below lower quota
                    a_h_q_afscs_w_no_cadets_to_donate.append(a)  # Then remove this AFSC from consideration
            # if all cadets in afsc_matches[a_h_q] contains all of the cadets in immobile cadets, continue.
            if a not in a_h_q_afscs_w_no_cadets_to_donate:
                distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[
                    a] * 100  # Calculate the percent from lower quota
        # lower_quota_ratios = []
        lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        # lower_quota_ratios = list(min(list(distance_from_quota.values())))
        # print(lower_quota_ratios, list(distance_from_quota.values()))
        a_l_q = [k for k, v in distance_from_quota.items() if v == min(
            distance_from_quota.values()) and v < 100]  # v<100 prevents continued filling after all afscs have met lower quota
        a_l_q = a_l_q[0]

        # Select AFSC most over quota [a_h_q]
        a_h_q = [k for k, v in distance_from_quota.items() if v == max(
            distance_from_quota.values()) > 100]  # v>100 ensures we only pull from AFSCs that are over their quota.
        # a_h_q = [k for k, v in distance_from_quota.items() if v == max(distance_from_quota.values()) and v > 100]
        if not a_h_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
            # print("I'm out of AFSCs over 100 to choose from: ", distance_from_quota)
            break  # Return to start to see if all AFSC lower quotas are >=100.
        a_h_q = a_h_q[0]  # Just take the first item in the list regardless

        # Select lowest scored cadet from most over-quota AFSC.
        a_h_cadet_score_from_afsc = []
        a_h_cadet_list = []
        lowest_score_cadet_for_afsc = {}

        # Building dict of all cadets in highest from quota
        for c in afsc_matches[a_h_q]:
            if c not in immobile_cadets:
                a_h_cadet_list.append(c)  # create a list of cadets matched to a_h_q
                a_h_cadet_score_from_afsc.append(afsc_scoring[a_h_q][c])  # create a list o
            # if next_cadet == True:
            #     distance_from_quota = distance_from_quota
            #     break
            # if next_cadet == False, we are still trying to match this cadet to the lowest quota AFSC. Keep trying.
        while len(a_h_cadet_score_from_afsc) > 0:
            lowest_score_cadet_for_afsc = dict(zip(a_h_cadet_list,
                                                   a_h_cadet_score_from_afsc))  # zip the two together to make a dictionary to pull the minimum value from.
            c = [k for k, v in lowest_score_cadet_for_afsc.items() if v == min(lowest_score_cadet_for_afsc.values())][0]
            if next_cadet == True:  # if next_iteration == True
                break

            while next_cadet == False:  # Keep trying to match cadet

                if c in afsc_scoring[
                    a_l_q]:  # If the cadet is scored by the lowest quota ratio AFSC (i.e. the cadet is eligible for the AFSC)
                    # Append this cadet to lowest quota AFSC.
                    afsc_matches[a_l_q].append(c)
                    # Remove this cadet from the highest quota AFSC.
                    afsc_matches[a_h_q].remove(c)
                    # Update M.
                    M[c] = a_l_q
                    # Track all cadets who have moved in a list
                    moved_cadets.append(c)
                    next_cadet = True  # We've finished getting this cadet match. Let's move on to the next cadet. Need to break out to the very first while loop.
                    # Could make next iteration
                    break  # first break to get back to the "for c in afsc_matches" loop. Then we'll break again to get to first while loop.

                # Select the next afsc the cadet is eligible for and remove the AFSC they were not eligible for.

                else:  # else: cadet is not in the lowest afsc (a_l_q). Check the next lowest AFSC.
                    # Select the next lowest quota AFSC and check for cadet membership in the above if statement.
                    if a_l_q in distance_from_quota:
                        del distance_from_quota[a_l_q]
                        a_l_q = [k for k, v in distance_from_quota.items() if
                                 v == min(distance_from_quota.values()) and v < 100]
                        if not a_l_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
                            next_cadet = True
                            del lowest_score_cadet_for_afsc[c]
                            immobile_cadets.append(c)
                            break
                        a_l_q = a_l_q[0]
                        next_cadet = False
                        continue

        if len(a_h_cadet_score_from_afsc) == 0:
            a_h_q_afscs_w_no_cadets_to_donate.append(a_h_q)
            print('No More Cadets in ', a_h_q)
            print(iter_num)
        # lower_quota_ratios = [ratio for afsc, ratio in distance_from_quota.items()]
        # lower_quota_ratios = []
        # lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        print(iter_num)
        continue

    total_iterations = iter_num
    print('Total Iterations:', iter_num)
    print(moved_cadets)
    for a in afsc_matches:
        distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[a] * 100
    print(distance_from_quota)
    print("Total time:", time.time() - start_time)
    print('Done')

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
        cadet_ranking[c] = afscs_from_cadets
    # Find blocking pairs

    # Initialize the keys in case some of the AFSCs are left unfilled (i.e. empty such as 32EXE with the OLEA data.)
    lowest_scoring_cadet_in_afsc = {}
    for a in afsc_matches:
        lowest_scoring_cadet_in_afsc[a] = []

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            if c in afsc_matches[a]:
                cadet_score_list.append(afsc_scoring[a][c])
            lowest_score = min(cadet_score_list)
            lowest_scoring_cadet_in_afsc[a] = lowest_score


    print("Working on finding blocking pairs...")
    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if not lowest_scoring_cadet_in_afsc[a]: # if lowest_scoring_cadet list empty, skip it.
                                    continue
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 OLEA Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])

    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 2 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 2 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 2 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 2 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 2 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 2 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")


    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()

    return unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary


def AFSC_Choose_After_Initial_Match(instance, capacity_type, scoring_files):
    """AFSC choosing their most desired cadet algorithm after the initial HR match. Testing the OLEA rank system."""

    # AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking, \
    # unmatched_ordered_cadets = OLEA_Scoring_Build_func(instance)
    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)
    print('Algorithm: Classic HR:', 'capacity_type:', capacity_type)

    # Using the two if statements so I can maintain the scoring_files name so that each graph prints accurately based on input.
    # scoring_files_dict calls the function to build the appropriate ranking structure for use in the rest of the algorithm.

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    # New Version
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches
    cannot_match_due_to_cap = []

    # Pre-process the problematic AFSCs according to the cadet's preference and the minimums required by each AFSC.

    # New Version
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    cadets_f_32EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '32EXE':
                if cadet_scoring[cadet]['32EXE'] == 8.0:
                    cadets_f_32EXE[cadet] = cadet_scoring[cadet]['32EXE']


    ordered_cadets_f_32EXE = {k: v for k, v in sorted(cadets_f_32EXE.items(), key=lambda item: item[1], reverse=True)}
    print('32EXE cadets',ordered_cadets_f_32EXE)

    import itertools
    ordered_cadets_f_32EXE_3c = dict(
        itertools.islice(ordered_cadets_f_32EXE.items(), instance.parameters['quota_min'][14]))

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0 and cadet not in ordered_cadets_f_32EXE_3c:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                if cadet not in ordered_cadets_f_32EXE_3c:
                    # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                    cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_32EXE_3c = {}

    for c in ordered_cadets_f_32EXE:
        if c not in ordered_cadets_f_62EXC_20c and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_32EXE_3c[c] = cadet_scoring[c]['32EXE']

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_32EXE and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(),
                         instance.parameters['quota_min'][24]))  # taking min from 62EXC

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_32EXE_3c.keys() and c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    thirty_two_EXE_cadets = []
    for c in filtered_cadets_f_32EXE_3c.keys():
        thirty_two_EXE_cadets.append(c)
        afsc_matches['32EXE'] = thirty_two_EXE_cadets
        M[c] = '32EXE'

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    sixty_two_EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        sixty_two_EXE_cadets.append(c)
        afsc_matches['62EXE'] = sixty_two_EXE_cadets
        M[c] = '62EXE'

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print("Total iterations:", iter_num)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of cannot_match cadets:', len(cannot_match))

    print('Working on the best cadet for each AFSC in cannot_match list...')
    best_cadet_start_time = time.time()

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    for c in cannot_match[:]:
        a = highest_percentile_afsc(c, instance)
        # Check if AFSC is at capacity. If not, add cadet.
        if len(afsc_matches[a]) < afsc_capacities[a]:
            afsc_matches[a].append(c)
            M[c] = a
            cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                a]]
            if len(cannot_match) != 0:
                c = cannot_match[0]
        # Check if AFSC is over capacity. If less than 120%, add cadet.
        elif len(afsc_matches[a]) >= afsc_capacities[a]:
            if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                afsc_capacities[a] = afsc_capacities[a] + 1
                afsc_matches[a].append(c)
                # Update M and update cannot_match list
                M[c] = a
                cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                    a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                if len(cannot_match) != 0:
                    c = cannot_match[0]
            else:
                cannot_match_due_to_cap.append(c)
                cannot_match[:] = [c for c in cannot_match if c not in cannot_match_due_to_cap]
                if len(cannot_match) != 0:
                    c = cannot_match[0]

    for c in cannot_match_due_to_cap[:]:
        percentile_dict = next_highest_percentile_afsc(c, instance)
        a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
        while c in cannot_match_due_to_cap:
            # delete minimum key : value pair from dict
            if len(afsc_matches[a]) < afsc_capacities[a]:
                M[c] = a
                cannot_match_due_to_cap[:] = [c for c in cannot_match_due_to_cap if c not in M]
                continue
            if len(afsc_matches[a]) >= afsc_capacities[a]:
                if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match_due_to_cap[:] = [c for c in cannot_match_due_to_cap if c not in M]
                else:
                    del percentile_dict[a]
                    a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
                    continue

    best_cadet_end_time = time.time()
    total_time_for_best_cadet = best_cadet_end_time - best_cadet_start_time
    print('Total time to find the best cadet for each AFSC in the cannot_match list:', total_time_for_best_cadet)

    # Finding Blocking Pairs

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            if c in afsc_matches[a]:
                cadet_score_list.append(afsc_scoring[a][c])
            lowest_score = min(cadet_score_list, default=cadet_score_list)
            lowest_scoring_cadet_in_afsc[a] = lowest_score


    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 OLEA Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])

    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # # append new solution to current solutions
    # solution_name = 'OLEA - AFSCs Choose After HR'
    # if solution_name in df_solutions.columns:
    #     solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    # append new solution to current solutions
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()

    return unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary


def AFSC_Choose_After_Initial_Match_Without_Cap(instance, capacity_type, scoring_files):
    """AFSC choosing their most desired cadet algorithm after the initial HR match. Testing the OLEA rank system."""

    # AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking, \
    # unmatched_ordered_cadets = OLEA_Scoring_Build_func(instance)
    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)
    print('Algorithm: Classic HR:', 'capacity_type:', capacity_type)

    # Using the two if statements so I can maintain the scoring_files name so that each graph prints accurately based on input.
    # scoring_files_dict calls the function to build the appropriate ranking structure for use in the rest of the algorithm.

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    # New Version
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches
    cannot_match_due_to_cap = []

    # Pre-process the problematic AFSCs according to the cadet's preference and the minimums required by each AFSC.

    # New Version
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    cadets_f_32EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '32EXE':
                if cadet_scoring[cadet]['32EXE'] == 8.0:
                    cadets_f_32EXE[cadet] = cadet_scoring[cadet]['32EXE']


    ordered_cadets_f_32EXE = {k: v for k, v in sorted(cadets_f_32EXE.items(), key=lambda item: item[1], reverse=True)}
    print('32EXE cadets',ordered_cadets_f_32EXE)

    import itertools
    ordered_cadets_f_32EXE_3c = dict(
        itertools.islice(ordered_cadets_f_32EXE.items(), instance.parameters['quota_min'][14]))

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0 and cadet not in ordered_cadets_f_32EXE_3c:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                if cadet not in ordered_cadets_f_32EXE_3c:
                    # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                    cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_32EXE_3c = {}

    for c in ordered_cadets_f_32EXE:
        if c not in ordered_cadets_f_62EXC_20c and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_32EXE_3c[c] = cadet_scoring[c]['32EXE']

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_32EXE and c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(),
                         instance.parameters['quota_min'][24]))  # taking min from 62EXC

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_32EXE_3c.keys() and c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    thirty_two_EXE_cadets = []
    for c in filtered_cadets_f_32EXE_3c.keys():
        thirty_two_EXE_cadets.append(c)
        afsc_matches['32EXE'] = thirty_two_EXE_cadets
        M[c] = '32EXE'

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    sixty_two_EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        sixty_two_EXE_cadets.append(c)
        afsc_matches['62EXE'] = sixty_two_EXE_cadets
        M[c] = '62EXE'

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print("Total iterations:", iter_num)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of cannot_match cadets:', len(cannot_match))

    print('Working on the best cadet for each AFSC in cannot_match list...')
    best_cadet_start_time = time.time()

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    for c in cannot_match[:]:
        a = highest_percentile_afsc(c, instance)
        # Check if AFSC is at capacity. If not, add cadet.
        if len(afsc_matches[a]) < afsc_capacities[a]:
            afsc_matches[a].append(c)
            M[c] = a
            cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                a]]
            if len(cannot_match) != 0:
                c = cannot_match[0]

    for c in cannot_match[:]:
        percentile_dict = next_highest_percentile_afsc(c, instance)
        a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
        while c in cannot_match:
            # delete minimum key : value pair from dict
            if len(afsc_matches[a]) < afsc_capacities[a]:
                M[c] = a
                cannot_match[:] = [c for c in cannot_match if c not in M]
                continue
            if len(afsc_matches[a]) >= afsc_capacities[a]:
                afsc_capacities[a] = afsc_capacities[a] + 1
                afsc_matches[a].append(c)
                M[c] = a
                cannot_match[:] = [c for c in cannot_match if c not in M]
            else:
                del percentile_dict[a]
                a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
                continue

    best_cadet_end_time = time.time()
    total_time_for_best_cadet = best_cadet_end_time - best_cadet_start_time
    print('Total time to find the best cadet for each AFSC in the cannot_match list:', total_time_for_best_cadet)

    # Finding Blocking Pairs

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            if c in afsc_matches[a]:
                cadet_score_list.append(afsc_scoring[a][c])
            lowest_score = min(cadet_score_list, default=cadet_score_list)
            lowest_scoring_cadet_in_afsc[a] = lowest_score


    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 OLEA Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])

    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # # append new solution to current solutions
    # solution_name = 'OLEA - AFSCs Choose After HR'
    # if solution_name in df_solutions.columns:
    #     solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    # append new solution to current solutions
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        solution_name = 'OLEA - Alg 3 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - Max Capacity'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - LQs'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        solution_name = 'Tiered Degree - Alg 3 - Uncapacitated'
        if solution_name in df_solutions.columns:
            solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()

    return unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary


def build_plotly_graphs(instance, build_graphs=True):

    if build_graphs == True:

        from collections import Counter
        import plotly.express as px

        AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking,\
        unmatched_ordered_cadets = OLEA_Scoring_Build_func(instance)
        afsc_matches = afsc_matches_build(instance)
        unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = HR_Classic_Algorithm(instance)
        unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = Pre_Processing_Algorithm(instance)
        unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = AFSC_Choose_After_Initial_Match(instance)

        cadets_matched_afsc_index = {}
        for a in afsc_matches:
            cadets_matched_afsc_index[a] = {}
            for c in afsc_matches[a]:
                if a in cadet_ranking[c]:
                    cadet_index = cadet_ranking[c].index(a)
                    cadets_matched_afsc_index[a][c] = cadet_index

        # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

        AFSC_index_counts_dict = {}

        for a in afsc_matches:
            AFSC_index_counts_dict[a] = {}
            AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
            AFSC_index_counts_dict[a] = AFSC_index_counts

        # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
        AFSC_index_counts_chart = {}
        for a in afsc_matches:
            AFSC_index_counts_chart[a] = {}
            for k, v in AFSC_index_counts_dict[a].items():
                k_str = "Choice " + str(k + 1)
                AFSC_index_counts_chart[a][k_str] = v

        plt.rcParams["figure.figsize"] = (10, 8)
        stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
        stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

        # Plotly graph below

        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart",
                     labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 31", ]})


        # -------------------------- NEXT FIGURE ------------------------------------ #



        # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

        # cadet_scoring = {}
        # for i in range(instance.parameters['N']):
        #     c = 'c' + str(i)
        #     cadet_scoring[c] = {}
        #     for j in range(instance.parameters['M']):
        #         a = instance.parameters['afsc_vector'][j]
        #         score = cadet_score_from_AFSC[i][j]
        #         if score > 0:  # i.e., the cade is not ineligible.
        #             cadet_scoring[c][a] = score
        #
        # for c in range(instance.parameters['N']):
        #     AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        #     # print(np.where(instance.parameters['utility'][c] > 0))
        #     for a in AFSC_list[0]:
        #         if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
        #             cadet_score_from_AFSC[c][a] = -1

        cadet_scores = {}
        for c in range(instance.parameters['N']):
            cadet_scored_afscs = []
            for a in range(instance.parameters['M']):
                if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                    cadet_scored_afscs.append(
                        5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
                elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                    cadet_scored_afscs.append(
                        4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
                elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                    cadet_scored_afscs.append(
                        3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
                elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                    cadet_scored_afscs.append(2 + np.random.random())
                elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                    cadet_scored_afscs.append(1 + np.random.random())
                elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                    cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
                else:
                    # cadet_scores.append(instance.parameters['merit'][c]*0)
                    cadet_scored_afscs.append(0)
                # If cadet score is below 0 then remove them from matching possibilities.

                cadet_scores[c] = cadet_scored_afscs

        # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
        cadet_score_from_AFSC = {}
        for c in range(instance.parameters['N']):
            cadet_of_AFSC_score_list = []
            for a in AFSC_scores.keys():
                s = cadet_scores[c][a]
                cadet_of_AFSC_score_list.append(s)
                cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

        # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

        for c in range(instance.parameters['N']):
            AFSC_list = np.where(instance.parameters['utility'][c] > 0)
            # print(np.where(instance.parameters['utility'][c] > 0))
            for a in AFSC_list[0]:
                if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                    cadet_score_from_AFSC[c][a] = -1

        cadet_scoring = {}
        for i in range(instance.parameters['N']):
            c = 'c' + str(i)
            cadet_scoring[c] = {}
            # print(c)
            for j in range(instance.parameters['M']):
                a = instance.parameters['afsc_vector'][j]
                score = cadet_score_from_AFSC[i][j]
                if score > 0:
                    cadet_scoring[c][a] = score

        print(cadet_scoring['c0'])
        # Need to eliminate the change for over 7 if preferenced.

        cadet_degree_matches = {}
        for a in afsc_matches:
            cadet_degree_matches[a] = {}
            for c in afsc_matches[a]:
                if a in cadet_scoring[c]:
                    if cadet_scoring[c][a] >= 7:
                        cadet_degree_matches[a][c] = 'Voluntary Match'
                    if 5 <= cadet_scoring[c][a] < 6:
                        cadet_degree_matches[a][c] = 'Mandatory Vol'
                    if 4 <= cadet_scoring[c][a] < 5:
                        cadet_degree_matches[a][c] = 'Desired Vol'
                    if 3 <= cadet_scoring[c][a] < 4:
                        cadet_degree_matches[a][c] = 'Permitted Vol'
                    if 2 <= cadet_scoring[c][a] < 3:
                        cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                    if 1 <= cadet_scoring[c][a] < 2:
                        cadet_degree_matches[a][c] = 'Desired Non-Vol'
                    if 0.5 <= cadet_scoring[c][a] < 1:
                        cadet_degree_matches[a][c] = 'Permitted Non-Vol'

        cadet_degree_counts_dict = {}

        for a in afsc_matches:
            cadet_degree_counts_dict[a] = {}
            cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
            cadet_degree_counts_dict[a] = cadet_degree_counts

        cadet_degree_counts_chart = {}
        for a in afsc_matches:
            cadet_degree_counts_chart[a] = {}
            for k, v in cadet_degree_counts_dict[a].items():
                cadet_degree_counts_chart[a][k] = v

        cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
        cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T

        # Plotly graph below

        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier with AFSC Match",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders= {"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                           'Mandatory Non-Vol', 'Desired Non-Vol','Permitted Non-Vol']})



        # Dict 1
        cadet_choice_dict = {}
        for a in afsc_matches:
            for c in afsc_matches[a]:
                if a in cadet_ranking[c]:
                    cadet_index = cadet_ranking[c].index(a)
                    cadet_index += 1
                    cadet_choice_dict[c] = cadet_index

        # Dict 2
        cadet_merit_dict = {}
        for a in afsc_matches:
            for c in afsc_matches[a]:
                if a in cadet_ranking[c]:
                    c_merit = int(c[1:])
                    cadet_merit = instance.parameters['merit'][c_merit]
                    cadet_merit_dict[c] = cadet_merit

        # Dict 3
        cadet_to_matched_afsc_dict = {}
        for a in afsc_matches:
            for c in afsc_matches[a]:
                if a in cadet_ranking[c]:
                    cadet_to_matched_afsc_dict[c] = a

        merit_vs_choice_chart_df = pd.DataFrame(
            {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
             'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

        merit_vs_choice_chart_df.head()

        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC")


        # -------------------------- NEXT FIGURE ------------------------------------ #

        box_merit_plot_df = pd.DataFrame(
            {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC")

        # -------------------------- NEXT FIGURE ------------------------------------ #

        # Make merit plot
        import matplotlib.font_manager as font_manager
        from matplotlib import rcParams

        afscs = instance.parameters['afsc_vector']
        afsc_lower_quotas = instance.parameters["quota_min"]
        unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

        afsc_avg_merit = {}

        rcParams['axes.spines.top'] = False
        rcParams['axes.spines.right'] = False

        font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
        for font in font_manager.findSystemFonts(font_dir):
            font_manager.fontManager.addfont(font)

        rcParams['font.family'] = 'Open Sans'

        for a in afsc_matches:
            if unchanging_lower_quotas[a] > 40:
                list_of_cadet_merit_in_matched_afsc = []
                for c in afsc_matches[a]:
                    c = int(c[1:])
                    list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
            else:
                continue
            afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
            afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
            afsc_avg_merit[a] = afsc_avg_cadet_merit

        print('Average Merit per AFSC:', afsc_avg_merit)
        # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
        # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
        afsc_merit_list = [m for m in afsc_avg_merit.values()]
        mean_merit = np.mean(afsc_merit_list)
        print("Mean Merit of Solution for Large AFSCs:", mean_merit)
        median_merit = np.median(afsc_merit_list)
        print('Median Merit of Solution for Large AFSCs:', median_merit)

        fig = plt.figure(figsize=(15, 9))
        plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
        plt.axhline(y=0.35, color='black', linestyle='--')
        plt.axhline(y=0.65, color='black', linestyle='--')
        plt.xlabel('Large AFSC', size=14)
        plt.ylabel('Average Merit of AFSC', size=14)
        plt.title('Average Merit of Large AFSCs', size=16)
        plt.ylim(.2, 0.8)
        merit_plot = plt.show()


    return stacked_fig, fig_degree, fig_choice_merit, box_fig, merit_plot


def algorithm_1_for_thesis(instance, capacity_type, scoring_files):
    global fig_degree, fig_choice_merit, box_fig, stacked_fig
    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)


    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)



    print('Running HR algorithm...')
    afsc_matches, cannot_match, M, blocking_pairs_dictionary = HR_Classic_Algorithm(instance, capacity_type, scoring_files)
    print('HR algorithm complete')


    # Build the graphs needed

    # stacked_fig, fig_degree, fig_choice_merit, box_fig, merit_plot = build_plotly_graphs(instance, build_graphs=True)
    #
    # stacked_fig.show()
    # fig_degree.show()
    # fig_choice_merit.show()
    # box_fig.show()

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    from collections import Counter
    import plotly.express as px
    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index

    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts

    # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

    # Plotly graph below

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Max - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - LQ - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Uncapacitated - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - Max - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 31", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - LQ - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - Uncapacitated - Classic HR",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    stacked_fig.show()
    # -------------------------- NEXT FIGURE ------------------------------------ #

    sum_of_cadet_choice_chart = stacked_bar_chart_dict_df.sum().to_frame()
    sum_of_cadet_choice_chart.reset_index(inplace=True)

    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.rename(columns={'index': 'Choice', 0: 'Sum of Cadet Choice'})

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['Sum of Cadet Choice'].div(
        (instance.parameters['N']), axis=0)

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['% of Total Cadets'] * 100

    sum_of_cadet_choice_chart_mapping = pd.DataFrame(
        {'Choice': ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                    "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                    "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                    "Choice 18",
                    "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                    "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                    "Choice 29", "Choice 30", "Choice 31", "Choice 32"]})

    sort_mapping = sum_of_cadet_choice_chart_mapping.reset_index().set_index('Choice')
    sum_of_cadet_choice_chart['Choice_order'] = sum_of_cadet_choice_chart['Choice'].map(sort_mapping['index'])
    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.sort_values('Choice_order')

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Max - Classic HR', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - LQ - Classic HR', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Uncapacitated - Classic HR', text_auto=True)

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Max - Classic HR', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - LQ - Classic HR', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Uncapacitated - Classic HR', text_auto=True)

    cadet_choice_sums_fig.show()



    # -------------------------- NEXT FIGURE ------------------------------------ #

    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        if instance.parameters['utility'][c].any() == -50:
            print(True)
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1


    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score

    # Need to eliminate the change for over 7 if preferenced.

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] <= 7:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'


    # Force search and pull the cadets who didn't give a utility rating. They count as voluntary for needs of the AF.

    cadet_utilities = {}
    for c in range(instance.parameters['N']):
        cadet = 'c' + str(c)
        cadet_utilities[cadet] = []
        utility_list = []
        for a in range(instance.parameters['M']):
            utility = instance.parameters['utility'][c][a]
            utility = float(utility)
            utility_list.append(utility)
        cadet_utilities[cadet] = utility_list

    cadets_with_no_utility = []
    for c in cadet_ranking.keys():
        if all(value == 0 for value in cadet_utilities[c]):
            cadets_with_no_utility.append(c)

    for a in afsc_matches:
        for c in afsc_matches[a]:
            if c in cadets_with_no_utility:
                cadet_degree_matches[a][c] = 'Permitted Vol'

# Added above code blocks to reflect all cadets with no utility as "voluntary" matches based on needs of the AF


    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - OLEA - Max - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - OLEA - LQ - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - OLEA - Uncapacitated - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - Tiered Degree - Max - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - Tiered Degree - LQ - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier Match - Tiered Degree - Uncapacitated - Classic HR",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    fig_degree.show()

    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - Max - Classic HR')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - LQ - Classic HR')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - Uncapacitated - Classic HR')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - Max - Classic HR')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - LQ - Classic HR')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - Uncapacitated - Classic HR')


    fig_choice_merit.show()


    # -------------------------- NEXT FIGURE ------------------------------------ #

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Max - Classic HR')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - LQ - Classic HR')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Uncapacitated - Classic HR')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Max - Classic HR')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - LQ - Classic HR')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Uncapacitated - Classic HR')


    box_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    # Make merit plot
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

    afsc_avg_merit = {}

    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Max - Classic HR', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - LQ - Classic HR', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Uncapacitated - Classic HR', size=16)
    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Max - Classic HR', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - LQ - Classic HR', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Uncapacitated - Classic HR', size=16)
    plt.ylim(.2, 0.8)
    merit_plot = plt.show()


def algorithm_2_for_thesis(instance, capacity_type, scoring_files):
    global fig_degree, fig_choice_merit, box_fig, stacked_fig

    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))


    print('Running Pre-Processing algorithm...')
    unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = Pre_Processing_Algorithm(instance, capacity_type, scoring_files)
    print('Pre-Processing algorithm complete')

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)

    # Build the graphs needed
    #
    from collections import Counter
    import plotly.express as px
    # stacked_fig, fig_degree, fig_choice_merit, box_fig, merit_plot = build_plotly_graphs(instance, build_graphs=True)
    #
    # stacked_fig.show()
    # fig_degree.show()
    # fig_choice_merit.show()
    # box_fig.show()

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    cadet_ranking = scoring_files_dict["cadet_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]


    # cadet_scoring = {}
    # for i in range(instance.parameters['N']):
    #     # print(i)
    #     #     #print(cadet_score_from_AFSC[c])
    #     c = 'c' + str(i)
    #     cadet_scoring[c] = {}
    #     # print(c)
    #     for j in range(instance.parameters['M']):
    #         a = instance.parameters['afsc_vector'][j]
    #         # print(a)
    #         score = cadet_score_from_AFSC[i][j]
    #         if score > 0:  # added
    #             cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
        cadet_ranking[c] = afscs_from_cadets

    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index

    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts

    # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

    # Plotly graph below

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Max - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - LQ - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Uncapacitated - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - Max - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - LQ - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df,
                             title="Cadet Choice Chart - Tiered Degree - Uncapacitated - Alg 2",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    stacked_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    sum_of_cadet_choice_chart = stacked_bar_chart_dict_df.sum().to_frame()
    sum_of_cadet_choice_chart.reset_index(inplace=True)

    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.rename(columns={'index': 'Choice', 0: 'Sum of Cadet Choice'})

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['Sum of Cadet Choice'].div(
        (instance.parameters['N']), axis=0)

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['% of Total Cadets'] * 100

    sum_of_cadet_choice_chart_mapping = pd.DataFrame(
        {'Choice': ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                    "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                    "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                    "Choice 18",
                    "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                    "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                    "Choice 29", "Choice 30", "Choice 31", "Choice 32"]})

    sort_mapping = sum_of_cadet_choice_chart_mapping.reset_index().set_index('Choice')
    sum_of_cadet_choice_chart['Choice_order'] = sum_of_cadet_choice_chart['Choice'].map(sort_mapping['index'])
    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.sort_values('Choice_order')

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Max - Alg 2', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - LQ - Alg 2', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Uncapacitated - Alg 2', text_auto=True)

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Max - Alg 2', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - LQ - Alg 2', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Uncapacitated - Alg 2', text_auto=True)

    cadet_choice_sums_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #



    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score

    # Need to eliminate the change for over 7 if preferenced.

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] <= 7:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'

    # Force search and pull the cadets who didn't give a utility rating. They count as voluntary for needs of the AF.

    cadet_utilities = {}
    for c in range(instance.parameters['N']):
        cadet = 'c' + str(c)
        cadet_utilities[cadet] = []
        utility_list = []
        for a in range(instance.parameters['M']):
            utility = instance.parameters['utility'][c][a]
            utility = float(utility)
            utility_list.append(utility)
        cadet_utilities[cadet] = utility_list

    cadets_with_no_utility = []
    for c in cadet_ranking.keys():
        if all(value == 0 for value in cadet_utilities[c]):
            cadets_with_no_utility.append(c)

    for a in afsc_matches:
        for c in afsc_matches[a]:
            if c in cadets_with_no_utility:
                cadet_degree_matches[a][c] = 'Permitted Vol'

    # Added above code blocks to reflect all cadets with no utility as "voluntary" matches based on needs of the AF

    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Max - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - LQ - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Uncapacitated - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Max - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - LQ - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Uncapacitated - Alg 2",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    fig_degree.show()

    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - Max - Alg 2')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - LQ - Alg 2')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - OLEA - Uncapacitated - Alg 2')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - Max - Alg 2')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - LQ - Alg 2')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC", title= 'Cadet Merit vs Choice - Tiered Degree - Uncapacitated - Alg 2')



    fig_choice_merit.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Max - Alg 2')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - LQ - Alg 2')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Uncapacitated - Alg 2')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Max - Alg 2')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - LQ - Alg 2')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Uncapacitated - Alg 2')


    box_fig.show()
    # -------------------------- NEXT FIGURE ------------------------------------ #

    # Make merit plot
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

    afsc_avg_merit = {}

    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Max - Alg 2', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - LQ - Alg 2', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Uncapacitated - Alg 2', size=16)
    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Max - Alg 2', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - LQ - Alg 2', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Uncapacitated - Alg 2', size=16)
    plt.ylim(.2, 0.8)
    merit_plot = plt.show()


def algorithm_3_for_thesis(instance, capacity_type, scoring_files):
    global fig_degree, fig_choice_merit, box_fig, stacked_fig

    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    print('Running AFSC Choosing on Second Half algorithm...')
    unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = AFSC_Choose_After_Initial_Match(instance, capacity_type, scoring_files)
    print('AFSC Choosing on Second Half algorithm complete')

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)
    # Build the graphs needed
    #

    from collections import Counter
    import plotly.express as px
    # stacked_fig, fig_degree, fig_choice_merit, box_fig, merit_plot = build_plotly_graphs(instance, build_graphs=True)
    #
    # stacked_fig.show()
    # fig_degree.show()
    # fig_choice_merit.show()
    # box_fig.show()

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
        cadet_ranking[c] = afscs_from_cadets

    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index

    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts

    # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

    # Plotly graph below

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Max - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - LQ - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Uncapacitated - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - Max - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - LQ - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df,
                             title="Cadet Choice Chart - Tiered Degree - Uncapacitated - Alg 3",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    stacked_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    sum_of_cadet_choice_chart = stacked_bar_chart_dict_df.sum().to_frame()
    sum_of_cadet_choice_chart.reset_index(inplace=True)

    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.rename(columns={'index': 'Choice', 0: 'Sum of Cadet Choice'})

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['Sum of Cadet Choice'].div(
        (instance.parameters['N']), axis=0)

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['% of Total Cadets'] * 100

    sum_of_cadet_choice_chart_mapping = pd.DataFrame(
        {'Choice': ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                    "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                    "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                    "Choice 18",
                    "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                    "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                    "Choice 29", "Choice 30", "Choice 31", "Choice 32"]})

    sort_mapping = sum_of_cadet_choice_chart_mapping.reset_index().set_index('Choice')
    sum_of_cadet_choice_chart['Choice_order'] = sum_of_cadet_choice_chart['Choice'].map(sort_mapping['index'])
    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.sort_values('Choice_order')

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Max - Alg 3', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - LQ - Alg 3', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Uncapacitated - Alg 3', text_auto=True)

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Max - Alg 3', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - LQ - Alg 3', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Uncapacitated - Alg 3', text_auto=True)

    cadet_choice_sums_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score


    # Need to eliminate the change for over 7 if preferenced.

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] <= 7:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'

    # Force search and pull the cadets who didn't give a utility rating. They count as voluntary for needs of the AF.

    cadet_utilities = {}
    for c in range(instance.parameters['N']):
        cadet = 'c' + str(c)
        cadet_utilities[cadet] = []
        utility_list = []
        for a in range(instance.parameters['M']):
            utility = instance.parameters['utility'][c][a]
            utility = float(utility)
            utility_list.append(utility)
        cadet_utilities[cadet] = utility_list

    cadets_with_no_utility = []
    for c in cadet_ranking.keys():
        if all(value == 0 for value in cadet_utilities[c]):
            cadets_with_no_utility.append(c)

    for a in afsc_matches:
        for c in afsc_matches[a]:
            if c in cadets_with_no_utility:
                cadet_degree_matches[a][c] = 'Permitted Vol'

# Added above code blocks to reflect all cadets with no utility as "voluntary" matches based on needs of the AF

    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T


    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Max - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - LQ - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Uncapacitated - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Max - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - LQ - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Uncapacitated - Alg 3",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    fig_degree.show()


    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - Max - Alg 3')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - LQ - Alg 3')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - Uncapacitated - Alg 3')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - Max - Alg 3')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - LQ - Alg 3')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - Uncapacitated - Alg 3')

    fig_choice_merit.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Max - Alg 3')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - LQ - Alg 3')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Uncapacitated - Alg 3')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Max - Alg 3')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - LQ - Alg 3')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Uncapacitated - Alg 3')

    box_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    # Make merit plot
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

    afsc_avg_merit = {}

    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Max - Alg 3', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - LQ - Alg 3', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Uncapacitated - Alg 3', size=16)
    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Max - Alg 3', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - LQ - Alg 3', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Uncapacitated - Alg 3', size=16)
    plt.ylim(.2, 0.8)
    merit_plot = plt.show()


def algorithm_4_for_thesis(instance, capacity_type, scoring_files):
    global fig_degree, fig_choice_merit, box_fig, stacked_fig

    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    print('Running AFSC Choosing on Second Half algorithm...')
    unmatched_ordered_cadets, afsc_matches, cannot_match, M, blocking_pairs_dictionary = AFSC_Choose_After_Initial_Match_Without_Cap(instance, capacity_type, scoring_files)
    print('AFSC Choosing on Second Half algorithm complete')

    if scoring_files == 'OLEA_scoring_files':
        scoring_files_dict = OLEA_Scoring_Build_func(instance)

    if scoring_files == 'tiered_degree_scoring_files':
        scoring_files_dict = Tiered_Degree_Scoring_Build(instance)
    # Build the graphs needed
    #

    from collections import Counter
    import plotly.express as px
    # stacked_fig, fig_degree, fig_choice_merit, box_fig, merit_plot = build_plotly_graphs(instance, build_graphs=True)
    #
    # stacked_fig.show()
    # fig_degree.show()
    # fig_choice_merit.show()
    # box_fig.show()

    AFSC_scores = scoring_files_dict["AFSC_scores"]
    ranking = scoring_files_dict["ranking"]
    afsc_scoring = scoring_files_dict["afsc_scoring"]
    cadet_scores = scoring_files_dict["cadet_scores"]
    cadet_score_from_AFSC = scoring_files_dict["cadet_score_from_AFSC"]
    cadet_scoring = scoring_files_dict["cadet_scoring"]
    afsc_ranking = scoring_files_dict["afsc_ranking"]
    unmatched_ordered_cadets = scoring_files_dict["unmatched_ordered_cadets"]

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
        cadet_ranking[c] = afscs_from_cadets

    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index

    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts

    # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

    # Plotly graph below

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Max - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - LQ - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - OLEA - Uncapacitated - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - Max - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart - Tiered Degree - LQ - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        stacked_fig = px.bar(stacked_bar_chart_dict_df,
                             title="Cadet Choice Chart - Tiered Degree - Uncapacitated - Alg 4",
                             labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},

                             color_discrete_sequence=['#0508b8', '#1910d8', '#3c19f0', '#6b1cfb', '#981cfd', '#bf1cfd',
                                                      '#dd2bfd', '#f246fe', '#fc67fd', '#fe88fc', '#fea5fd', '#febefe',
                                                      '#fec3fe',

                                                      'rgb(253, 237, 176)', 'rgb(250, 205, 145)', 'rgb(246, 173, 119)',
                                                      'rgb(240, 142, 98)', 'rgb(231, 109, 84)', 'rgb(216, 80, 83)',
                                                      'rgb(195, 56, 90)', 'rgb(168, 40, 96)', 'rgb(138, 29, 99)',
                                                      'rgb(107, 24, 93)', 'rgb(76, 21, 80)', 'rgb(47, 15, 61)',

                                                      'rgb(0, 0, 0)', 'rgb(16, 16, 16)', 'rgb(38, 38, 38)',
                                                      'rgb(59, 59, 59)', 'rgb(81, 80, 80)', 'rgb(102, 101, 101)',
                                                      'rgb(124, 123, 122)', 'rgb(146, 146, 145)', 'rgb(171, 171, 170)',
                                                      'rgb(197, 197, 195)', 'rgb(224, 224, 223)', 'rgb(254, 254, 253)'
                                                      ],
                             category_orders={
                                 "variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                              "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                                              "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                              "Choice 18",
                                              "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                              "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                              "Choice 29", "Choice 30", "Choice 31", "Choice 32", ]})

    stacked_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    sum_of_cadet_choice_chart = stacked_bar_chart_dict_df.sum().to_frame()
    sum_of_cadet_choice_chart.reset_index(inplace=True)

    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.rename(columns={'index': 'Choice', 0: 'Sum of Cadet Choice'})

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['Sum of Cadet Choice'].div(
        (instance.parameters['N']), axis=0)

    sum_of_cadet_choice_chart['% of Total Cadets'] = sum_of_cadet_choice_chart['% of Total Cadets'] * 100

    sum_of_cadet_choice_chart_mapping = pd.DataFrame(
        {'Choice': ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                    "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11", "Choice 12",
                    "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                    "Choice 18",
                    "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                    "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                    "Choice 29", "Choice 30", "Choice 31", "Choice 32"]})

    sort_mapping = sum_of_cadet_choice_chart_mapping.reset_index().set_index('Choice')
    sum_of_cadet_choice_chart['Choice_order'] = sum_of_cadet_choice_chart['Choice'].map(sort_mapping['index'])
    sum_of_cadet_choice_chart = sum_of_cadet_choice_chart.sort_values('Choice_order')

    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Max - Alg 4', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - LQ - Alg 4', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - OLEA - Uncapacitated - Alg 4', text_auto=True)

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Max - Alg 4', text_auto=True)

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - LQ - Alg 4', text_auto=True)

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        cadet_choice_sums_fig = px.bar(sum_of_cadet_choice_chart, x='Choice', y='Sum of Cadet Choice',
                                       color='% of Total Cadets', color_continuous_scale='Agsunset',
                                       title='Sum of Cadet Choice - Tiered Degree - Uncapacitated - Alg 4', text_auto=True)

    cadet_choice_sums_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(
                    3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score


    # Need to eliminate the change for over 7 if preferenced.

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] <= 7:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'

    # Force search and pull the cadets who didn't give a utility rating. They count as voluntary for needs of the AF.

    cadet_utilities = {}
    for c in range(instance.parameters['N']):
        cadet = 'c' + str(c)
        cadet_utilities[cadet] = []
        utility_list = []
        for a in range(instance.parameters['M']):
            utility = instance.parameters['utility'][c][a]
            utility = float(utility)
            utility_list.append(utility)
        cadet_utilities[cadet] = utility_list

    cadets_with_no_utility = []
    for c in cadet_ranking.keys():
        if all(value == 0 for value in cadet_utilities[c]):
            cadets_with_no_utility.append(c)

    for a in afsc_matches:
        for c in afsc_matches[a]:
            if c in cadets_with_no_utility:
                cadet_degree_matches[a][c] = 'Permitted Vol'

# Added above code blocks to reflect all cadets with no utility as "voluntary" matches based on needs of the AF

    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T


    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Max - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - LQ - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - OLEA - Uncapacitated - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Max - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - LQ - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df,
                            title="Cadet Degree Tier Match - Tiered Degree - Uncapacitated - Alg 4",
                            labels={"value": "Number of Cadets", "index": "AFSC",
                                    "variable": "Cadet Degree Tier and Vol Status"},

                            color_discrete_map={'Mandatory Vol': 'rgb(0,0,255)', 'Desired Vol': 'rgb(51,153,255)',
                                                'Permitted Vol': 'rgb(102,204,255)',
                                                'Mandatory Non-Vol': 'rgb(255,102,204)',
                                                'Desired Non-Vol': 'rgb(255,102,102)',
                                                'Permitted Non-Vol': 'rgb(255,0,0)'},

                            category_orders={"variable": ['Mandatory Vol', 'Desired Vol', 'Permitted Vol',
                                                          'Mandatory Non-Vol', 'Desired Non-Vol', 'Permitted Non-Vol']})

    fig_degree.show()


    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - Max - Alg 4')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - LQ - Alg 4')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - OLEA - Uncapacitated - Alg 4')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - Max - Alg 4')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - LQ - Alg 4')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC",
                                      title='Cadet Merit vs Choice - Tiered Degree - Uncapacitated - Alg 4')

    fig_choice_merit.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    # Plotly graph below
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Max - Alg 4')

    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - LQ - Alg 4')

    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - OLEA - Uncapacitated - Alg 4')

    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Max - Alg 4')

    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - LQ - Alg 4')

    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC", title='Cadet Merit Distribution - Tiered Degree - Uncapacitated - Alg 4')

    box_fig.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    # Make merit plot
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

    afsc_avg_merit = {}

    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    if capacity_type == 'max' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Max - Alg 4', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - LQ - Alg 4', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'OLEA_scoring_files':
        plt.title('Average Merit of Large AFSCs - OLEA - Uncapacitated - Alg 4', size=16)
    if capacity_type == 'max' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Max - Alg 4', size=16)
    if capacity_type == 'lower_quotas' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - LQ - Alg 4', size=16)
    if capacity_type == 'uncapacitated' and scoring_files == 'tiered_degree_scoring_files':
        plt.title('Average Merit of Large AFSCs - Tiered Degree - Uncapacitated - Alg 4', size=16)
    plt.ylim(.2, 0.8)
    merit_plot = plt.show()


def alg_function_test(instance, capacity_type):
    AFSC_scores_list = AFSC_scoring_data_structure(instance)
    cadet_scores_list, unmatched_ordered_cadets = cadet_scoring_data_structure(instance)
    cadet_score_from_AFSC, cadet_scoring, cadet_ranking, afsc_scoring, afsc_ranking = build_ranking_data_structures(
        instance)
    afsc_capacities = set_afsc_capacities(instance, capacity_type)
    afsc_matches = afsc_matches_build(instance)

    start_time = time.time()
    Classical_HR_algorithm_time = time.time()

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print("Total iterations:", iter_num)

    print("Time to run matching algorithm:", Classical_HR_algorithm_time - start_time)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    # print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    return M, cannot_match, afsc_matches


def alg_AFSC_choose_best_cadet_to_match_all(instance, capacity_type='lower_quotas'):
    # AFSC_scores_list = AFSC_scoring_data_structure(instance)
    # cadet_scores_list, unmatched_ordered_cadets_list = cadet_scoring_data_structure(instance)
    # cadet_score_from_AFSC, cadet_scoring, cadet_ranking, afsc_scoring, afsc_ranking = build_ranking_data_structures(
    #     instance)
    AFSC_scores, ranking, afsc_scoring, cadet_scores, cadet_score_from_AFSC, cadet_scoring, cadet_ranking, unmatched_ordered_cadets = OLEA_Scoring_Build_func(instance)
    afsc_capacities = set_afsc_capacities(instance)
    afsc_matches = afsc_matches_build(instance)

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    matching_data_structure_creation = time.time()

    start_time = time.time()
    Classical_HR_algorithm_time = time.time()
    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches
    blocking_pairs = []
    cannot_match_due_to_cap = []


    # Adding the new pre-processing elements

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(), instance.parameters['quota_min'][24]))

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        EXE_cadets.append(c)
        afsc_matches['62EXE'] = EXE_cadets
        M[c] = '62EXE'

    ###### END pre-processing code

# # Old Version
#     # initialize parameters
    print_details = False
    use_test_cadets = False
    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    # print(unmatched_ordered_cadets)
    test_cadets = ['c144', 'c82', 'c951', 'c1182', 'c460', 'c1183']
    if use_test_cadets:
        unmatched_ordered_cadets = test_cadets
    while len(unmatched_ordered_cadets) > 0 and iter_num <= iter_limit:  # while there are still cadets to match
        if print_details: print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
        if print_details: print('Length of perm unmatched cadets:', len(cannot_match))
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            if print_details: print('\nnext cadet', c)  # Can comment out
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
            # print(unmatched_ordered_cadets) # Can comment out
        if len(cadet_ranking[c]) == 0:  # if no more AFSCs in ranking list
            if print_details: print('Cannot match', c)
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
        print('Trying to match', c, a)
        if c == 'c1464':
            print('\n\n\nXXXXX ', c,
                  'XXXXXXXX-----------------------------------------------------------------------------------------------------------------------------\n\n\n')
        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            # if print_details: print("M", M)
            if print_details: print('Matched', c, 'and', a)
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            # afsc_matches = {13H: ["c1", "c2", "c3",...]}
            # sort afsc_matches[a] list based on afsc_ranking list of cadets
            # afsc_scoring[a][c] = s
            # print(afsc_ranking[a])
            #      afsc_matches[a] = [afsc_ranking[b] for _, b in zip(afsc_matches[a], afsc_ranking[a])]
            # TODO: sort matches by cadet scores, descending
            # print('matches 1:', afsc_matches) # Can comment out
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop
        else:  # if at capacity
            if print_details: print(a, "at capacity")
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            if print_details: print("AFSC_matches", afsc_matches[a])
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                if print_details: print(c_, 'higher than', c)
                # if print_details: print(cadet_ranking[c])
                if print_details: print("Line 61: Removing", a, 'from', c, 's list')
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                # if print_details: print(a, 'removed from', c,'s list')
                # M[c_] = a #added this define M[c_] so I could remove from it but c_ is already matched so this may be a problem.
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop
            else:  # if c ranks higher than c_
                if print_details: print(c_, 'lower than', c)
                if print_details: print('Line 69: Removing', c_, 'from', a)
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                if c_ == 'c144': print(cadet_ranking[c_],
                                       '-----------------******************--------------------------------')
                if print_details: print("Removing", a, 'from', c_, 's list')
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                # if print_details: print("Cadet AFSC List:", cadet_ranking[c_])
                M.pop(c_, print(c_))  # remove c_ from M - Gives error that can't use pop to remove a string.
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                if len(unmatched_ordered_cadets) == 0 and next_cadet == False:  # added to get the last cadet added as unmatched
                    cannot_match.append(c)  # added to get the last cadet added as unmatched
                continue  # go to beginning of while loop

#-----------------------------------------------#

# New version as of 21 Oct 22
    # initialize parameters
    #
    # iter_num = 0
    # iter_limit = 10000
    # next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    # while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
    #     iter_num += 1
    #     if next_cadet:
    #         c = unmatched_ordered_cadets[0]  # next cadet to match
    #         unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
    #
    #     if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
    #         cannot_match.append(c)  # add them to cannot match list
    #         next_cadet = True
    #         continue  # go to beginning of while loop
    #     a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
    #
    #     if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
    #         M[c] = a  # add match to M
    #         afsc_matches[a].append(c)  # insert into AFSC a's matches
    #         next_cadet = True  # move to next cadet
    #         continue  # go to beginning of while loop
    #
    #     else:  # if at capacity
    #         c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
    #         for c_hat in afsc_matches[a]:
    #             if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
    #                 c_ = c_hat
    #         if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
    #             cadet_ranking[c].remove(a)  # remove a from c's ranking list
    #             next_cadet = False  # keep trying to match this cadet
    #             continue  # go to beginning of while loop
    #
    #         else:  # if c ranks higher than c_
    #             afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
    #             cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
    #             M.pop(c_)  # remove c_ from M
    #             M[c] = a  # add match to M
    #             next_cadet = False
    #             afsc_matches[a].append(c)  # insert into AFSC a's matches
    #             c = c_
    #             continue

    print("Total iterations:", iter_num)

    print("Time to run matching algorithm:", Classical_HR_algorithm_time - start_time)

    print('Length of cadets still remaining to match in unmatched list:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    # print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))
    print('Working on the best cadet for each AFSC in cannot_match list...')
    best_cadet_start_time = time.time()

    for c in cannot_match[:]:
        a = highest_percentile_afsc(c, instance)
        # Check if AFSC is at capacity. If not, add cadet.
        if len(afsc_matches[a]) < afsc_capacities[a]:
            afsc_matches[a].append(c)
            M[c] = a
            cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                a]]
            if len(cannot_match) != 0:
                c = cannot_match[0]
        # Check if AFSC is over capacity. If less than 120%, add cadet.
        elif len(afsc_matches[a]) >= afsc_capacities[a]:
            if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                afsc_capacities[a] = afsc_capacities[a] + 1
                afsc_matches[a].append(c)
                # Update M and update cannot_match list
                M[c] = a
                cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                    a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                if len(cannot_match) != 0:
                    c = cannot_match[0]
            else:
                cannot_match_due_to_cap.append(c)
                cannot_match[:] = [c for c in cannot_match if c not in cannot_match_due_to_cap]
                if len(cannot_match) != 0:
                    c = cannot_match[0]

    for c in cannot_match_due_to_cap[:]:
        percentile_dict = next_highest_percentile_afsc(c, instance)
        a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
        while c in cannot_match_due_to_cap:
            # delete minimum key : value pair from dict
            if len(afsc_matches[a]) < afsc_capacities[a]:
                M[c] = a
                cannot_match_due_to_cap[:] = [c for c in cannot_match_due_to_cap if c not in M]
                continue
            if len(afsc_matches[a]) >= afsc_capacities[a]:
                if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match_due_to_cap[:] = [c for c in cannot_match_due_to_cap if c not in M]
                else:
                    del percentile_dict[a]
                    a = [k for k, v in percentile_dict.items() if v == min(percentile_dict.values())][0]
                    continue
    best_cadet_end_time = time.time()
    total_time_for_best_cadet = best_cadet_end_time - best_cadet_start_time
    print('Total time to find the best cadet for each AFSC in the cannot_match list:', total_time_for_best_cadet)

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            cadet_score_list.append(afsc_scoring[a][c])
        lowest_score = min(cadet_score_list, default=cadet_score_list)
        lowest_scoring_cadet_in_afsc[a] = lowest_score

    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    afsc_avg_merit = {}

    # Creates font like that of Plotly graphs - 'Open Sans'
    import matplotlib.font_manager
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams
    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    # matplotlib.font_manager.findSystemFonts(fontpaths=None, fontext='ttf')[:10]

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=16)
    plt.ylabel('Average Merit of AFSC', size=16)
    plt.title('Average Merit of Large AFSCs', size=18)
    plt.ylim(.2, 0.8)
    plt.show()


    #
    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    # save current solutions
    # df_solutions = pd.read_excel(file_name, sheet_name='Solutions', engine='openpyxl', index_col=0)
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'afsc_chooses_best_cadet_to_match_all_w_pre_process'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})


    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='Cadet', right_on='HR Lower Quotas w/ Max Capacities')

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")


    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_index=True, right_index=True)
    # print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='HR Lower Quotas w/ Unlimited Capacities', right_on='HR Lower Quotas w/ Max Capacities')
    # The above line did append but it was the exact same column and went from 0 to 183,222 cadets.

    # save current and new solutions by deleting current sheet and creating a new one
    # ExcelWorkbook = load_workbook(file_name)
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    # print(a)
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index


    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    from collections import Counter
    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts

    # Build dictionary for the chart to have a legend that reflects the cadet's ith choice of the AFSC they are matched with.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T

    # Plotly graph below
    import plotly.express as px
    fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart",
                 labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"},
                 category_orders={"variable": ["Choice 1", "Choice 2", "Choice 3", "Choice 4", "Choice 5", "Choice 6",
                                               "Choice 7", "Choice 8", "Choice 9", "Choice 10", "Choice 11",
                                               "Choice 12",
                                               "Choice 13", "Choice 14", "Choice 15", "Choice 16", "Choice 17",
                                               "Choice 18",
                                               "Choice 19", "Choice 20", "Choice 21", "Choice 22", "Choice 23",
                                               "Choice 24", "Choice 25", "Choice 26", "Choice 27", "Choice 28",
                                               "Choice 29", "Choice 30", "Choice 31", "Choice 31", ]})
    fig.show()
    # -------------------------- NEXT FIGURE ------------------------------------ #

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets


    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] < 6:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'

    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T

    # Plotly graph below

    import plotly.express as px
    fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier with AFSC Match",
                        labels={"value": "Number of Cadets", "index": "AFSC",
                                "variable": "Cadet Degree Tier and Vol Status"})
    fig_degree.show()

    # Must initialize cadet ranking dictionary again because we eliminate values in the HR algorithm. This is used in the cadet choice dict.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    # print(a)
                    afscs_from_cadets.append(a)
                    # print(afscs_from_cadets)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    fig_choice_merit = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC")
    fig_choice_merit.show()

    # -------------------------- NEXT FIGURE ------------------------------------ #

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC")
    box_fig.show()


def func_classic_hr_alg_1_0(instance):
    """This is the pure HR algorithm using deferred acceptance and max capacities"""

    import time
    import numpy as np
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections
    import matplotlib.pyplot as plt

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:
    # My ranking for AFSCs based on merit.

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]
    # print(cadet_merit_list)
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # print(sorted_merit_list)

    # for m in sorted_merit_list:
    # for c in cadet_list:
    # unmatched_cadets_list.append(c)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # print(unmatched_ordered_cadets)

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result
    # print(ranking[a])
    # print(ranking.keys())
    # print('-------------------------')
    # for k, v in ranking.items():
    #     print(k, v)

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    unchanging_lower_quotas
    afsc_avg_merit = {}

    # Creates font like that of Plotly graphs - 'Open Sans'
    import matplotlib.font_manager
    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams
    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    # matplotlib.font_manager.findSystemFonts(fontpaths=None, fontext='ttf')[:10]

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=16)
    plt.ylabel('Average Merit of AFSC', size=16)
    plt.title('Average Merit of Large AFSCs', size=18)
    plt.ylim(.2, 0.8)
    plt.show()

    # #return M
    #
    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    # df_solutions = pd.read_excel(file_name, sheet_name='Solutions', engine='openpyxl', index_col=0)
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_classic_hr_alg_1_0'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='Cadet', right_on='HR Lower Quotas w/ Max Capacities')

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_index=True, right_index=True)
    # print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='HR Lower Quotas w/ Unlimited Capacities', right_on='HR Lower Quotas w/ Max Capacities')
    # The above line did append but it was the exact same column and went from 0 to 183,222 cadets.

    # save current and new solutions by deleting current sheet and creating a new one
    # ExcelWorkbook = load_workbook(file_name)
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_0(instance):
    """This algorithm has the capacities manually altered to satisfy the exact conditions of 2023 instance
     to match all cadets using only the classic HR algorithm and modified upper capacities.

     THIS IS NOT WORKING CORRECTLY. TO fix, must explore all unmatched cadets and increment capacities up 1 past max
     capacity which should be what I already did, however, cadets are still entering cannot match. This iteration was
     performed to show it was possible to match all cadets for 2023 but holds no value because we aren't going to give
     the low merit, unmatched cadets their first choice AFSCs out of convenience."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40,
                      126, 48, 24, 4, 101, 77, 38]
    afsc_capacities = dict(zip(afscs, pgl_capacities))

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:
    # My ranking for AFSCs based on merit.

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]
    # print(cadet_merit_list)
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # print(sorted_merit_list)

    # for m in sorted_merit_list:
    # for c in cadet_list:
    # unmatched_cadets_list.append(c)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # print(unmatched_ordered_cadets)

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result
    # print(ranking[a])
    # print(ranking.keys())
    # print('-------------------------')
    # for k, v in ranking.items():
    #     print(k, v)

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    print('\nM', M)
    print("Total iterations for hr_alg_1_0_0:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))
    print('Length of Match:', len(M))

    print("Total time:", time.time() - start_time)

    """
    Script for appending sheet with data to an existing Excel workbook.
    """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    # df_solutions = pd.read_excel(file_name, sheet_name='Solutions', engine='openpyxl', index_col=0)
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_0'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='Cadet', right_on='HR Lower Quotas w/ Max Capacities')

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_index=True, right_index=True)
    # print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='HR Lower Quotas w/ Unlimited Capacities', right_on='HR Lower Quotas w/ Max Capacities')
    # The above line did append but it was the exact same column and went from 0 to 183,222 cadets.

    # save current and new solutions by deleting current sheet and creating a new one
    # ExcelWorkbook = load_workbook(file_name)
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_1(instance):
    """This algorithm loops through the remaining 25 cadets for instance data 2023 with max capacities."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:
    # My ranking for AFSCs based on merit.

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]
    # print(cadet_merit_list)
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # print(sorted_merit_list)

    # for m in sorted_merit_list:
    # for c in cadet_list:
    # unmatched_cadets_list.append(c)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # print(unmatched_ordered_cadets)

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result
    # print(ranking[a])
    # print(ranking.keys())
    # print('-------------------------')
    # for k, v in ranking.items():
    #     print(k, v)

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    best_unmatched_cadet = {}  # This is a new dict that has keys of previously unmatched cadets with a value of the AFSC that scored the cadet the highest.
    # unmatched_afsc_score = {} # This is a new dict that has the AFSC as the key with the afsc's associated value that it scored the cadet. This allow me to then pull the max value for the AFSC that scored the cadet the highest of all eligible AFSCS and return the key (which is the AFSC) associated with that value to add it as the value for the best_unmatched_cadet dict. Then I can update the matchings based on this dictionary and update AFSC_matches list.
    for c in cannot_match:
        afscs_score_of_cannot_match_cadets = []
        print('-------------------------------------', c)
        unmatched_afsc_score = {}
        for a in afsc_matches:
            if c in afsc_ranking[a]:
                if len(afsc_matches[a]) >= afsc_capacities[a]:
                    if c in afsc_scoring[a]:
                        iter_num += 1
                        score_array = afsc_scoring[a][c]
                        print(a, c, score_array)
                        unmatched_afsc_score[a] = afsc_scoring[a][c]

                        # afsc_desiring_unmatched_cadet = max(unmatched_afsc_score, key=unmatched_afsc_score.get) #gets the first highest score in the dictionary
                        afsc_desiring_unmatched_cadet = [k for k, v in unmatched_afsc_score.items() if v == max(
                            unmatched_afsc_score.values())]  # gets all of the max values in the dict so I can see ties.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[
                    0]  # pulling index so I don't return a list, but instead a str

        if len(afsc_desiring_unmatched_cadet) > 1:  # if I have a tie, break tie with capacity
            iter_num += 1
            least_capacitated_afsc = {}  # Create dict of all the capacities of the AFSCs being considered with ties.
            for a in afsc_desiring_unmatched_cadet:
                least_capacitated_afsc[a] = len(afsc_matches[a]) / afsc_capacities[a] * 100  # Calculate the capacities
                print('least capacity:', least_capacitated_afsc)
                afsc_desiring_unmatched_cadet = [k for k, v in least_capacitated_afsc.items() if v == min(
                    least_capacitated_afsc.values())]  # Select the AFSC with the lowest overage in their capacity. This is the AFSC the cadet will match with.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[0]

                print(best_unmatched_cadet[c])
                print(afsc_desiring_unmatched_cadet[0])
                if len(afsc_desiring_unmatched_cadet) > 1:
                    print(afsc_desiring_unmatched_cadet)
                    afsc_desiring_unmatched_cadet = afsc_desiring_unmatched_cadet[
                        0]  # pulling the first AFSC in a tie where the score tied AND the AFSC's capacity was tied. Give the first AFSC we encounter the cadet in this case.
                    best_unmatched_cadet[
                        c] = afsc_desiring_unmatched_cadet  # since we already accessed the first AFSC,just update it.

    print('Previous Capacities', afsc_capacities)
    # Need to update afsc_matches and afsc_capacities
    for c in best_unmatched_cadet:
        for a in afsc_matches:
            if a in best_unmatched_cadet[c] == a in afsc_matches:
                afsc_matches[a].append(c)
                afsc_capacities[a] = afsc_capacities[a] + 1
    print('New AFSC Capacities:', afsc_capacities)

    # Need to update M[c]
    for c in best_unmatched_cadet:
        M[c] = best_unmatched_cadet[c]

    # Need to update cannot_match
    for c in best_unmatched_cadet:
        cannot_match.remove(c)

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    """
    Script for appending sheet with data to an existing Excel workbook.
    """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_1_1(instance):
    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:
    # My ranking for AFSCs based on merit.

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]
    # print(cadet_merit_list)
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)
    # print(sorted_merit_list)

    # for m in sorted_merit_list:
    # for c in cadet_list:
    # unmatched_cadets_list.append(c)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # print(unmatched_ordered_cadets)

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result
    # print(ranking[a])
    # print(ranking.keys())
    # print('-------------------------')
    # for k, v in ranking.items():
    #     print(k, v)

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:10]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

            elif c in afsc_ranking[a][0:100]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

            elif c in afsc_ranking[a][0:500]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

            elif c in afsc_ranking[a][0:1000]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

            elif c in afsc_ranking[a][0:1500]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]  # Must use list
                    # comprehension to build a copy of cannot_match in-place to not have the skipping
                    # effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    """
     Script for appending sheet with data to an existing Excel workbook.
     """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_1_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_2(instance):
    """This algorithm sets all of the capcities to equal the minimum quota for the 2023 data instance, allowing for all
     preferences to be included in the cadet's preference list."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
                      34, 12, 2, 69, 50, 34]
    afsc_capacities = dict(zip(afscs, pgl_capacities))

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    """
     Script for appending sheet with data to an existing Excel workbook.
     """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_2'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_2_1(instance):
    """ This has all capacities set to min's for 2023 data and AFSC's choosing from top 10, 50,100... 1500 cadets."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections
    import matplotlib.pyplot as plt
    import numpy as np

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
                      34, 12, 2, 69, 50, 34]
    afsc_capacities = dict(zip(afscs, pgl_capacities))
    afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue


    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))
    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:10]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:100]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:500]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:1000]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]  # Must use list
                    # comprehension to build a copy of cannot_match in-place to not have the skipping
                    # effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]

            # elif c in afsc_ranking[a][0:len(cadet_ranking.keys())]:
            #     if len(afsc_matches[a]) < afsc_capacities[a]:
            #         afsc_matches[a].append(c)
            #     elif len(afsc_matches[a]) >= afsc_capacities[a]:
            #         afsc_capacities[a] = afsc_capacities[a] + 1
            #         afsc_matches[a].append(c)
            #         M[c] = a
            #         cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]  # Must use list
            #         # comprehension to build a copy of cannot_match in-place to not have the skipping
            #         # effect that occurs when removing from a list that you are iterating over.
            #         if len(cannot_match) != 0:
            #             c = cannot_match[0]

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    unchanging_lower_quotas
    afsc_avg_merit = {}

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC')
    plt.ylabel('Average Merit of AFSC')
    plt.title('Average Merit of Large AFSCs')
    plt.ylim(.2, 0.8)
    plt.show()

    """
     Script for appending sheet with data to an existing Excel workbook.
     """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_2_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_2_1_1(instance):
    """ This is the tiered AFSC choice with a 20% cap on each AFSC's overclassification. This forces the overclassification
    to be spread across AFSCs which balances the merit within the required [0.35 and 0.65] threshold. This does not solve
    the lower quota problem related to 62EXC and 62EXE. It creates 209 blocking pairs for 2023."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections
    import matplotlib.pyplot as plt
    import numpy as np

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
                      34, 12, 2, 69, 50, 34]
    afsc_capacities = dict(zip(afscs, pgl_capacities))

    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_overclass_limit = dict(zip(afscs, capacities))

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    # Set up slices to iterate through
    # c_10 = slice(0, 11)
    # c_100 = slice(0, 101)
    # c_500 = slice(0, 501)
    # c_1000 = slice(0, 1001)
    # c_end = slice(0, len(cadet_ranking.keys()) + 1)
    #
    # interval_list = [c_10, c_100, c_500, c_1000, c_end]
    # for c in cannot_match[:]:
    #     for a in afsc_matches:
    #         for interval in range(len(interval_list)):
    #             if c in afsc_ranking[a][interval_list[interval]]:
    #                 if len(afsc_matches[a]) < afsc_capacities[a]:
    #                     afsc_matches[a].append(c)
    #                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
    #                     if afsc_capacities[a] <= 1.2 * afsc_overclass_limit[a]:
    #                         afsc_capacities[a] = afsc_capacities[a] + 1
    #                         afsc_matches[a].append(c)
    #                         M[c] = a
    #                         cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
    #                             a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
    #                         if len(cannot_match) != 0:
    #                             c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:10]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                        afsc_capacities[a] = afsc_capacities[a] + 1
                        afsc_matches[a].append(c)
                        M[c] = a
                        cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                            a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                        if len(cannot_match) != 0:
                            c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:100]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                        afsc_capacities[a] = afsc_capacities[a] + 1
                        afsc_matches[a].append(c)
                        M[c] = a
                        cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                            a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                        if len(cannot_match) != 0:
                            c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:500]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                        afsc_capacities[a] = afsc_capacities[a] + 1
                        afsc_matches[a].append(c)
                        M[c] = a
                        cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                            a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                        if len(cannot_match) != 0:
                            c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:1000]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                        afsc_capacities[a] = afsc_capacities[a] + 1
                        afsc_matches[a].append(c)
                        M[c] = a
                        cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                            a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                        if len(cannot_match) != 0:
                            c = cannot_match[0]

    for c in cannot_match[:]:
        for a in afsc_matches:
            if c in afsc_ranking[a][0:]:
                if len(afsc_matches[a]) < afsc_capacities[a]:
                    afsc_matches[a].append(c)
                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    if len(afsc_matches[a]) <= 1.2 * afsc_overclass_limit[a]:
                        afsc_capacities[a] = afsc_capacities[a] + 1
                        afsc_matches[a].append(c)
                        M[c] = a
                        cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]  # Must use list
                        # comprehension to build a copy of cannot_match in-place to not have the skipping
                        # effect that occurs when removing from a list that you are iterating over.
                        if len(cannot_match) != 0:
                            c = cannot_match[0]

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    afsc_avg_merit = {}

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC')
    plt.ylabel('Average Merit of AFSC')
    plt.title('Average Merit of Large AFSCs')
    plt.ylim(.2, 0.8)
    plt.show()

    # Check blocking pairs --------------------------------------------------------------
    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        # print(a)
        cadet_score_list = []
        for c in afsc_matches[a]:
            # print(c)
            cadet_score_list.append(afsc_scoring[a][c])
            # print(cadet_score_list)
        lowest_score = min(cadet_score_list)
        lowest_scoring_cadet_in_afsc[a] = lowest_score
        # print(lowest_score)
    # print(lowest_scoring_cadet_in_afsc)

    # --------------------------------------------------------------
    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.

    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            # print('cadet', c)
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            # print('cadet',c, 'not in afsc',a)
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    # print('cadet',c,"'s score for afsc",a,':', cadet_scoring[c][a])
                                    # print('cadet',c,"'s score for afsc",current_afsc,':', cadet_scoring[c][current_afsc])
                                    # # print('looking at cadet',c)
                                    # print('-----------')
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a
            # print(AFSCs_cadets_desire_morethan_current_afsc)
            # print('--------------------------------------------------------------------')

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            # afscs_that_desire_cadets_they_are_not_matched_with = []
            # print('cadet', c)
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                # print('current1:',current_afsc)
                # print('afsc1', a)
                for a in afsc_matches:
                    if a != current_afsc:

                        # print('current2:',current_afsc)
                        # print('afsc2', a)
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set
                                    # print('afsc',a, 'scores cadet', c, 'higher than their lowest scored cadet',lowest_scoring_cadet_in_afsc[a] )

    # print("dict of cadet keys with list of afscs that desire the cadet more than the cadet's current afsc desires the cadet",cadets_other_AFSCs_desire_more)

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('d1', d1)
    print('d2', d2)
    print('Blocking pairs dictionary:', blocking_pairs_dictionary)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    """
     Script for appending sheet with data to an existing Excel workbook.
     """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_2_1_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_0_2_2(instance):
    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
                      34, 12, 2, 69, 50, 34]
    afsc_capacities = dict(zip(afscs, pgl_capacities))
    afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    for c in cannot_match[:]:
        if len(cannot_match) == 0:
            break
        for a in afsc_matches:
            if c in afsc_ranking[a][0:]:
                iter_num += 1
                if len(afsc_matches[a]) < afsc_capacities[a]:  # This is possible because we were preventing blocking
                    # pairs before, now we are creating blocking pairs by allowing to pull cadets that the AFSC scored.
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]
                    c = cannot_match[0]

                elif len(afsc_matches[a]) >= afsc_capacities[a]:
                    afsc_capacities[a] = afsc_capacities[a] + 1
                    afsc_matches[a].append(c)
                    M[c] = a
                    cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
                        a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
                    if len(cannot_match) != 0:
                        c = cannot_match[0]
                    else:
                        break

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    unchanging_lower_quotas
    afsc_avg_merit = {}

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC')
    plt.ylabel('Average Merit of AFSC')
    plt.title('Average Merit of Large AFSCs')
    plt.ylim(.2, 0.8)
    plt.show()

    """
     Script for appending sheet with data to an existing Excel workbook.
     """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_0_2_2'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


# def func_hr_alg_1_1(instance):
    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    """This alg only considers a cadet's first 6 preferences. Max capacities set."""
    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        # print(a)
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            # print(c)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                # print(cadet_score_of_afsc)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            # print(cadets_from_afscs)
            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    for c in cadet_ranking:  # MODIFICATION
        cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    """
        Script for appending sheet with data to an existing Excel workbook.
        """

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


# def func_hr_alg_1_1_1(instance):
#     import time
#     import pandas as pd
#     from openpyxl import load_workbook
#     from datetime import datetime
#     import collections
#
#     start_time = time.time()
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # My Score Calculations for CFMs
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     # My Score Calculations for CFMs
#     # My Score Calculations for CFMs
#     import numpy as np
#     np.random.seed(2)
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # How AFSCs rank cadets within each AFSC:
#
#     AFSC_scores = {}
#     for a in range(instance.parameters['M']):
#         afsc_scored_cadets = []
#         for c in range(instance.parameters['N']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
#             else:
#                 # cadet_scores.append(instance.parameters['merit'][c]*0)
#                 afsc_scored_cadets.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             AFSC_scores[a] = afsc_scored_cadets
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     cadet_scores = {}
#     for c in range(instance.parameters['N']):
#         cadet_scored_afscs = []
#         for a in range(instance.parameters['M']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(2 + np.random.random())
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(1 + np.random.random())
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
#             else:
#                 # cadet_scores.append(instance.parameters['merit'][c]*0)
#                 cadet_scored_afscs.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             cadet_scores[c] = cadet_scored_afscs
#
#     # Try to make full list of all cadets in merit order for unmatched cadet list
#     cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
#     cadet_merit_list = [m for m in instance.parameters['merit']]
#     sorted_merit_list = sorted(cadet_merit_list, reverse=True)
#
#     unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
#                                 sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
#
#     # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
#     ranking = {}
#     for a in range(instance.parameters['M']):
#         # print(a)
#         sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
#         result = []
#         for i, (value, index) in enumerate(sorted_scores):
#             result.append(index)
#         ranking[a] = result
#
#     # G2G - This is relied upon to  build the other dictionaries as well.
#     # The below builds the score that a cadet assigns to each AFSC.
#     cadet_score_from_AFSC = {}
#     # c0_scores_list = []
#     for c in range(instance.parameters['N']):
#         cadet_of_AFSC_score_list = []
#         for a in AFSC_scores.keys():
#             s = cadet_scores[c][a]
#             # print(s)
#             cadet_of_AFSC_score_list.append(s)
#             # print(cadet_of_AFSC_score_list)
#             cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list
#
#     # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.
#
#     for c in range(instance.parameters['N']):
#         AFSC_list = np.where(instance.parameters['utility'][c] > 0)
#         # print(np.where(instance.parameters['utility'][c] > 0))
#         for a in AFSC_list[0]:
#             if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
#                 cadet_score_from_AFSC[c][a] = -1
#             # print()
#             else:
#                 cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7
#
#     cadet_scoring = {}
#     for i in range(instance.parameters['N']):
#         # print(i)
#         #     #print(cadet_score_from_AFSC[c])
#         c = 'c' + str(i)
#         cadet_scoring[c] = {}
#         # print(c)
#         for j in range(instance.parameters['M']):
#             a = instance.parameters['afsc_vector'][j]
#             # print(a)
#             score = cadet_score_from_AFSC[i][j]
#             if score > 0:  # added
#                 cadet_scoring[c][a] = score  # added
#
#     cadet_ranked_AFSC = {}
#     for c in range(instance.parameters['N']):
#         sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
#         result2 = []
#         for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
#             result2.append(index)
#         cadet_ranked_AFSC[c] = result2
#
#     # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
#     cadet_ranking = {}
#     for i in range(instance.parameters['N']):
#         # print(i)
#         afscs_from_cadets = []
#         c = 'c' + str(i)
#         for j in range(instance.parameters['M']):
#             # print(j)
#             AFSC = cadet_ranked_AFSC[i][j]
#             # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
#             a = instance.parameters['afsc_vector'][AFSC]
#             if a in cadet_scoring[c].keys():
#                 if cadet_scoring[c][a] > 0:
#                     afscs_from_cadets.append(a)
#
#             # print(afscs_from_cadets)
#             # print(AFSC)
#             cadet_ranking[c] = afscs_from_cadets
#
#     # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.
#
#     afsc_scoring = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         afsc_scoring[a] = {}
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             # print(c)
#             cadet_score_of_afsc = AFSC_scores[j][i]
#             if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
#                 c = 'c' + str(i)
#                 # print(cadet_score_of_afsc)
#                 afsc_scoring[a][c] = cadet_score_of_afsc
#
#     # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.
#
#     afsc_ranking = {}
#     for j in range(instance.parameters['M']):
#         cadets_from_afscs = []
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             if c in afsc_scoring[a].keys():
#                 if afsc_scoring[a][c] > 0:
#                     cadets_from_afscs.append(c)
#
#             # print(cadets_from_afscs)
#             afsc_ranking[a] = cadets_from_afscs
#
#     # #Actual Max Capacities for AFSCs
#     afscs = instance.parameters['afsc_vector']
#     capacities = instance.parameters["quota_max"]
#     afsc_capacities = dict(zip(afscs, capacities))
#
#     # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
#     # afscs = instance.parameters['afsc_vector']
#     # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
#     # afsc_capacities = dict(zip(afscs, pgl_capacities))
#     # afsc_capacities
#
#     # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
#     # afscs = instance.parameters['afsc_vector']
#     # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
#     # afsc_capacities = dict(zip(afscs, pgl_capacities))
#     # afsc_capacities
#
#     afsc_matches = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             i = []
#         afsc_matches[a] = i
#
#     # Pseudo code for CAMP HR Algorithm
#     # import tqdm from tqdm_notebook as tqdm
#     # input data
#     # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
#     # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
#     # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
#     # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
#     # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
#     # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity
#
#     matching_data_structure_creation = time.time()
#     print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)
#
#     for c in cadet_ranking:  # MODIFICATION
#         cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices
#
#     # initialize parameters
#     unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
#     cannot_match = []  # cadets with empty ranking lists
#     M = {}  # matches
#
#     iter_num = 0
#     iter_limit = 10000
#     next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
#     while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
#         iter_num += 1
#         if next_cadet:
#             c = unmatched_ordered_cadets[0]  # next cadet to match
#             unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
#
#         if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
#             cannot_match.append(c)  # add them to cannot match list
#             next_cadet = True
#             continue  # go to beginning of while loop
#         a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
#
#         if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
#             M[c] = a  # add match to M
#             afsc_matches[a].append(c)  # insert into AFSC a's matches
#             next_cadet = True  # move to next cadet
#             continue  # go to beginning of while loop
#
#         else:  # if at capacity
#             c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
#             for c_hat in afsc_matches[a]:
#                 if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
#                     c_ = c_hat
#             if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
#                 cadet_ranking[c].remove(a)  # remove a from c's ranking list
#                 next_cadet = False  # keep trying to match this cadet
#                 continue  # go to beginning of while loop
#
#             else:  # if c ranks higher than c_
#                 afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
#                 cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
#                 M.pop(c_)  # remove c_ from M
#                 M[c] = a  # add match to M
#                 next_cadet = False
#                 afsc_matches[a].append(c)  # insert into AFSC a's matches
#                 c = c_
#                 continue
#
#     print('\nM', M)
#     print("Total iterations:", iter_num)
#
#     Classical_HR_algorithm_time = time.time()
#     print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)
#
#     print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
#     print('Length of perm unmatched cadets:', len(cannot_match))
#
#     print("Total time:", time.time() - start_time)
#
#     # Trying to loop through those who could not be matched to match them with the career field that wants them.
#     print('# unmatched', len(cannot_match))
#
#     for c in cannot_match[:]:
#         if len(cannot_match) == 0:
#             break
#         for a in afsc_matches:
#             if c in afsc_ranking[a][0:]:
#                 iter_num += 1
#                 if len(afsc_matches[a]) < afsc_capacities[a]:  # This is possible because we were preventing blocking
#                     # pairs before, now we are creating blocking pairs by allowing to pull cadets that the AFSC scored.
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[a]]
#                     c = cannot_match[0]
#
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#                     else:
#                         break
#
#     print('Total Iterations', iter_num)
#     print('Cannot Match still:', len(cannot_match))
#     print('Length of M:', len(M))
#     # print('AFSC_Matches:', afsc_matches)
#
#     afsc_matches_len = []
#     for a in afsc_matches:
#         a = ('Matched:', len(afsc_matches[a]))
#         afsc_matches_len.append(a)
#
#     quota_min = []
#     for q in instance.parameters["quota_min"]:
#         q = ('min:', q)
#         quota_min.append(q)
#
#     max_quotas = []
#     for m in instance.parameters["quota_max"]:
#         m = ('max:', m)
#         max_quotas.append(m)
#
#     quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
#     print(quota_min_and_match_results)
#
#     print(M)
#
#     for c in cadet_ranking.keys():
#         if c not in M:
#             print("I'm still not in the overall Match! ==>", c)
#     # Appending sheet with data to an existing Excel workbook.
#
#     # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
#     new_M = {}
#     for c in M.keys():
#         i = int(c[1:])
#         new_M[i] = M[c]  # getting the index of cadet in M
#
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     ordered_M = dict(ordered_M)  # convert the ordering to a dictionary
#
#     all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
#     empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
#     solution_missing_cadets = ordered_M
#
#     # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
#     complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}
#
#     file_name = 'Real 2023 MacDonald Results.xlsx'
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     path = os.getcwd()
#
#     results_by_index = []
#     for i in complete_cadet_M_list.keys():
#         results_by_index.append(complete_cadet_M_list[i])
#     print(results_by_index)
#     # save current solutions
#     df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
#                                  engine='openpyxl', index_col=0)
#
#     # append new solution to current solutions
#     solution_name = 'func_hr_alg_1_1_1'
#     if solution_name in df_solutions.columns:
#         solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")
#
#     matching_list = results_by_index
#     df_new_solution = pd.DataFrame({solution_name: matching_list})
#     print(df_new_solution)
#
#     df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
#     print(df_solutions)
#
#     # save current and new solutions by deleting current sheet and creating a new one
#     ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
#     if 'Solutions' in ExcelWorkbook.sheetnames:
#         ExcelWorkbook.remove(ExcelWorkbook['Solutions'])
#
#     writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
#                             engine='openpyxl')  # Had to manually give directory
#     writer.book = ExcelWorkbook
#     df_solutions.to_excel(writer, sheet_name='Solutions')
#     # df_new_solution.to_excel(writer, sheet_name='Solutions')
#     writer.save()
#     writer.close()
#

# def func_hr_alg_1_1_1_1(instance):
#     """ This algorithm runs through a cadet's top 6 choices and considers an AFSCs, top 10, 50, 100,...1500 cadets to
#      match. Max capacities for 2023 data. """
#
#     import time
#     import pandas as pd
#     from openpyxl import load_workbook
#     from datetime import datetime
#     import collections
#
#     start_time = time.time()
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # My Score Calculations for CFMs
#     import numpy as np
#     np.random.seed(2)
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # How AFSCs rank cadets within each AFSC:
#
#     AFSC_scores = {}
#     for a in range(instance.parameters['M']):
#         afsc_scored_cadets = []
#         for c in range(instance.parameters['N']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
#             else:
#                 afsc_scored_cadets.append(0)
#             # If AFSC score is below 0 then remove them from matching possibilities.
#
#             AFSC_scores[a] = afsc_scored_cadets
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     cadet_scores = {}
#     for c in range(instance.parameters['N']):
#         cadet_scored_afscs = []
#         for a in range(instance.parameters['M']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(2 + np.random.random())
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(1 + np.random.random())
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
#             else:
#                 cadet_scored_afscs.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             cadet_scores[c] = cadet_scored_afscs
#
#     # Try to make full list of all cadets in merit order for unmatched cadet list
#     cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
#     cadet_merit_list = [m for m in instance.parameters['merit']]
#     sorted_merit_list = sorted(cadet_merit_list, reverse=True)
#
#     unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
#                                 sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
#
#     # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
#     ranking = {}
#     for a in range(instance.parameters['M']):
#         sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
#         result = []
#         for i, (value, index) in enumerate(sorted_scores):
#             result.append(index)
#         ranking[a] = result
#
#     # G2G - This is relied upon to  build the other dictionaries as well.
#     # The below builds the score that a cadet assigns to each AFSC.
#     cadet_score_from_AFSC = {}
#     for c in range(instance.parameters['N']):
#         cadet_of_AFSC_score_list = []
#         for a in AFSC_scores.keys():
#             s = cadet_scores[c][a]
#             cadet_of_AFSC_score_list.append(s)
#             cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list
#
#     # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.
#
#     for c in range(instance.parameters['N']):
#         AFSC_list = np.where(instance.parameters['utility'][c] > 0)
#         for a in AFSC_list[0]:
#             if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
#                 cadet_score_from_AFSC[c][a] = -1
#             else:
#                 cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7
#
#     cadet_scoring = {}
#     for i in range(instance.parameters['N']):
#         c = 'c' + str(i)
#         cadet_scoring[c] = {}
#         for j in range(instance.parameters['M']):
#             a = instance.parameters['afsc_vector'][j]
#             # print(a)
#             score = cadet_score_from_AFSC[i][j]
#             if score > 0:  # added
#                 cadet_scoring[c][a] = score  # added
#
#     cadet_ranked_AFSC = {}
#     for c in range(instance.parameters['N']):
#         sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
#         result2 = []
#         for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
#             result2.append(index)
#         cadet_ranked_AFSC[c] = result2
#
#     # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
#     cadet_ranking = {}
#     for i in range(instance.parameters['N']):
#         afscs_from_cadets = []
#         c = 'c' + str(i)
#         for j in range(instance.parameters['M']):
#             AFSC = cadet_ranked_AFSC[i][j]
#             # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
#             a = instance.parameters['afsc_vector'][AFSC]
#             if a in cadet_scoring[c].keys():
#                 if cadet_scoring[c][a] > 0:
#                     afscs_from_cadets.append(a)
#
#             cadet_ranking[c] = afscs_from_cadets
#
#     # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.
#
#     afsc_scoring = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         afsc_scoring[a] = {}
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             cadet_score_of_afsc = AFSC_scores[j][i]
#             if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
#                 c = 'c' + str(i)
#                 afsc_scoring[a][c] = cadet_score_of_afsc
#
#     # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.
#
#     afsc_ranking = {}
#     for j in range(instance.parameters['M']):
#         cadets_from_afscs = []
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             if c in afsc_scoring[a].keys():
#                 if afsc_scoring[a][c] > 0:
#                     cadets_from_afscs.append(c)
#
#             afsc_ranking[a] = cadets_from_afscs
#
#     # #Actual Max Capacities for AFSCs
#     afscs = instance.parameters['afsc_vector']
#     capacities = instance.parameters["quota_max"]
#     afsc_capacities = dict(zip(afscs, capacities))
#
#     # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
#     # afscs = instance.parameters['afsc_vector']
#     # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
#     # afsc_capacities = dict(zip(afscs, pgl_capacities))
#     # afsc_capacities
#
#     # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
#     # afscs = instance.parameters['afsc_vector']
#     # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
#     # afsc_capacities = dict(zip(afscs, pgl_capacities))
#     # afsc_capacities
#
#     afsc_matches = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             i = []
#         afsc_matches[a] = i
#
#     # Pseudo code for CAMP HR Algorithm
#     # import tqdm from tqdm_notebook as tqdm
#     # input data
#     # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
#     # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
#     # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
#     # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
#     # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
#     # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity
#
#     matching_data_structure_creation = time.time()
#     print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)
#
#     for c in cadet_ranking:  # MODIFICATION
#         cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices
#
#     # initialize parameters
#     unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
#     cannot_match = []  # cadets with empty ranking lists
#     M = {}  # matches
#
#     iter_num = 0
#     iter_limit = 10000
#     next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
#     while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
#         iter_num += 1
#         if next_cadet:
#             c = unmatched_ordered_cadets[0]  # next cadet to match
#             unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
#
#         if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
#             cannot_match.append(c)  # add them to cannot match list
#             next_cadet = True
#             continue  # go to beginning of while loop
#         a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
#
#         if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
#             M[c] = a  # add match to M
#             afsc_matches[a].append(c)  # insert into AFSC a's matches
#             next_cadet = True  # move to next cadet
#             continue  # go to beginning of while loop
#
#         else:  # if at capacity
#             c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
#             for c_hat in afsc_matches[a]:
#                 if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
#                     c_ = c_hat
#             if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
#                 cadet_ranking[c].remove(a)  # remove a from c's ranking list
#                 next_cadet = False  # keep trying to match this cadet
#                 continue  # go to beginning of while loop
#
#             else:  # if c ranks higher than c_
#                 afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
#                 cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
#                 M.pop(c_)  # remove c_ from M
#                 M[c] = a  # add match to M
#                 next_cadet = False
#                 afsc_matches[a].append(c)  # insert into AFSC a's matches
#                 c = c_
#                 continue
#
#     print('\nM', M)
#     print("Total iterations:", iter_num)
#
#     Classical_HR_algorithm_time = time.time()
#     print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)
#
#     print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
#     print('Length of perm unmatched cadets:', len(cannot_match))
#
#     print("Total time:", time.time() - start_time)
#
#     # Trying to loop through those who could not be matched to match them with the career field that wants them.
#     print('# unmatched', len(cannot_match))
#     for c in cannot_match[:]:
#         for a in afsc_matches:
#             if c in afsc_ranking[a][0:10]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:100]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:500]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:1000]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:1500]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#     print('Cannot Match still:', len(cannot_match))
#     # print('AFSC_Matches:', afsc_matches)
#
#     afsc_matches_len = []
#     for a in afsc_matches:
#         a = ('Matched:', len(afsc_matches[a]))
#         afsc_matches_len.append(a)
#
#     quota_min = []
#     for q in instance.parameters["quota_min"]:
#         q = ('min:', q)
#         quota_min.append(q)
#
#     max_quotas = []
#     for m in instance.parameters["quota_max"]:
#         m = ('max:', m)
#         max_quotas.append(m)
#
#     quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
#     print(quota_min_and_match_results)
#
#     print(M)
#     # return M
#
#     """
#     Script for appending sheet with data to an existing Excel workbook.
#     """
#
#     # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
#     new_M = {}
#     for c in M.keys():
#         i = int(c[1:])
#         new_M[i] = M[c]  # getting the index of cadet in M
#
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     ordered_M = dict(ordered_M)  # convert the ordering to a dictionary
#
#     all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
#     empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
#     solution_missing_cadets = ordered_M
#
#     # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
#     complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}
#
#     file_name = 'Real 2023 MacDonald Results.xlsx'
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     path = os.getcwd()
#
#     results_by_index = []
#     for i in complete_cadet_M_list.keys():
#         results_by_index.append(complete_cadet_M_list[i])
#     print(results_by_index)
#     # save current solutions
#     df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
#                                  engine='openpyxl', index_col=0)
#
#     # append new solution to current solutions
#     solution_name = 'func_hr_alg_1_1_1_1'
#     if solution_name in df_solutions.columns:
#         solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")
#
#     matching_list = results_by_index
#     df_new_solution = pd.DataFrame({solution_name: matching_list})
#     print(df_new_solution)
#
#     df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
#     print(df_solutions)
#
#     # save current and new solutions by deleting current sheet and creating a new one
#     ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
#     if 'Solutions' in ExcelWorkbook.sheetnames:
#         ExcelWorkbook.remove(ExcelWorkbook['Solutions'])
#
#     writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
#                             engine='openpyxl')  # Had to manually give directory
#     writer.book = ExcelWorkbook
#     df_solutions.to_excel(writer, sheet_name='Solutions')
#     # df_new_solution.to_excel(writer, sheet_name='Solutions')
#     writer.save()
#     writer.close()


# def func_hr_alg_1_1_2(instance):
#     """ This algorithm runs through a cadet's top 6 choices with min capacities for 2023 data. Does not loop through cadets.  """
#
#     import time
#     import pandas as pd
#     from openpyxl import load_workbook
#     from datetime import datetime
#     import collections
#
#     start_time = time.time()
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # My Score Calculations for CFMs
#     import numpy as np
#     np.random.seed(2)
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # How AFSCs rank cadets within each AFSC:
#
#     AFSC_scores = {}
#     for a in range(instance.parameters['M']):
#         afsc_scored_cadets = []
#         for c in range(instance.parameters['N']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
#             else:
#                 afsc_scored_cadets.append(0)
#             # If AFSC score is below 0 then remove them from matching possibilities.
#
#             AFSC_scores[a] = afsc_scored_cadets
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     cadet_scores = {}
#     for c in range(instance.parameters['N']):
#         cadet_scored_afscs = []
#         for a in range(instance.parameters['M']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(2 + np.random.random())
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(1 + np.random.random())
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
#             else:
#                 cadet_scored_afscs.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             cadet_scores[c] = cadet_scored_afscs
#
#     # Try to make full list of all cadets in merit order for unmatched cadet list
#     cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
#     cadet_merit_list = [m for m in instance.parameters['merit']]
#     sorted_merit_list = sorted(cadet_merit_list, reverse=True)
#
#     unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
#                                 sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
#
#     # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
#     ranking = {}
#     for a in range(instance.parameters['M']):
#         sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
#         result = []
#         for i, (value, index) in enumerate(sorted_scores):
#             result.append(index)
#         ranking[a] = result
#
#     # G2G - This is relied upon to  build the other dictionaries as well.
#     # The below builds the score that a cadet assigns to each AFSC.
#     cadet_score_from_AFSC = {}
#     for c in range(instance.parameters['N']):
#         cadet_of_AFSC_score_list = []
#         for a in AFSC_scores.keys():
#             s = cadet_scores[c][a]
#             cadet_of_AFSC_score_list.append(s)
#             cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list
#
#     # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.
#
#     for c in range(instance.parameters['N']):
#         AFSC_list = np.where(instance.parameters['utility'][c] > 0)
#         for a in AFSC_list[0]:
#             if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
#                 cadet_score_from_AFSC[c][a] = -1
#             else:
#                 cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7
#
#     cadet_scoring = {}
#     for i in range(instance.parameters['N']):
#         c = 'c' + str(i)
#         cadet_scoring[c] = {}
#         for j in range(instance.parameters['M']):
#             a = instance.parameters['afsc_vector'][j]
#             # print(a)
#             score = cadet_score_from_AFSC[i][j]
#             if score > 0:  # added
#                 cadet_scoring[c][a] = score  # added
#
#     cadet_ranked_AFSC = {}
#     for c in range(instance.parameters['N']):
#         sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
#         result2 = []
#         for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
#             result2.append(index)
#         cadet_ranked_AFSC[c] = result2
#
#     # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
#     cadet_ranking = {}
#     for i in range(instance.parameters['N']):
#         afscs_from_cadets = []
#         c = 'c' + str(i)
#         for j in range(instance.parameters['M']):
#             AFSC = cadet_ranked_AFSC[i][j]
#             # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
#             a = instance.parameters['afsc_vector'][AFSC]
#             if a in cadet_scoring[c].keys():
#                 if cadet_scoring[c][a] > 0:
#                     afscs_from_cadets.append(a)
#
#             cadet_ranking[c] = afscs_from_cadets
#
#     # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.
#
#     afsc_scoring = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         afsc_scoring[a] = {}
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             cadet_score_of_afsc = AFSC_scores[j][i]
#             if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
#                 c = 'c' + str(i)
#                 afsc_scoring[a][c] = cadet_score_of_afsc
#
#     # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.
#
#     afsc_ranking = {}
#     for j in range(instance.parameters['M']):
#         cadets_from_afscs = []
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             if c in afsc_scoring[a].keys():
#                 if afsc_scoring[a][c] > 0:
#                     cadets_from_afscs.append(c)
#
#             afsc_ranking[a] = cadets_from_afscs
#
#     # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
#     afscs = instance.parameters['afsc_vector']
#     pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
#                       34, 12, 2, 69, 50, 34]
#     afsc_capacities = dict(zip(afscs, pgl_capacities))
#     afsc_capacities
#
#     afsc_matches = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             i = []
#         afsc_matches[a] = i
#
#     # Pseudo code for CAMP HR Algorithm
#     # import tqdm from tqdm_notebook as tqdm
#     # input data
#     # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
#     # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
#     # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
#     # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
#     # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
#     # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity
#
#     matching_data_structure_creation = time.time()
#     print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)
#
#     for c in cadet_ranking:  # MODIFICATION
#         cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices
#
#     # initialize parameters
#     unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
#     cannot_match = []  # cadets with empty ranking lists
#     M = {}  # matches
#
#     iter_num = 0
#     iter_limit = 10000
#     next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
#     while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
#         iter_num += 1
#         if next_cadet:
#             c = unmatched_ordered_cadets[0]  # next cadet to match
#             unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
#
#         if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
#             cannot_match.append(c)  # add them to cannot match list
#             next_cadet = True
#             continue  # go to beginning of while loop
#         a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
#
#         if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
#             M[c] = a  # add match to M
#             afsc_matches[a].append(c)  # insert into AFSC a's matches
#             next_cadet = True  # move to next cadet
#             continue  # go to beginning of while loop
#
#         else:  # if at capacity
#             c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
#             for c_hat in afsc_matches[a]:
#                 if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
#                     c_ = c_hat
#             if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
#                 cadet_ranking[c].remove(a)  # remove a from c's ranking list
#                 next_cadet = False  # keep trying to match this cadet
#                 continue  # go to beginning of while loop
#
#             else:  # if c ranks higher than c_
#                 afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
#                 cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
#                 M.pop(c_)  # remove c_ from M
#                 M[c] = a  # add match to M
#                 next_cadet = False
#                 afsc_matches[a].append(c)  # insert into AFSC a's matches
#                 c = c_
#                 continue
#
#     print('\nM', M)
#     print("Total iterations:", iter_num)
#
#     Classical_HR_algorithm_time = time.time()
#     print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)
#
#     print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
#     print('Length of perm unmatched cadets:', len(cannot_match))
#
#     print("Total time:", time.time() - start_time)
#
#     # Trying to loop through those who could not be matched to match them with the career field that wants them.
#     print('# unmatched', len(cannot_match))
#
#     print('Cannot Match still:', len(cannot_match))
#     # print('AFSC_Matches:', afsc_matches)
#
#     afsc_matches_len = []
#     for a in afsc_matches:
#         a = ('Matched:', len(afsc_matches[a]))
#         afsc_matches_len.append(a)
#
#     quota_min = []
#     for q in instance.parameters["quota_min"]:
#         q = ('min:', q)
#         quota_min.append(q)
#
#     max_quotas = []
#     for m in instance.parameters["quota_max"]:
#         m = ('max:', m)
#         max_quotas.append(m)
#
#     quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
#     print(quota_min_and_match_results)
#
#     print(M)
#     # return M
#
#     # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
#     new_M = {}
#     for c in M.keys():
#         i = int(c[1:])
#         new_M[i] = M[c]  # getting the index of cadet in M
#
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     ordered_M = dict(ordered_M)  # convert the ordering to a dictionary
#
#     all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
#     empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
#     solution_missing_cadets = ordered_M
#
#     # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
#     complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}
#
#     file_name = 'Real 2023 MacDonald Results.xlsx'
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     path = os.getcwd()
#
#     results_by_index = []
#     for i in complete_cadet_M_list.keys():
#         results_by_index.append(complete_cadet_M_list[i])
#     print(results_by_index)
#     # save current solutions
#     df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
#                                  engine='openpyxl', index_col=0)
#
#     # append new solution to current solutions
#     solution_name = 'func_hr_alg_1_1_2'
#     if solution_name in df_solutions.columns:
#         solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")
#
#     matching_list = results_by_index
#     df_new_solution = pd.DataFrame({solution_name: matching_list})
#     print(df_new_solution)
#
#     df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
#     print(df_solutions)
#
#     # save current and new solutions by deleting current sheet and creating a new one
#     ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
#     if 'Solutions' in ExcelWorkbook.sheetnames:
#         ExcelWorkbook.remove(ExcelWorkbook['Solutions'])
#
#     writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
#                             engine='openpyxl')  # Had to manually give directory
#     writer.book = ExcelWorkbook
#     df_solutions.to_excel(writer, sheet_name='Solutions')
#     # df_new_solution.to_excel(writer, sheet_name='Solutions')
#     writer.save()
#     writer.close()


# def func_hr_alg_1_1_2_1(instance):
#     """ This algorithm runs through a cadet's top 6 choices and considers an AFSCs who prefer any cadets to
#       match. Minimum capacities  for 2023 data. """
#
#     import time
#     import pandas as pd
#     from openpyxl import load_workbook
#     from datetime import datetime
#     import collections
#
#     start_time = time.time()
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # My Score Calculations for CFMs
#     import numpy as np
#     np.random.seed(2)
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # How AFSCs rank cadets within each AFSC:
#
#     AFSC_scores = {}
#     for a in range(instance.parameters['M']):
#         afsc_scored_cadets = []
#         for c in range(instance.parameters['N']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
#             else:
#                 afsc_scored_cadets.append(0)
#             # If AFSC score is below 0 then remove them from matching possibilities.
#
#             AFSC_scores[a] = afsc_scored_cadets
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     cadet_scores = {}
#     for c in range(instance.parameters['N']):
#         cadet_scored_afscs = []
#         for a in range(instance.parameters['M']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(2 + np.random.random())
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(1 + np.random.random())
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
#             else:
#                 cadet_scored_afscs.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             cadet_scores[c] = cadet_scored_afscs
#
#     # Try to make full list of all cadets in merit order for unmatched cadet list
#     cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
#     cadet_merit_list = [m for m in instance.parameters['merit']]
#     sorted_merit_list = sorted(cadet_merit_list, reverse=True)
#
#     unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
#                                 sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
#
#     # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
#     ranking = {}
#     for a in range(instance.parameters['M']):
#         sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
#         result = []
#         for i, (value, index) in enumerate(sorted_scores):
#             result.append(index)
#         ranking[a] = result
#
#     # G2G - This is relied upon to  build the other dictionaries as well.
#     # The below builds the score that a cadet assigns to each AFSC.
#     cadet_score_from_AFSC = {}
#     for c in range(instance.parameters['N']):
#         cadet_of_AFSC_score_list = []
#         for a in AFSC_scores.keys():
#             s = cadet_scores[c][a]
#             cadet_of_AFSC_score_list.append(s)
#             cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list
#
#     # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.
#
#     for c in range(instance.parameters['N']):
#         AFSC_list = np.where(instance.parameters['utility'][c] > 0)
#         for a in AFSC_list[0]:
#             if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
#                 cadet_score_from_AFSC[c][a] = -1
#             else:
#                 cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7
#
#     cadet_scoring = {}
#     for i in range(instance.parameters['N']):
#         c = 'c' + str(i)
#         cadet_scoring[c] = {}
#         for j in range(instance.parameters['M']):
#             a = instance.parameters['afsc_vector'][j]
#             # print(a)
#             score = cadet_score_from_AFSC[i][j]
#             if score > 0:  # added
#                 cadet_scoring[c][a] = score  # added
#
#     cadet_ranked_AFSC = {}
#     for c in range(instance.parameters['N']):
#         sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
#         result2 = []
#         for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
#             result2.append(index)
#         cadet_ranked_AFSC[c] = result2
#
#     # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
#     cadet_ranking = {}
#     for i in range(instance.parameters['N']):
#         afscs_from_cadets = []
#         c = 'c' + str(i)
#         for j in range(instance.parameters['M']):
#             AFSC = cadet_ranked_AFSC[i][j]
#             # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
#             a = instance.parameters['afsc_vector'][AFSC]
#             if a in cadet_scoring[c].keys():
#                 if cadet_scoring[c][a] > 0:
#                     afscs_from_cadets.append(a)
#
#             cadet_ranking[c] = afscs_from_cadets
#
#     # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.
#
#     afsc_scoring = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         afsc_scoring[a] = {}
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             cadet_score_of_afsc = AFSC_scores[j][i]
#             if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
#                 c = 'c' + str(i)
#                 afsc_scoring[a][c] = cadet_score_of_afsc
#
#     # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.
#
#     afsc_ranking = {}
#     for j in range(instance.parameters['M']):
#         cadets_from_afscs = []
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             if c in afsc_scoring[a].keys():
#                 if afsc_scoring[a][c] > 0:
#                     cadets_from_afscs.append(c)
#
#             afsc_ranking[a] = cadets_from_afscs
#
#     # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
#     afscs = instance.parameters['afsc_vector']
#     pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
#                       34, 12, 2, 69, 50, 34]
#     afsc_capacities = dict(zip(afscs, pgl_capacities))
#     afsc_capacities
#
#     afsc_matches = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             i = []
#         afsc_matches[a] = i
#
#     # Pseudo code for CAMP HR Algorithm
#     # import tqdm from tqdm_notebook as tqdm
#     # input data
#     # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
#     # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
#     # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
#     # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
#     # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
#     # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity
#
#     matching_data_structure_creation = time.time()
#     print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)
#
#     for c in cadet_ranking:  # MODIFICATION
#         cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices
#
#     # initialize parameters
#     unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
#     cannot_match = []  # cadets with empty ranking lists
#     M = {}  # matches
#
#     iter_num = 0
#     iter_limit = 10000
#     next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
#     while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
#         iter_num += 1
#         if next_cadet:
#             c = unmatched_ordered_cadets[0]  # next cadet to match
#             unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
#
#         if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
#             cannot_match.append(c)  # add them to cannot match list
#             next_cadet = True
#             continue  # go to beginning of while loop
#         a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
#
#         if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
#             M[c] = a  # add match to M
#             afsc_matches[a].append(c)  # insert into AFSC a's matches
#             next_cadet = True  # move to next cadet
#             continue  # go to beginning of while loop
#
#         else:  # if at capacity
#             c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
#             for c_hat in afsc_matches[a]:
#                 if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
#                     c_ = c_hat
#             if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
#                 cadet_ranking[c].remove(a)  # remove a from c's ranking list
#                 next_cadet = False  # keep trying to match this cadet
#                 continue  # go to beginning of while loop
#
#             else:  # if c ranks higher than c_
#                 afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
#                 cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
#                 M.pop(c_)  # remove c_ from M
#                 M[c] = a  # add match to M
#                 next_cadet = False
#                 afsc_matches[a].append(c)  # insert into AFSC a's matches
#                 c = c_
#                 continue
#
#     print('\nM', M)
#     print("Total iterations:", iter_num)
#
#     Classical_HR_algorithm_time = time.time()
#     print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)
#
#     print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
#     print('Length of perm unmatched cadets:', len(cannot_match))
#
#     print("Total time:", time.time() - start_time)
#
#     # Trying to loop through those who could not be matched to match them with the career field that wants them.
#     print('# unmatched', len(cannot_match))
#     for c in cannot_match[:]:
#         for a in afsc_matches:
#             if c in afsc_ranking[a][0:]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#     print('Cannot Match still:', len(cannot_match))
#     # print('AFSC_Matches:', afsc_matches)
#
#     afsc_matches_len = []
#     for a in afsc_matches:
#         a = ('Matched:', len(afsc_matches[a]))
#         afsc_matches_len.append(a)
#
#     quota_min = []
#     for q in instance.parameters["quota_min"]:
#         q = ('min:', q)
#         quota_min.append(q)
#
#     max_quotas = []
#     for m in instance.parameters["quota_max"]:
#         m = ('max:', m)
#         max_quotas.append(m)
#
#     quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
#     print(quota_min_and_match_results)
#
#     print(M)
#     # return M
#
#     # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
#     new_M = {}
#     for c in M.keys():
#         i = int(c[1:])
#         new_M[i] = M[c]  # getting the index of cadet in M
#
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     ordered_M = dict(ordered_M)  # convert the ordering to a dictionary
#
#     all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
#     empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
#     solution_missing_cadets = ordered_M
#
#     # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
#     complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}
#
#     file_name = 'Real 2023 MacDonald Results.xlsx'
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     path = os.getcwd()
#
#     results_by_index = []
#     for i in complete_cadet_M_list.keys():
#         results_by_index.append(complete_cadet_M_list[i])
#     print(results_by_index)
#     # save current solutions
#     df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
#                                  engine='openpyxl', index_col=0)
#
#     # append new solution to current solutions
#     solution_name = 'func_hr_alg_1_1_2_1'
#     if solution_name in df_solutions.columns:
#         solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")
#
#     matching_list = results_by_index
#     df_new_solution = pd.DataFrame({solution_name: matching_list})
#     print(df_new_solution)
#
#     df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
#     print(df_solutions)
#
#     # save current and new solutions by deleting current sheet and creating a new one
#     ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
#     if 'Solutions' in ExcelWorkbook.sheetnames:
#         ExcelWorkbook.remove(ExcelWorkbook['Solutions'])
#
#     writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
#                             engine='openpyxl')  # Had to manually give directory
#     writer.book = ExcelWorkbook
#     df_solutions.to_excel(writer, sheet_name='Solutions')
#     # df_new_solution.to_excel(writer, sheet_name='Solutions')
#     writer.save()
#     writer.close()
#
#
# def func_hr_alg_1_1_2_2(instance):
#     """ This algorithm runs through a cadet's top 6 choices and considers an AFSCs, top 10, 50, 100,...1500 cadets to
#       match. Minimum capacities for 2023 data. """
#
#     import time
#     import pandas as pd
#     from openpyxl import load_workbook
#     from datetime import datetime
#     import collections
#
#     start_time = time.time()
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # My Score Calculations for CFMs
#     import numpy as np
#     np.random.seed(2)
#
#     instance_creation_time = time.time()
#     print("Time to create instance", instance_creation_time - start_time)
#
#     # How AFSCs rank cadets within each AFSC:
#
#     AFSC_scores = {}
#     for a in range(instance.parameters['M']):
#         afsc_scored_cadets = []
#         for c in range(instance.parameters['N']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
#             else:
#                 afsc_scored_cadets.append(0)
#             # If AFSC score is below 0 then remove them from matching possibilities.
#
#             AFSC_scores[a] = afsc_scored_cadets
#
#     # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
#     cadet_scores = {}
#     for c in range(instance.parameters['N']):
#         cadet_scored_afscs = []
#         for a in range(instance.parameters['M']):
#             if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
#                 cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
#             elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(2 + np.random.random())
#             elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(1 + np.random.random())
#             elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
#                 cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
#             else:
#                 cadet_scored_afscs.append(0)
#             # If cadet score is below 0 then remove them from matching possibilities.
#
#             cadet_scores[c] = cadet_scored_afscs
#
#     # Try to make full list of all cadets in merit order for unmatched cadet list
#     cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
#     cadet_merit_list = [m for m in instance.parameters['merit']]
#     sorted_merit_list = sorted(cadet_merit_list, reverse=True)
#
#     unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
#                                 sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
#
#     # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
#     ranking = {}
#     for a in range(instance.parameters['M']):
#         sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
#         result = []
#         for i, (value, index) in enumerate(sorted_scores):
#             result.append(index)
#         ranking[a] = result
#
#     # G2G - This is relied upon to  build the other dictionaries as well.
#     # The below builds the score that a cadet assigns to each AFSC.
#     cadet_score_from_AFSC = {}
#     for c in range(instance.parameters['N']):
#         cadet_of_AFSC_score_list = []
#         for a in AFSC_scores.keys():
#             s = cadet_scores[c][a]
#             cadet_of_AFSC_score_list.append(s)
#             cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list
#
#     # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.
#
#     for c in range(instance.parameters['N']):
#         AFSC_list = np.where(instance.parameters['utility'][c] > 0)
#         for a in AFSC_list[0]:
#             if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
#                 cadet_score_from_AFSC[c][a] = -1
#             else:
#                 cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7
#
#     cadet_scoring = {}
#     for i in range(instance.parameters['N']):
#         c = 'c' + str(i)
#         cadet_scoring[c] = {}
#         for j in range(instance.parameters['M']):
#             a = instance.parameters['afsc_vector'][j]
#             # print(a)
#             score = cadet_score_from_AFSC[i][j]
#             if score > 0:  # added
#                 cadet_scoring[c][a] = score  # added
#
#     cadet_ranked_AFSC = {}
#     for c in range(instance.parameters['N']):
#         sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
#         result2 = []
#         for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
#             result2.append(index)
#         cadet_ranked_AFSC[c] = result2
#
#     # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
#     cadet_ranking = {}
#     for i in range(instance.parameters['N']):
#         afscs_from_cadets = []
#         c = 'c' + str(i)
#         for j in range(instance.parameters['M']):
#             AFSC = cadet_ranked_AFSC[i][j]
#             # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
#             a = instance.parameters['afsc_vector'][AFSC]
#             if a in cadet_scoring[c].keys():
#                 if cadet_scoring[c][a] > 0:
#                     afscs_from_cadets.append(a)
#
#             cadet_ranking[c] = afscs_from_cadets
#
#     # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.
#
#     afsc_scoring = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         afsc_scoring[a] = {}
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             cadet_score_of_afsc = AFSC_scores[j][i]
#             if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
#                 c = 'c' + str(i)
#                 afsc_scoring[a][c] = cadet_score_of_afsc
#
#     # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.
#
#     afsc_ranking = {}
#     for j in range(instance.parameters['M']):
#         cadets_from_afscs = []
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             cadet = ranking[j][i]
#             c = 'c' + str(cadet)
#             if c in afsc_scoring[a].keys():
#                 if afsc_scoring[a][c] > 0:
#                     cadets_from_afscs.append(c)
#
#             afsc_ranking[a] = cadets_from_afscs
#
#     # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
#     afscs = instance.parameters['afsc_vector']
#     pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
#                       34, 12, 2, 69, 50, 34]
#     afsc_capacities = dict(zip(afscs, pgl_capacities))
#     afsc_capacities
#
#     afsc_matches = {}
#     for j in range(instance.parameters['M']):
#         a = instance.parameters['afsc_vector'][j]
#         # print(a)
#         for i in range(instance.parameters['N']):
#             i = []
#         afsc_matches[a] = i
#
#     # Pseudo code for CAMP HR Algorithm
#     # import tqdm from tqdm_notebook as tqdm
#     # input data
#     # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
#     # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
#     # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
#     # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
#     # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
#     # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity
#
#     matching_data_structure_creation = time.time()
#     print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)
#
#     for c in cadet_ranking:  # MODIFICATION
#         cadet_ranking[c] = cadet_ranking[c][:6]  # Looking at only the cadet's first 6 choices
#
#     # initialize parameters
#     unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
#     cannot_match = []  # cadets with empty ranking lists
#     M = {}  # matches
#
#     iter_num = 0
#     iter_limit = 10000
#     next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
#     while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
#         iter_num += 1
#         if next_cadet:
#             c = unmatched_ordered_cadets[0]  # next cadet to match
#             unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
#
#         if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
#             cannot_match.append(c)  # add them to cannot match list
#             next_cadet = True
#             continue  # go to beginning of while loop
#         a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
#
#         if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
#             M[c] = a  # add match to M
#             afsc_matches[a].append(c)  # insert into AFSC a's matches
#             next_cadet = True  # move to next cadet
#             continue  # go to beginning of while loop
#
#         else:  # if at capacity
#             c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
#             for c_hat in afsc_matches[a]:
#                 if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
#                     c_ = c_hat
#             if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
#                 cadet_ranking[c].remove(a)  # remove a from c's ranking list
#                 next_cadet = False  # keep trying to match this cadet
#                 continue  # go to beginning of while loop
#
#             else:  # if c ranks higher than c_
#                 afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
#                 cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
#                 M.pop(c_)  # remove c_ from M
#                 M[c] = a  # add match to M
#                 next_cadet = False
#                 afsc_matches[a].append(c)  # insert into AFSC a's matches
#                 c = c_
#                 continue
#
#     print('\nM', M)
#     print("Total iterations:", iter_num)
#
#     Classical_HR_algorithm_time = time.time()
#     print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)
#
#     print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
#     print('Length of perm unmatched cadets:', len(cannot_match))
#
#     print("Total time:", time.time() - start_time)
#
#     # Trying to loop through those who could not be matched to match them with the career field that wants them.
#     print('# unmatched', len(cannot_match))
#     for c in cannot_match[:]:
#         for a in afsc_matches:
#             if c in afsc_ranking[a][0:10]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:100]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:500]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:1000]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#             elif c in afsc_ranking[a][0:1500]:
#                 if len(afsc_matches[a]) < afsc_capacities[a]:
#                     afsc_matches[a].append(c)
#                 elif len(afsc_matches[a]) >= afsc_capacities[a]:
#                     afsc_capacities[a] = afsc_capacities[a] + 1
#                     afsc_matches[a].append(c)
#                     M[c] = a
#                     cannot_match[:] = [c for c in cannot_match if c not in afsc_matches[
#                         a]]  # Must use list comprehension to build a copy of cannot_match in-place to not have the skipping effect that occurs when removing from a list that you are iterating over.
#                     if len(cannot_match) != 0:
#                         c = cannot_match[0]
#
#     print('Cannot Match still:', len(cannot_match))
#     # print('AFSC_Matches:', afsc_matches)
#
#     afsc_matches_len = []
#     for a in afsc_matches:
#         a = ('Matched:', len(afsc_matches[a]))
#         afsc_matches_len.append(a)
#
#     quota_min = []
#     for q in instance.parameters["quota_min"]:
#         q = ('min:', q)
#         quota_min.append(q)
#
#     max_quotas = []
#     for m in instance.parameters["quota_max"]:
#         m = ('max:', m)
#         max_quotas.append(m)
#
#     quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
#     print(quota_min_and_match_results)
#
#     print(M)
#     # return M
#
#     # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
#     new_M = {}
#     for c in M.keys():
#         i = int(c[1:])
#         new_M[i] = M[c]  # getting the index of cadet in M
#
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     ordered_M = dict(ordered_M)  # convert the ordering to a dictionary
#
#     all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
#     empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
#     solution_missing_cadets = ordered_M
#
#     # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
#     complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}
#
#     file_name = 'Real 2023 MacDonald Results.xlsx'
#     ordered_M = collections.OrderedDict(sorted(new_M.items()))
#     path = os.getcwd()
#
#     results_by_index = []
#     for i in complete_cadet_M_list.keys():
#         results_by_index.append(complete_cadet_M_list[i])
#     print(results_by_index)
#     # save current solutions
#     df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
#                                  engine='openpyxl', index_col=0)
#
#     # append new solution to current solutions
#     solution_name = 'func_hr_alg_1_1_2_2'
#     if solution_name in df_solutions.columns:
#         solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")
#
#     matching_list = results_by_index
#     df_new_solution = pd.DataFrame({solution_name: matching_list})
#     print(df_new_solution)
#
#     df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
#     print(df_solutions)
#
#     # save current and new solutions by deleting current sheet and creating a new one
#     ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
#     if 'Solutions' in ExcelWorkbook.sheetnames:
#         ExcelWorkbook.remove(ExcelWorkbook['Solutions'])
#
#     writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
#                             engine='openpyxl')  # Had to manually give directory
#     writer.book = ExcelWorkbook
#     df_solutions.to_excel(writer, sheet_name='Solutions')
#     # df_new_solution.to_excel(writer, sheet_name='Solutions')
#     writer.save()
#     writer.close()


def func_hr_alg_1_2(instance):
    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    best_unmatched_cadet = {}  # This is a new dict that has keys of previously unmatched cadets with a value of the AFSC that scored the cadet the highest.
    # unmatched_afsc_score = {} # This is a new dict that has the AFSC as the key with the afsc's associated value that it scored the cadet. This allow me to then pull the max value for the AFSC that scored the cadet the highest of all eligible AFSCS and return the key (which is the AFSC) associated with that value to add it as the value for the best_unmatched_cadet dict. Then I can update the matchings based on this dictionary and update AFSC_matches list.
    for c in cannot_match:
        afscs_score_of_cannot_match_cadets = []
        print('-------------------------------------', c)
        unmatched_afsc_score = {}
        for a in afsc_matches:
            # print(a,'*********************************')

            if c in afsc_ranking[a]:
                if len(afsc_matches[a]) >= afsc_capacities[a]:
                    if c in afsc_scoring[a]:
                        iter_num += 1
                        score_array = afsc_scoring[a][c]
                        print(a, c, score_array)
                        unmatched_afsc_score[a] = afsc_scoring[a][c]

                        # afsc_desiring_unmatched_cadet = max(unmatched_afsc_score, key=unmatched_afsc_score.get) #gets the first highest score in the dictionary
                        afsc_desiring_unmatched_cadet = [k for k, v in unmatched_afsc_score.items() if v == max(
                            unmatched_afsc_score.values())]  # gets all of the max values in the dict so I can see ties.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[
                    0]  # pulling index so I don't return a list, but instead a str

        if len(afsc_desiring_unmatched_cadet) > 1:  # if I have a tie, break tie with capacity
            iter_num += 1
            least_capacitated_afsc = {}  # Create dict of all the capacities of the AFSCs being considered with ties.
            for a in afsc_desiring_unmatched_cadet:
                least_capacitated_afsc[a] = len(afsc_matches[a]) / afsc_capacities[a] * 100  # Calculate the capacities
                print('least capacity:', least_capacitated_afsc)
                afsc_desiring_unmatched_cadet = [k for k, v in least_capacitated_afsc.items() if v == min(
                    least_capacitated_afsc.values())]  # Select the AFSC with the lowest overage in their capacity. This is the AFSC the cadet will match with.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[0]

                print("Is this happening for c867 or c928?")
                print(best_unmatched_cadet[c])
                print(afsc_desiring_unmatched_cadet[0])
                if len(afsc_desiring_unmatched_cadet) > 1:
                    print(afsc_desiring_unmatched_cadet)
                    afsc_desiring_unmatched_cadet = afsc_desiring_unmatched_cadet[
                        0]  # pulling the first AFSC in a tie where the score tied AND the AFSC's capacity was tied. Give the first AFSC we encounter the cadet in this case.
                    best_unmatched_cadet[
                        c] = afsc_desiring_unmatched_cadet  # since we already accessed the first AFSC,just update it.

    print('Previous Capacities', afsc_capacities)
    # Need to update afsc_matches and afsc_capacities
    for c in best_unmatched_cadet:
        for a in afsc_matches:
            if a in best_unmatched_cadet[c] == a in afsc_matches:
                afsc_matches[a].append(c)
                afsc_capacities[a] = afsc_capacities[a] + 1
    print('New AFSC Capacities:', afsc_capacities)

    # Need to update M[c]
    for c in best_unmatched_cadet:
        M[c] = best_unmatched_cadet[c]

    # Need to update cannot_match
    for c in best_unmatched_cadet:
        cannot_match.remove(c)

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_2'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_alg_1_3(instance):
    """ This algorithm calculates based on lower quotas set as the max capacities and all preferences. The algorithm
     them selects the AFSC's most preferred cadet from connot_match and matches them, incrementing the capacity by 1
     as required."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63,
                      34, 12, 2, 69, 50, 34]
    afsc_capacities = dict(zip(afscs, pgl_capacities))

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    best_unmatched_cadet = {}  # This is a new dict that has keys of previously unmatched cadets with a value of the AFSC that scored the cadet the highest.
    # unmatched_afsc_score = {} # This is a new dict that has the AFSC as the key with the afsc's associated value that it scored the cadet. This allow me to
    # then pull the max value for the AFSC that scored the cadet the highest of all eligible AFSCS and return the key (which is the AFSC) associated with that value to add it as the value for the best_unmatched_cadet dict. Then I can update the matchings based on this dictionary and update AFSC_matches list.
    for c in cannot_match:
        afscs_score_of_cannot_match_cadets = []
        print('-------------------------------------', c)
        unmatched_afsc_score = {}
        for a in afsc_matches:
            # print(a,'*********************************')

            if c in afsc_ranking[a]:
                if len(afsc_matches[a]) >= afsc_capacities[a]:
                    if c in afsc_scoring[a]:
                        iter_num += 1
                        score_array = afsc_scoring[a][c]
                        print(a, c, score_array)
                        unmatched_afsc_score[a] = afsc_scoring[a][c]

                        # afsc_desiring_unmatched_cadet = max(unmatched_afsc_score, key=unmatched_afsc_score.get) #gets the first highest score in the dictionary
                        afsc_desiring_unmatched_cadet = [k for k, v in unmatched_afsc_score.items() if v == max(
                            unmatched_afsc_score.values())]  # gets all of the max values in the dict so I can see ties.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[
                    0]  # pulling index so I don't return a list, but instead a str

        if len(afsc_desiring_unmatched_cadet) > 1:  # if I have a tie, break tie with capacity
            iter_num += 1
            least_capacitated_afsc = {}  # Create dict of all the capacities of the AFSCs being considered with ties.
            for a in afsc_desiring_unmatched_cadet:
                least_capacitated_afsc[a] = len(afsc_matches[a]) / afsc_capacities[a] * 100  # Calculate the capacities
                print('least capacity:', least_capacitated_afsc)
                afsc_desiring_unmatched_cadet = [k for k, v in least_capacitated_afsc.items() if v == min(
                    least_capacitated_afsc.values())]  # Select the AFSC with the lowest overage in their capacity. This is the AFSC the cadet will match with.
                best_unmatched_cadet[c] = afsc_desiring_unmatched_cadet[0]

                print("Is this happening for c867 or c928?")
                print(best_unmatched_cadet[c])
                print(afsc_desiring_unmatched_cadet[0])
                if len(afsc_desiring_unmatched_cadet) > 1:
                    print(afsc_desiring_unmatched_cadet)
                    afsc_desiring_unmatched_cadet = afsc_desiring_unmatched_cadet[
                        0]  # pulling the first AFSC in a tie where the score tied AND the AFSC's capacity was tied. Give the first AFSC we encounter the cadet in this case.
                    best_unmatched_cadet[
                        c] = afsc_desiring_unmatched_cadet  # since we already accessed the first AFSC,just update it.

    print('Previous Capacities', afsc_capacities)
    # Need to update afsc_matches and afsc_capacities
    for c in best_unmatched_cadet:
        for a in afsc_matches:
            if a in best_unmatched_cadet[c] == a in afsc_matches:
                afsc_matches[a].append(c)
                afsc_capacities[a] = afsc_capacities[a] + 1
    print('New AFSC Capacities:', afsc_capacities)

    # Need to update M[c]
    for c in best_unmatched_cadet:
        M[c] = best_unmatched_cadet[c]

    # Need to update cannot_match
    for c in best_unmatched_cadet:
        cannot_match.remove(c)

    print('Cannot Match still:', len(cannot_match))
    # print('AFSC_Matches:', afsc_matches)

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    # return M

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_1_3'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_lq_alg_2_0(instance):
    """This function conducts lower quotas with AFSC capacity set at 500 for all AFSCs to allow all cadets to obtain
    their #1 choice. This cycles through the highest quota AFSC and tries to give up it's lowest ranked cadet to the
    AFSC farthest from meeting its lower quota."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    # start_time = time.time()
    # import os
    #
    # # Obtain initial working directory
    # dir_path = os.getcwd() + '/'
    # print('initial working directory:', dir_path)
    #
    # # Get main afccp folder path
    # index = dir_path.find('afccp')
    # dir_path = dir_path[:index + 6]
    #
    # # Update working directory
    # os.chdir(dir_path)
    # print('updated working directory:', dir_path)
    #
    # # Import the problem class
    # from afccp.core.problem_class import CadetCareerProblem
    #
    # instance = CadetCareerProblem("2023", printing=True)  # Must be one string, no spaces
    #
    # instance_creation_time = time.time()
    # print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # # Actual Max Capacities for AFSCs
    # afscs = instance.parameters['afsc_vector']
    # capacities = instance.parameters["quota_max"]
    # afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the really high overall capacity for each AFSC to allow all members to get first choice.
    afscs = instance.parameters['afsc_vector']
    pgl_capacities = [500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500,
                      500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500]
    afsc_capacities = dict(zip(afscs, pgl_capacities))
    print('afsc_capacities:', afsc_capacities)

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    # print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##

    print('##-----------------------------------------------------------##' * 3)

    # Create dictionary of lower quotas for each AFSC
    afscs = instance.parameters['afsc_vector']
    lower_quotas = instance.parameters['quota_min']
    afsc_lower_quotas = dict(zip(afscs, lower_quotas))
    afsc_lower_quotas

    moved_cadets = []
    immobile_cadets = []  # cadets that can't be moved
    a_h_q_afscs_w_no_cadets_to_donate = []
    iter_num = 0
    total_iterations = 0
    # while afsc lower quotas are not met or exceeded:
    distance_from_quota = {}
    lower_quota_ratios = [20]
    # next_cadet = False  # Set to look at the next cadet.
    # while distance_from_quota.values() < 100:
    # (v < 100 for v in distance_from_quota.values()) or iter_num <= iter_limit:

    iter_limit = 200
    # while lower_quota_ratios < [100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0]: #Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
    while (lower_quota_ratios[
               0] < 99.0):  # iter_num < 1000 # Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
        # next_iteration = False # added
        next_cadet = False
        iter_num += 1

        # Select AFSC most under quota [a_l_q]
        distance_from_quota = {}
        for a in afsc_matches:
            if len(afsc_matches[a]) / afsc_lower_quotas[a] * 100 >= 100:  # if they are more than 100% of lower quota
                if (len(afsc_matches[a]) - 1) / afsc_lower_quotas[
                    a] * 100 < 100:  # If taking one cadet will put them below lower quota
                    a_h_q_afscs_w_no_cadets_to_donate.append(a)  # Then remove this AFSC from consideration
            # if all cadets in afsc_matches[a_h_q] contains all of the cadets in immobile cadets, continue.
            if a not in a_h_q_afscs_w_no_cadets_to_donate:
                distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[
                    a] * 100  # Calculate the percent from lower quota
        # lower_quota_ratios = []
        lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        # lower_quota_ratios = list(min(list(distance_from_quota.values())))
        print(lower_quota_ratios, list(distance_from_quota.values()))
        a_l_q = [k for k, v in distance_from_quota.items() if v == min(
            distance_from_quota.values()) and v < 100]  # v<100 prevents continued filling after all afscs have met lower quota
        a_l_q = a_l_q[0]

        # Select AFSC most over quota [a_h_q]
        a_h_q = [k for k, v in distance_from_quota.items() if v == max(
            distance_from_quota.values()) > 100]  # v>100 ensures we only pull from AFSCs that are over their quota.
        # a_h_q = [k for k, v in distance_from_quota.items() if v == max(distance_from_quota.values()) and v > 100]
        if not a_h_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
            print("I'm out of AFSCs over 100 to choose from: ", distance_from_quota)
            break  # Return to start to see if all AFSC lower quotas are >=100.
        a_h_q = a_h_q[0]  # Just take the first item in the list regardless

        # Select lowest scored cadet from most over-quota AFSC.
        a_h_cadet_score_from_afsc = []
        a_h_cadet_list = []
        lowest_score_cadet_for_afsc = {}

        # Building dict of all cadets in highest from quota
        for c in afsc_matches[a_h_q]:
            if c not in immobile_cadets:
                a_h_cadet_list.append(c)  # create a list of cadets matched to a_h_q
                a_h_cadet_score_from_afsc.append(afsc_scoring[a_h_q][c])  # create a list o
            # if next_cadet == True:
            #     distance_from_quota = distance_from_quota
            #     break
            # if next_cadet == False, we are still trying to match this cadet to the lowest quota AFSC. Keep trying.
        while len(a_h_cadet_score_from_afsc) > 0:
            lowest_score_cadet_for_afsc = dict(zip(a_h_cadet_list,
                                                   a_h_cadet_score_from_afsc))  # zip the two together to make a dictionary to pull the minimum value from.
            c = [k for k, v in lowest_score_cadet_for_afsc.items() if v == min(lowest_score_cadet_for_afsc.values())][0]
            if next_cadet == True:  # if next_iteration == True
                break

            while next_cadet == False:  # Keep trying to match cadet

                if c in afsc_scoring[
                    a_l_q]:  # If the cadet is scored by the lowest quota ratio AFSC (i.e. the cadet is eligible for the AFSC)
                    # Append this cadet to lowest quota AFSC.
                    afsc_matches[a_l_q].append(c)
                    # Remove this cadet from the highest quota AFSC.
                    afsc_matches[a_h_q].remove(c)
                    # Update M.
                    M[c] = a_l_q
                    # Track all cadets who have moved in a list
                    moved_cadets.append(c)
                    next_cadet = True  # We've finished getting this cadet match. Let's move on to the next cadet. Need to break out to the very first while loop.
                    # Could make next iteration
                    break  # first break to get back to the "for c in afsc_matches" loop. Then we'll break again to get to first while loop.

                # Select the next afsc the cadet is eligible for and remove the AFSC they were not eligible for.

                else:  # else: cadet is not in the lowest afsc (a_l_q). Check the next lowest AFSC.
                    # Select the next lowest quota AFSC and check for cadet membership in the above if statement.
                    if a_l_q in distance_from_quota:
                        del distance_from_quota[a_l_q]
                        a_l_q = [k for k, v in distance_from_quota.items() if
                                 v == min(distance_from_quota.values()) and v < 100]
                        if not a_l_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
                            next_cadet = True
                            del lowest_score_cadet_for_afsc[c]
                            immobile_cadets.append(c)
                            break
                        a_l_q = a_l_q[0]
                        next_cadet = False
                        continue

        if len(a_h_cadet_score_from_afsc) == 0:
            a_h_q_afscs_w_no_cadets_to_donate.append(a_h_q)
            print('No More Cadets in ', a_h_q)
            print(iter_num)
        # lower_quota_ratios = [ratio for afsc, ratio in distance_from_quota.items()]
        # lower_quota_ratios = []
        # lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        print(iter_num)
        continue

    total_iterations = iter_num
    print('Total Iterations:', iter_num)
    print(moved_cadets)
    for a in afsc_matches:
        distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[a] * 100
    print(distance_from_quota)
    print("Total time:", time.time() - start_time)
    print('Done')

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))

    afsc_avg_merit = {}

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC')
    plt.ylabel('Average Merit of AFSC')
    plt.title('Average Merit of Large AFSCs')
    plt.ylim(.2, 0.8)
    plt.show()
    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_2_0'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_lq_alg_2_1(instance):
    """This function conducts lower quotas with AFSC capacity set at true max capacities for all AFSCs to distribute
     cadets across AFSCs before cycling through AFSC that did not meet their lower quota. The function then cycles
     through the highest quota AFSC and tries to give up it's lowest ranked cadet to the AFSC farthest from meeting its
      lower quota. Best result so far!"""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    start_time = time.time()
    import os

    # # Obtain initial working directory
    # dir_path = os.getcwd() + '/'
    # print('initial working directory:', dir_path)
    #
    # # Get main afccp folder path
    # index = dir_path.find('afccp')
    # dir_path = dir_path[:index + 6]
    #
    # # Update working directory
    # os.chdir(dir_path)
    # print('updated working directory:', dir_path)
    #
    # # Import the problem class
    # from afccp.core.problem_class import CadetCareerProblem
    #
    # instance = CadetCareerProblem("2023", printing=True)  # Must be one string, no spaces
    #
    # instance_creation_time = time.time()
    # print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # # #Modified capacities with the really high overall capacity for each AFSC to allow all members to get first choice.
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500,
    #                   500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # print('afsc_capacities:', afsc_capacities)

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    # print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##

    print('##-----------------------------------------------------------##' * 3)

    # Create dictionary of lower quotas for each AFSC
    afscs = instance.parameters['afsc_vector']
    lower_quotas = instance.parameters['quota_min']
    afsc_lower_quotas = dict(zip(afscs, lower_quotas))
    afsc_lower_quotas

    moved_cadets = []
    immobile_cadets = []  # cadets that can't be moved
    a_h_q_afscs_w_no_cadets_to_donate = []
    iter_num = 0
    total_iterations = 0
    # while afsc lower quotas are not met or exceeded:
    distance_from_quota = {}
    lower_quota_ratios = [20]
    # next_cadet = False  # Set to look at the next cadet.
    # while distance_from_quota.values() < 100:
    # (v < 100 for v in distance_from_quota.values()) or iter_num <= iter_limit:

    iter_limit = 200
    # while lower_quota_ratios < [100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0]: #Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
    while (lower_quota_ratios[
               0] < 99.0):  # iter_num < 1000 # Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
        # next_iteration = False # added
        next_cadet = False
        iter_num += 1

        # Select AFSC most under quota [a_l_q]
        distance_from_quota = {}
        for a in afsc_matches:
            if len(afsc_matches[a]) / afsc_lower_quotas[a] * 100 >= 100:  # if they are more than 100% of lower quota
                if (len(afsc_matches[a]) - 1) / afsc_lower_quotas[
                    a] * 100 < 100:  # If taking one cadet will put them below lower quota
                    a_h_q_afscs_w_no_cadets_to_donate.append(a)  # Then remove this AFSC from consideration
            # if all cadets in afsc_matches[a_h_q] contains all of the cadets in immobile cadets, continue.
            if a not in a_h_q_afscs_w_no_cadets_to_donate:
                distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[
                    a] * 100  # Calculate the percent from lower quota
        # lower_quota_ratios = []
        lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        # lower_quota_ratios = list(min(list(distance_from_quota.values())))
        print(lower_quota_ratios, list(distance_from_quota.values()))
        a_l_q = [k for k, v in distance_from_quota.items() if v == min(
            distance_from_quota.values()) and v < 100]  # v<100 prevents continued filling after all afscs have met lower quota
        a_l_q = a_l_q[0]

        # Select AFSC most over quota [a_h_q]
        a_h_q = [k for k, v in distance_from_quota.items() if v == max(
            distance_from_quota.values()) > 100]  # v>100 ensures we only pull from AFSCs that are over their quota.
        # a_h_q = [k for k, v in distance_from_quota.items() if v == max(distance_from_quota.values()) and v > 100]
        if not a_h_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
            print("I'm out of AFSCs over 100 to choose from: ", distance_from_quota)
            break  # Return to start to see if all AFSC lower quotas are >=100.
        a_h_q = a_h_q[0]  # Just take the first item in the list regardless

        # Select lowest scored cadet from most over-quota AFSC.
        a_h_cadet_score_from_afsc = []
        a_h_cadet_list = []
        lowest_score_cadet_for_afsc = {}

        # Building dict of all cadets in highest from quota
        for c in afsc_matches[a_h_q]:
            if c not in immobile_cadets:
                a_h_cadet_list.append(c)  # create a list of cadets matched to a_h_q
                a_h_cadet_score_from_afsc.append(afsc_scoring[a_h_q][c])  # create a list o
            # if next_cadet == True:
            #     distance_from_quota = distance_from_quota
            #     break
            # if next_cadet == False, we are still trying to match this cadet to the lowest quota AFSC. Keep trying.
        while len(a_h_cadet_score_from_afsc) > 0:
            lowest_score_cadet_for_afsc = dict(zip(a_h_cadet_list,
                                                   a_h_cadet_score_from_afsc))  # zip the two together to make a dictionary to pull the minimum value from.
            c = [k for k, v in lowest_score_cadet_for_afsc.items() if v == min(lowest_score_cadet_for_afsc.values())][0]
            if next_cadet == True:  # if next_iteration == True
                break

            while next_cadet == False:  # Keep trying to match cadet

                if c in afsc_scoring[
                    a_l_q]:  # If the cadet is scored by the lowest quota ratio AFSC (i.e. the cadet is eligible for the AFSC)
                    # Append this cadet to lowest quota AFSC.
                    afsc_matches[a_l_q].append(c)
                    # Remove this cadet from the highest quota AFSC.
                    afsc_matches[a_h_q].remove(c)
                    # Update M.
                    M[c] = a_l_q
                    # Track all cadets who have moved in a list
                    moved_cadets.append(c)
                    next_cadet = True  # We've finished getting this cadet match. Let's move on to the next cadet. Need to break out to the very first while loop.
                    # Could make next iteration
                    break  # first break to get back to the "for c in afsc_matches" loop. Then we'll break again to get to first while loop.

                # Select the next afsc the cadet is eligible for and remove the AFSC they were not eligible for.

                else:  # else: cadet is not in the lowest afsc (a_l_q). Check the next lowest AFSC.
                    # Select the next lowest quota AFSC and check for cadet membership in the above if statement.
                    if a_l_q in distance_from_quota:
                        del distance_from_quota[a_l_q]
                        a_l_q = [k for k, v in distance_from_quota.items() if
                                 v == min(distance_from_quota.values()) and v < 100]
                        if not a_l_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
                            next_cadet = True
                            del lowest_score_cadet_for_afsc[c]
                            immobile_cadets.append(c)
                            break
                        a_l_q = a_l_q[0]
                        next_cadet = False
                        continue

        if len(a_h_cadet_score_from_afsc) == 0:
            a_h_q_afscs_w_no_cadets_to_donate.append(a_h_q)
            print('No More Cadets in ', a_h_q)
            print(iter_num)
        # lower_quota_ratios = [ratio for afsc, ratio in distance_from_quota.items()]
        # lower_quota_ratios = []
        # lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        print(iter_num)
        continue

    total_iterations = iter_num
    print('Total Iterations:', iter_num)
    print(moved_cadets)
    for a in afsc_matches:
        distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[a] * 100
    print(distance_from_quota)
    print("Total time:", time.time() - start_time)
    print('Done')

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    unchanging_lower_quotas
    afsc_avg_merit = {}

    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams
    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    plt.title('Average Merit of Large AFSCs', size=16)
    plt.ylim(.2, 0.8)
    plt.show()

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'func_hr_alg_2_1'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def func_hr_lq_alg_2_1_1_w_preprocess(instance):
    """ This function runs the CAMP HR with Lower Quotas at max capacity after pre-processing the required cadets for
    AFSC 62EXC and 62EXE. This algorithm fills all lower quotas when used with the 2023 cadet data."""

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections
    import matplotlib.pyplot as plt

    # start_time = time.time()
    # import os
    #
    # # Obtain initial working directory
    # dir_path = os.getcwd() + '/'
    # print('initial working directory:', dir_path)
    #
    # # Get main afccp folder path
    # index = dir_path.find('afccp')
    # dir_path = dir_path[:index + 6]
    #
    # # Update working directory
    # os.chdir(dir_path)
    # print('updated working directory:', dir_path)
    #
    # # Import the problem class
    # from afccp.core.problem_class import CadetCareerProblem
    #
    # instance = CadetCareerProblem("2023", printing=True)  # Must be one string, no spaces
    #
    # instance_creation_time = time.time()
    # print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # # #Modified capacities with the really high overall capacity for each AFSC to allow all members to get first choice.
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500,
    #                   500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # print('afsc_capacities:', afsc_capacities)

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    # Adding the new pre-processing elements

    cadets_f_62EXC = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXC':
                if cadet_scoring[cadet]['62EXC'] >= 7.0:
                    cadets_f_62EXC[cadet] = cadet_scoring[cadet]['62EXC']

    ordered_cadets_f_62EXC = {k: v for k, v in sorted(cadets_f_62EXC.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXC)

    import itertools
    ordered_cadets_f_62EXC_20c = dict(
        itertools.islice(ordered_cadets_f_62EXC.items(), instance.parameters['quota_min'][24]))

    cadets_f_62EXE = {}
    for cadet, afsc_w_scores in cadet_scoring.items():
        for key in afsc_w_scores:
            if key == '62EXE':
                # print(cadet,':', cadet_scoring[cadet]['62EXE'])
                cadets_f_62EXE[cadet] = cadet_scoring[cadet]['62EXE']

    ordered_cadets_f_62EXE = {k: v for k, v in sorted(cadets_f_62EXE.items(), key=lambda item: item[1], reverse=True)}
    # print(ordered_cadets_f_62EXE)
    # print(len(ordered_cadets_f_62EXE))

    # Need to take only 53 because we have to leave 3 cadets for 32EXE
    ordered_cadets_f_62EXE_53c = dict(
        itertools.islice(ordered_cadets_f_62EXE.items(), instance.parameters['quota_min'][25]))

    filtered_cadets_f_62EXC_20c = {}

    for c in ordered_cadets_f_62EXC:
        if c not in ordered_cadets_f_62EXE_53c:
            filtered_cadets_f_62EXC_20c[c] = cadet_scoring[c]['62EXC']

    # force it to just take the 20 cadets you want.
    filtered_cadets_f_62EXC_20c = dict(
        itertools.islice(filtered_cadets_f_62EXC_20c.items(), instance.parameters['quota_min'][24]))

    # filter out the cadets who are duplicates in the unmatched list so that they aren't double counted.
    pre_processed_unmatched_ordered_cadets = []
    for c in unmatched_ordered_cadets:
        if c not in filtered_cadets_f_62EXC_20c.keys() and c not in ordered_cadets_f_62EXE_53c.keys():
            pre_processed_unmatched_ordered_cadets.append(c)
    unmatched_ordered_cadets = pre_processed_unmatched_ordered_cadets
    # print(len(unmatched_ordered_cadets))

    # Match cadets to AFSC 62EXC - update afsc_matches and M
    EXC_cadets = []
    for c in filtered_cadets_f_62EXC_20c.keys():
        EXC_cadets.append(c)
        afsc_matches['62EXC'] = EXC_cadets
        M[c] = '62EXC'

    # Match cadets to AFSC 62EXE - update afsc_matches and M
    EXE_cadets = []
    for c in ordered_cadets_f_62EXE_53c.keys():
        EXE_cadets.append(c)
        afsc_matches['62EXE'] = EXE_cadets
        M[c] = '62EXE'

    ###### END pre-processing code

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches

    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    while len(unmatched_ordered_cadets) > 0 or next_cadet == False and iter_num <= iter_limit:
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list

        if len(cadet_ranking[c]) == 0 and c not in cannot_match:  # if no more AFSCs in ranking list
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c

        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop

        else:  # if at capacity
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop

            else:  # if c ranks higher than c_
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                M.pop(c_)  # remove c_ from M
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                continue

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    # print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##
    ##-----------------------------------------------------------##
    ##__________________________________________________________##

    print('##-----------------------------------------------------------##' * 3)

    # Create dictionary of lower quotas for each AFSC
    afscs = instance.parameters['afsc_vector']
    lower_quotas = instance.parameters['quota_min']
    afsc_lower_quotas = dict(zip(afscs, lower_quotas))
    afsc_lower_quotas

    moved_cadets = []
    immobile_cadets = []  # cadets that can't be moved
    a_h_q_afscs_w_no_cadets_to_donate = []
    iter_num = 0
    total_iterations = 0
    # while afsc lower quotas are not met or exceeded:
    distance_from_quota = {}
    lower_quota_ratios = [20]
    # next_cadet = False  # Set to look at the next cadet.
    # while distance_from_quota.values() < 100:
    # (v < 100 for v in distance_from_quota.values()) or iter_num <= iter_limit:

    iter_limit = 200
    # while lower_quota_ratios < [100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0, 100.0]: #Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
    while (lower_quota_ratios[
               0] < 99.0):  # iter_num < 1000 # Make a list of lower quota ratios and pull the values from the list to update the lower quota ratios list after each run.
        # next_iteration = False # added
        next_cadet = False
        iter_num += 1

        # Select AFSC most under quota [a_l_q]
        distance_from_quota = {}
        for a in afsc_matches:
            if len(afsc_matches[a]) / afsc_lower_quotas[a] * 100 >= 100:  # if they are more than 100% of lower quota
                if (len(afsc_matches[a]) - 1) / afsc_lower_quotas[
                    a] * 100 < 100:  # If taking one cadet will put them below lower quota
                    a_h_q_afscs_w_no_cadets_to_donate.append(a)  # Then remove this AFSC from consideration
            # if all cadets in afsc_matches[a_h_q] contains all of the cadets in immobile cadets, continue.
            if a not in a_h_q_afscs_w_no_cadets_to_donate:
                distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[
                    a] * 100  # Calculate the percent from lower quota
        # lower_quota_ratios = []
        lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        # lower_quota_ratios = list(min(list(distance_from_quota.values())))
        print(lower_quota_ratios, list(distance_from_quota.values()))
        a_l_q = [k for k, v in distance_from_quota.items() if v == min(
            distance_from_quota.values()) and v < 100]  # v<100 prevents continued filling after all afscs have met lower quota
        if not a_l_q:
            break
        else:
            a_l_q = a_l_q[0]

        # Select AFSC most over quota [a_h_q]
        a_h_q = [k for k, v in distance_from_quota.items() if v == max(
            distance_from_quota.values()) > 100]  # v>100 ensures we only pull from AFSCs that are over their quota.
        # a_h_q = [k for k, v in distance_from_quota.items() if v == max(distance_from_quota.values()) and v > 100]
        if not a_h_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
            print("I'm out of AFSCs over 100 to choose from: ", distance_from_quota)
            break  # Return to start to see if all AFSC lower quotas are >=100.
        a_h_q = a_h_q[0]  # Just take the first item in the list regardless

        # Select lowest scored cadet from most over-quota AFSC.
        a_h_cadet_score_from_afsc = []
        a_h_cadet_list = []
        lowest_score_cadet_for_afsc = {}

        # Building dict of all cadets in highest from quota
        for c in afsc_matches[a_h_q]:
            if c not in immobile_cadets:
                a_h_cadet_list.append(c)  # create a list of cadets matched to a_h_q
                a_h_cadet_score_from_afsc.append(afsc_scoring[a_h_q][c])  # create a list o
            # if next_cadet == True:
            #     distance_from_quota = distance_from_quota
            #     break
            # if next_cadet == False, we are still trying to match this cadet to the lowest quota AFSC. Keep trying.
        while len(a_h_cadet_score_from_afsc) > 0:
            lowest_score_cadet_for_afsc = dict(zip(a_h_cadet_list,
                                                   a_h_cadet_score_from_afsc))  # zip the two together to make a dictionary to pull the minimum value from.
            c = [k for k, v in lowest_score_cadet_for_afsc.items() if v == min(lowest_score_cadet_for_afsc.values())][0]
            if next_cadet == True:  # if next_iteration == True
                break

            while next_cadet == False:  # Keep trying to match cadet

                if c in afsc_scoring[
                    a_l_q]:  # If the cadet is scored by the lowest quota ratio AFSC (i.e. the cadet is eligible for the AFSC)
                    # Append this cadet to lowest quota AFSC.
                    afsc_matches[a_l_q].append(c)
                    # Remove this cadet from the highest quota AFSC.
                    afsc_matches[a_h_q].remove(c)
                    # Update M.
                    M[c] = a_l_q
                    # Track all cadets who have moved in a list
                    moved_cadets.append(c)
                    next_cadet = True  # We've finished getting this cadet match. Let's move on to the next cadet. Need to break out to the very first while loop.
                    # Could make next iteration
                    break  # first break to get back to the "for c in afsc_matches" loop. Then we'll break again to get to first while loop.

                # Select the next afsc the cadet is eligible for and remove the AFSC they were not eligible for.

                else:  # else: cadet is not in the lowest afsc (a_l_q). Check the next lowest AFSC.
                    # Select the next lowest quota AFSC and check for cadet membership in the above if statement.
                    if a_l_q in distance_from_quota:
                        del distance_from_quota[a_l_q]
                        a_l_q = [k for k, v in distance_from_quota.items() if
                                 v == min(distance_from_quota.values()) and v < 100]
                        if not a_l_q:  # This is the pythonic way to say, if the a_l_q list is empty... do...
                            next_cadet = True
                            del lowest_score_cadet_for_afsc[c]
                            immobile_cadets.append(c)
                            break
                        a_l_q = a_l_q[0]
                        next_cadet = False
                        continue

        if len(a_h_cadet_score_from_afsc) == 0:
            a_h_q_afscs_w_no_cadets_to_donate.append(a_h_q)
            print('No More Cadets in ', a_h_q)
            print(iter_num)
        # lower_quota_ratios = [ratio for afsc, ratio in distance_from_quota.items()]
        # lower_quota_ratios = []
        # lower_quota_ratios = [v for k, v in distance_from_quota.items() if v == min(distance_from_quota.values())]
        print(iter_num)
        continue

    total_iterations = iter_num
    print('Total Iterations:', iter_num)
    print(moved_cadets)
    for a in afsc_matches:
        distance_from_quota[a] = len(afsc_matches[a]) / afsc_lower_quotas[a] * 100
    print(distance_from_quota)
    print("Total time:", time.time() - start_time)
    print('Done')

    afscs = instance.parameters['afsc_vector']
    afsc_lower_quotas = instance.parameters["quota_min"]
    unchanging_lower_quotas = dict(zip(afscs, afsc_lower_quotas))
    unchanging_lower_quotas
    afsc_avg_merit = {}

    import matplotlib.font_manager as font_manager
    from matplotlib import rcParams
    rcParams['axes.spines.top'] = False
    rcParams['axes.spines.right'] = False

    font_dir = ['/Users/ianmacdonald/Desktop/AFIT/AFIT - 4th Quarter 2022/Open_Sans']
    for font in font_manager.findSystemFonts(font_dir):
        font_manager.fontManager.addfont(font)

    rcParams['font.family'] = 'Open Sans'

    for a in afsc_matches:
        if unchanging_lower_quotas[a] > 40:
            list_of_cadet_merit_in_matched_afsc = []
            for c in afsc_matches[a]:
                c = int(c[1:])
                list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        else:
            continue
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / len(afsc_matches[a])
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    print('Average Merit per AFSC:', afsc_avg_merit)
    # avg_merit_of_all_afscs = sum(afsc_avg_merit.values())/len(afsc_avg_merit)
    # print('Average Merit of All AFSCs:', avg_merit_of_all_afscs)
    afsc_merit_list = [m for m in afsc_avg_merit.values()]
    mean_merit = np.mean(afsc_merit_list)
    print("Mean Merit of Solution for Large AFSCs:", mean_merit)
    median_merit = np.median(afsc_merit_list)
    print('Median Merit of Solution for Large AFSCs:', median_merit)

    fig = plt.figure(figsize=(15, 9))
    plt.stem(afsc_avg_merit.keys(), afsc_avg_merit.values())
    plt.axhline(y=0.35, color='black', linestyle='--')
    plt.axhline(y=0.65, color='black', linestyle='--')
    plt.xlabel('Large AFSC', size=14)
    plt.ylabel('Average Merit of AFSC', size=14)
    plt.title('Average Merit of Large AFSCs', size=16)
    plt.ylim(.2, 0.8)
    plt.show()

    # Get all cadets in a match and pair them with their index number for their AFSC. If they were unmatched, assign them AFSC 99N.
    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of cadet in M

    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    ordered_M = dict(ordered_M)  # convert the ordering to a dictionary

    all_cadets = list(range(0, len(range(instance.parameters['N']))))  #
    empty_cadet_dict = dict(zip(all_cadets, [None] * len(all_cadets)))
    solution_missing_cadets = ordered_M

    # Assign any unmatched cadets to AFSC 99N and maintain cadet ordering.
    complete_cadet_M_list = {k: solution_missing_cadets.get(k, '99N') for k in empty_cadet_dict}

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in complete_cadet_M_list.keys():
        results_by_index.append(complete_cadet_M_list[i])
    print(results_by_index)
    # save current solutions
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'Pre-Processing with Max Capacitated HR with LQs'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)

    df_solutions = pd.concat([df_solutions, df_new_solution], axis=1, join="outer")
    print(df_solutions)

    # save current and new solutions by deleting current sheet and creating a new one
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


def old_hr_algorithm_version(): # dont_use_this
    print_details = False
    use_test_cadets = False
    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    # print(unmatched_ordered_cadets)
    test_cadets = ['c144', 'c82', 'c951', 'c1182', 'c460', 'c1183']
    if use_test_cadets:
        unmatched_ordered_cadets = test_cadets
    while len(unmatched_ordered_cadets) > 0 and iter_num <= iter_limit:  # while there are still cadets to match
        if print_details: print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
        if print_details: print('Length of perm unmatched cadets:', len(cannot_match))
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            if print_details: print('\nnext cadet', c)  # Can comment out
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
            # print(unmatched_ordered_cadets) # Can comment out
        if len(cadet_ranking[c]) == 0:  # if no more AFSCs in ranking list
            if print_details: print('Cannot match', c)
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
        print('Trying to match', c, a)
        if c == 'c1464':
            print('\n\n\nXXXXX ', c,
                  'XXXXXXXX-----------------------------------------------------------------------------------------------------------------------------\n\n\n')
        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            # if print_details: print("M", M)
            if print_details: print('Matched', c, 'and', a)
            iter_num += 1  ##$%#&#^$%^$%#$%&#&$#%^@$%^#$%@#$^@^$%^#$%^$%^#$%^#$%^#$%^#$%^#$%^#$%^$%^#$% added this. Delete.
            print('cadet', c, 'and afsc', a)
            print(iter_num)
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            # afsc_matches = {13H: ["c1", "c2", "c3",...]}
            # sort afsc_matches[a] list based on afsc_ranking list of cadets
            # afsc_scoring[a][c] = s
            # print(afsc_ranking[a])
            #      afsc_matches[a] = [afsc_ranking[b] for _, b in zip(afsc_matches[a], afsc_ranking[a])]
            # print('matches 1:', afsc_matches) # Can comment out
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop
        else:  # if at capacity
            if print_details: print(a, "at capacity")
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            if print_details: print("AFSC_matches", afsc_matches[a])
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                if print_details: print(c_, 'higher than', c)
                # if print_details: print(cadet_ranking[c])
                if print_details: print("Line 61: Removing", a, 'from', c, 's list')
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                # if print_details: print(a, 'removed from', c,'s list')
                # M[c_] = a #added this define M[c_] so I could remove from it but c_ is already matched so this may be a problem.
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop
            else:  # if c ranks higher than c_
                if print_details: print(c_, 'lower than', c)
                if print_details: print('Line 69: Removing', c_, 'from', a)
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                if c_ == 'c144': print(cadet_ranking[c_],
                                       '-----------------******************--------------------------------')
                if print_details: print("Removing", a, 'from', c_, 's list')
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                # if print_details: print("Cadet AFSC List:", cadet_ranking[c_])
                M.pop(c_, print(c_))  # remove c_ from M - Gives error that can't use pop to remove a string.
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                if len(unmatched_ordered_cadets) == 0 and next_cadet == False:  # added to get the last cadet added as unmatched
                    cannot_match.append(c)  # added to get the last cadet added as unmatched
                continue  # go to beginning of while loop


def func_blocking_pairs_check(afsc_matches, cadet_ranking, afsc_scoring, cadet_scoring):
    """This code finds the blocking pairs in a match. """
    # Create dictionary of the lowest scored cadet in each AFSC.

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        cadet_score_list = []
        for c in afsc_matches[a]:
            cadet_score_list.append(afsc_scoring[a][c])
        lowest_score = min(cadet_score_list)
        lowest_scoring_cadet_in_afsc[a] = lowest_score

    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.
    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                for a in afsc_matches:
                    if a != current_afsc:
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))

    return blocking_pairs_dictionary

    # blocking_pairs_dictionary = func_blocking_pairs_check(afsc_matches, cadet_ranking, afsc_scoring, cadet_scoring)

    # def func_blocking_pairs_check(instance):
    """This code finds the blocking pairs in a match. """

    # Create dictionary of the lowest scored cadet in each AFSC.

    lowest_scoring_cadet_in_afsc = {}

    for a in afsc_matches:
        # print(a)
        cadet_score_list = []
        for c in afsc_matches[a]:
            # print(c)
            cadet_score_list.append(afsc_scoring[a][c])
            # print(cadet_score_list)
        lowest_score = min(cadet_score_list)
        lowest_scoring_cadet_in_afsc[a] = lowest_score
        # print(lowest_score)
    # print(lowest_scoring_cadet_in_afsc)

    # --------------------------------------------------------------
    # Now let's find the AFSCs that prefer some cadet in their pref list to their lowest scoring matched cadet.

    AFSCs_cadets_desire_morethan_current_afsc = {}
    cadets_other_AFSCs_desire_more = {}
    for c in cadet_ranking.keys():
        if c not in cannot_match:
            # print('cadet', c)
            for a in afsc_matches:
                if c not in afsc_matches[a]:
                    current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                    if a != current_afsc:
                        if c not in afsc_matches[a] and c in afsc_matches[current_afsc]:
                            # print('cadet',c, 'not in afsc',a)
                            if a in cadet_scoring[c]:  # and current_afsc in cadet_scoring[c]:
                                if cadet_scoring[c][a] > cadet_scoring[c][current_afsc]:
                                    # print('cadet',c,"'s score for afsc",a,':', cadet_scoring[c][a])
                                    # print('cadet',c,"'s score for afsc",current_afsc,':', cadet_scoring[c][current_afsc])
                                    # # print('looking at cadet',c)
                                    # print('-----------')
                                    AFSCs_cadets_desire_morethan_current_afsc[c] = a
            # print(AFSCs_cadets_desire_morethan_current_afsc)
            # print('--------------------------------------------------------------------')

    for c in cadet_ranking.keys():
        if c not in cannot_match:
            # afscs_that_desire_cadets_they_are_not_matched_with = []
            # print('cadet', c)
            for a in afsc_matches:
                current_afsc = [a for a in afsc_matches if c in afsc_matches[a]][0]
                # print('current1:',current_afsc)
                # print('afsc1', a)
                for a in afsc_matches:
                    if a != current_afsc:

                        # print('current2:',current_afsc)
                        # print('afsc2', a)
                        if c not in afsc_matches[a]:
                            if c in afsc_scoring[a]:
                                afscs_that_desire_cadets_they_are_not_matched_with = []
                                if afsc_scoring[a][c] > lowest_scoring_cadet_in_afsc[a]:
                                    afscs_that_desire_cadets_they_are_not_matched_with.append(a)
                                    afscs_that_desire_cadets_they_are_not_matched_with_set = \
                                        list(set(afscs_that_desire_cadets_they_are_not_matched_with))[0]
                                    cadets_other_AFSCs_desire_more[
                                        c] = afscs_that_desire_cadets_they_are_not_matched_with_set
                                    # print('afsc',a, 'scores cadet', c, 'higher than their lowest scored cadet',lowest_scoring_cadet_in_afsc[a] )

    # print("dict of cadet keys with list of afscs that desire the cadet more than the cadet's current afsc desires the cadet",cadets_other_AFSCs_desire_more)

    # Now lets find if we have blocking pairs. If the cadet prefers the AFSC and the AFSC prefers the cadet to it's lowest scored AFSC, then we have a blocking pair with the cadet and AFSC.

    d1 = set(cadets_other_AFSCs_desire_more.items())
    d2 = set(AFSCs_cadets_desire_morethan_current_afsc.items())
    blocking_pairs_set = d1.intersection(d2)

    # Convert the blocking pairs set to a dictionary so we can reference it later.
    blocking_pairs_dictionary = dict((x, y) for x, y in blocking_pairs_set)
    print('d1', d1)
    print('d2', d2)
    print('Blocking pairs dictionary:', blocking_pairs_dictionary)
    print('Total Blocking Pairs:', len(blocking_pairs_dictionary))


def OLEA_scoring_placeholder(instance='instance'):
    # Get directory path
    import os
    dir_path = os.getcwd() + '/'

    # Get main afccp folder path
    index = dir_path.find('afccp')
    dir_path = dir_path[:index + 6]

    # Update working directory
    os.chdir(dir_path)

    # Import main problem class
    from afccp.core.problem_class import CadetCareerProblem

    # Create new random instance
    instance = CadetCareerProblem("2023OLEA", printing=True)  # 'Random', N=10, M=2, P=2, printing=True)
    # instance.set_instance_value_parameters()
    # instance.set_instance_solution("Solution H")
    # instance.display_all_results_graphs()
    #
    # instance.classic_hr_alg()
    # instance.hr_alg_1_0_2_1()
    # instance.hr_alg_1_0_2_2()
    # instance.hr_lq_alg_2_0()
    # instance.hr_lq_alg_2_1()

    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections
    import matplotlib.pyplot as plt

    start_time = time.time()

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                afsc_scored_cadets.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

        AFSC_scores[a] = afsc_scored_cadets

        # Create a dictionary with AFSC and the scores from my CFM scoring matrix so we can compare each cadet's value to the values provided by OLEA. We can then turn the OLEA rankings into 0s where ineligible.

    afsc_dict = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_dict[a] = i

    # Created new dictionary of AFSCs and cadet scores using CFM scoring matrix using the OLEA 2023 data. We can now compare each cadet in this dictionary for the particular AFSC and assign a 0 score to all the cadets in the afsc_scoring sheet from OLEA based on who holds a 0 in this dictionary.
    list_of_AFSCs = afsc_dict.keys()
    list_of_AFSC_scores = list(AFSC_scores.values())
    AFSCs_scores = dict(zip(list_of_AFSCs, list_of_AFSC_scores))

    # Trying to create AFSC_scoring dict to use with OLEA's output to compare cadet positions, assign 0s, and remove ineligibles.

    ranking = {}
    # for a in range(instance.parameters['M']):
    for a in afsc_dict:
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSCs_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    afsc_scoring = {}
    for a in afsc_dict:
        afsc_scoring[a] = {}
        for i, v in enumerate(AFSCs_scores[a]):
            cadet = ranking[a][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSCs_scores[a][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # Read in the file from OLEA from the ranking's sheet.
    path = os.getcwd()
    data_file = 'Real 2023OLEA.xlsx'
    afsc_ranking_df = pd.read_excel(os.path.join(path, "afccp", "resources", "matching", "instances", data_file),
                                    sheet_name='AFSC Ranking', engine='openpyxl')
    # afsc_ranking_df = pd.read_excel(data_file, sheet_name='AFSC Ranking')
    afsc_ranking_df.set_index('Cadet', inplace=True)
    afsc_scoring_OLEA = afsc_ranking_df.to_dict()

    # Create dictionary of only eligible cadets for each AFSC from OLEA's roster.
    for a in afsc_dict:
        # afsc_scoring[a] = {}
        for c in afsc_scoring[a]:
            if c in afsc_scoring[a] and c in afsc_scoring_OLEA[a]:
                cadet_percentile_from_afsc = afsc_scoring_OLEA[a][c]
                afsc_scoring[a][c] = cadet_percentile_from_afsc

    sorted_percentiles = {}
    for a in afsc_dict:
        # print(a)
        cadets_in_order_list = []
        for k, v in sorted(afsc_scoring[a].items(), key=lambda item: item[1], reverse=True):
            # print('k:',k, 'v', v )
            cadets_in_order_list.append(k)
        sorted_percentiles[a] = cadets_in_order_list
    print(sorted_percentiles['13H'], '***************')
    print(sorted_percentiles.keys())

    # Now we can take the index of each cadet
    for a in afsc_dict:
        for i, cadet in enumerate(sorted_percentiles[a], start=1):
            cadet_index = i
            c = cadet
            afsc_scoring[a][c] = cadet_index

    # Assign percentile as the cadet's score
    for a in afsc_dict:
        for c in afsc_scoring[a]:
            percentile = afsc_scoring[a][c] / len(afsc_scoring[a])
            afsc_scoring[a][c] = percentile

    # End OLEA Edits - previous code picks up below.
    # ----------------------------------------------------------------------------------------------------------
    # ----------------------------------------------------------------------------------------------------------

    # Original cadet scoring build items needed to finish OLEA build.
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    cadet_score_from_AFSC = {}
    # c0_scores_list = []
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            # print(s)
            cadet_of_AFSC_score_list.append(s)
            # print(cadet_of_AFSC_score_list)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            # print()
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        # print(i)
        #     #print(cadet_score_from_AFSC[c])
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            # print(j)
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    # print(cadet_list)
    cadet_merit_list = [m for m in instance.parameters['merit']]

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]
    # ----------------------------------------------------------------------------------------------------------
    # ----------------------------------------------------------------------------------------------------------

    # Original HR algorithm with OLEA list

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    # Pseudo code for CAMP HR Algorithm
    # import tqdm from tqdm_notebook as tqdm
    # input data
    # cadet_ranking = {'c1': ['a20', 'a3',...]} #dictionary of cadets, value = ordered list of AFSCs
    # cadet_scoring = {'c1': {'a1':4.9, 'a2':3.8,...}} #dictionary of cadets, value = dictionary of AFSCs, value = AFSC score
    # AFSC_ranking = {'a1': ['c543', 'c32',...]} #dictionary of AFSCs, value = ordered list of cadets
    # AFSC_scoring = {'a1': {'c1': 4.6, 'c2': 5.8,...}} #dictionary of AFSC, value = dictionary of cadets, value = cadet score
    # afsc_matches = {'a1': ['c34', 'c66',...]} #dictionary of AFSCs, value = ordered list of matched cadets
    # AFSC_capacities = {'a1': 240, 'a2': 224, ...} #dictionary of AFSCs, value = AFSC capacity

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches
    print_details = False
    use_test_cadets = False
    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    # print(unmatched_ordered_cadets)
    test_cadets = ['c144', 'c82', 'c951', 'c1182', 'c460', 'c1183']
    if use_test_cadets:
        unmatched_ordered_cadets = test_cadets
    while len(unmatched_ordered_cadets) > 0 and iter_num <= iter_limit:  # while there are still cadets to match
        if print_details: print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
        if print_details: print('Length of perm unmatched cadets:', len(cannot_match))
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            if print_details: print('\nnext cadet', c)  # Can comment out
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
            # print(unmatched_ordered_cadets) # Can comment out
        if len(cadet_ranking[c]) == 0:  # if no more AFSCs in ranking list
            if print_details: print('Cannot match', c)
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
        print('Trying to match', c, a)
        if c == 'c1464':
            print('\n\n\nXXXXX ', c,
                  'XXXXXXXX-----------------------------------------------------------------------------------------------------------------------------\n\n\n')
        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            # if print_details: print("M", M)
            if print_details: print('Matched', c, 'and', a)
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            # afsc_matches = {13H: ["c1", "c2", "c3",...]}
            # sort afsc_matches[a] list based on afsc_ranking list of cadets
            # afsc_scoring[a][c] = s
            # print(afsc_ranking[a])
            #      afsc_matches[a] = [afsc_ranking[b] for _, b in zip(afsc_matches[a], afsc_ranking[a])]
            # TODO: sort matches by cadet scores, descending
            # print('matches 1:', afsc_matches) # Can comment out
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop
        else:  # if at capacity
            if print_details: print(a, "at capacity")
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            if print_details: print("AFSC_matches", afsc_matches[a])
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                if print_details: print(c_, 'higher than', c)
                # if print_details: print(cadet_ranking[c])
                if print_details: print("Line 61: Removing", a, 'from', c, 's list')
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                # if print_details: print(a, 'removed from', c,'s list')
                # M[c_] = a #added this define M[c_] so I could remove from it but c_ is already matched so this may be a problem.
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop
            else:  # if c ranks higher than c_
                if print_details: print(c_, 'lower than', c)
                if print_details: print('Line 69: Removing', c_, 'from', a)
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                if print_details: print("Removing", a, 'from', c_, 's list')
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                # if print_details: print("Cadet AFSC List:", cadet_ranking[c_])
                M.pop(c_, print(c_))  # remove c_ from M - Gives error that can't use pop to remove a string.
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                if len(unmatched_ordered_cadets) == 0 and next_cadet == False:  # added to get the last cadet added as unmatched
                    cannot_match.append(c)  # added to get the last cadet added as unmatched
                continue  # go to beginning of while loop

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    print(M)


def make_cadet_choice_chart(afsc_matches, Scoring_Build, instance='instance'):
    ''' This makes the bar plot that shows what choice the cadets received in their assigned matching'''

    import pandas as pd
    import matplotlib.pyplot as plt
    from collections import Counter
    import plotly.express as px

    # Must re-initialize the cadet_ranking dictionary because the matching algorithm deletes keys from it as it iterates.
    # Will need to call the build_ranking_data_structures function for this!

    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    # print(a)
                    afscs_from_cadets.append(a)
                    # print(afscs_from_cadets)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    cadets_matched_afsc_index = {}
    for a in afsc_matches:
        cadets_matched_afsc_index[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadets_matched_afsc_index[a][c] = cadet_index

    # print(cadets_matched_afsc_index)

    # Now gather the counts of each index for each AFSC. Another dict with key AFSC and key:value pairs of index: count of occurrence.

    AFSC_index_counts_dict = {}

    for a in afsc_matches:
        AFSC_index_counts_dict[a] = {}
        AFSC_index_counts = dict(Counter(cadets_matched_afsc_index[a].values()))
        AFSC_index_counts_dict[a] = AFSC_index_counts
        # print('afsc',a, 'with counts', AFSC_index_counts_dict, '\n')

    # Build a dictionary to access the number of cadets who got their "ith" choice. Constructed as keys of AFSCs, with a
    # key value pair of a cadet's ith choice and the number of cadets who got that ith choice for the particular AFSC.

    # # This dict is useful so you can look up each AFSC and pull the raw number of choices.
    # for a in afsc_matches:
    #     for k, v in AFSC_index_counts_dict[a].items():
    #         # print('a',a,'k', k, 'v:', v )
    #         AFSC_index_counts_dict[a][k] = v

    # Build dictionary for the chart to have a legend that reflects the ith choice.
    AFSC_index_counts_chart = {}
    for a in afsc_matches:
        AFSC_index_counts_chart[a] = {}
        for k, v in AFSC_index_counts_dict[a].items():
            k_str = "Choice " + str(k + 1)
            AFSC_index_counts_chart[a][k_str] = v

    plt.rcParams["figure.figsize"] = (10, 8)
    stacked_bar_chart_dict_df = pd.DataFrame(AFSC_index_counts_chart)
    stacked_bar_chart_dict_df = stacked_bar_chart_dict_df.T
    # Uncomment the below lines to make the matplotlib plot.
    # stacked_bar_chart_dict_df['AFSCs'] = AFSC_index_counts_dict.keys()
    # stacked_bar_chart_dict_df.plot.bar(x='AFSCs', stacked=True, title='Test')
    # plt.legend(["1st Choice", '3rd Choice', '4th Choice', '8th Choice', '10th Choice', '11th Choice', '12th Choice'])
    # plt.show()

    # Plotly graph below

    # del stacked_bar_chart_dict_df['AFSCs']

    # Make plotly graph below! This makes a wide form graph.

    fig = px.bar(stacked_bar_chart_dict_df, title="Cadet Choice Chart",
                 labels={"value": "Number of Cadets", "index": "AFSC", "variable": "Cadet Preference"})
    fig.show()


def make_cadet_degree_chart(Scoring_Build, afsc_matches, instance='instance'):
    import plotly.express as px

    # Will need to re-initialize the cadet scoring matrix to remove cadet's preference with a score of 7 or greater.

    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                # cadet_scores.append(instance.parameters['merit'][c]*0)
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # The below builds the score that a cadet assigns to each AFSC. Omits if voluntary with a score of 7.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        # print(np.where(instance.parameters['utility'][c] > 0))
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        # print(c)
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            score = cadet_score_from_AFSC[i][j]
            if score > 0:
                cadet_scoring[c][a] = score

    cadet_degree_matches = {}
    for a in afsc_matches:
        cadet_degree_matches[a] = {}
        for c in afsc_matches[a]:
            if a in cadet_scoring[c]:
                if 5 <= cadet_scoring[c][a] < 6:
                    cadet_degree_matches[a][c] = 'Mandatory Vol'
                if 4 <= cadet_scoring[c][a] < 5:
                    cadet_degree_matches[a][c] = 'Desired Vol'
                if 3 <= cadet_scoring[c][a] < 4:
                    cadet_degree_matches[a][c] = 'Permitted Vol'
                if 2 <= cadet_scoring[c][a] < 3:
                    cadet_degree_matches[a][c] = 'Mandatory Non-Vol'
                if 1 <= cadet_scoring[c][a] < 2:
                    cadet_degree_matches[a][c] = 'Desired Non-Vol'
                if 0.5 <= cadet_scoring[c][a] < 1:
                    cadet_degree_matches[a][c] = 'Permitted Non-Vol'

    cadet_degree_counts_dict = {}

    for a in afsc_matches:
        cadet_degree_counts_dict[a] = {}
        cadet_degree_counts = dict(Counter(cadet_degree_matches[a].values()))
        cadet_degree_counts_dict[a] = cadet_degree_counts

    cadet_degree_counts_chart = {}
    for a in afsc_matches:
        cadet_degree_counts_chart[a] = {}
        for k, v in cadet_degree_counts_dict[a].items():
            cadet_degree_counts_chart[a][k] = v

    cadet_degree_stacked_bar_chart_dict_df = pd.DataFrame(cadet_degree_counts_chart)
    cadet_degree_stacked_bar_chart_dict_df = cadet_degree_stacked_bar_chart_dict_df.T

    # Plotly graph below

    fig_degree = px.bar(cadet_degree_stacked_bar_chart_dict_df, title="Cadet Degree Tier with AFSC Match",
                        labels={"value": "Number of Cadets", "index": "AFSC",
                                "variable": "Cadet Degree Tier and Vol Status"})
    # fig_degree.show()


def make_cadet_merit_choice_scatter(afsc_matches, Scoring_Build):
    import scipy.stats as stats
    from scipy.stats import spearmanr  # Spearman correlation

    # Make a several dictionaries and then combine them together to make the desired dataframe.
    # Dict 1 = cadet: index choice
    # Dict 2 = cadet: merit value
    # Dict 3 = cadet: afsc they are matched to

    # Must initialize cadet ranking dictionary again because we eliminate values in the HR algorithm. This is used in the cadet choice dict.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    # print(a)
                    afscs_from_cadets.append(a)
                    # print(afscs_from_cadets)

            # print(afscs_from_cadets)
            # print(AFSC)
            cadet_ranking[c] = afscs_from_cadets

    # Dict 1
    cadet_choice_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_index = cadet_ranking[c].index(a)
                cadet_index += 1
                cadet_choice_dict[c] = cadet_index

    # Dict 2
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 3
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    merit_vs_choice_chart_df = pd.DataFrame(
        {'Cadet Choice': pd.Series(cadet_choice_dict), 'Cadet Merit': pd.Series(cadet_merit_dict),
         'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    merit_vs_choice_chart_df.head()

    fig = px.scatter(merit_vs_choice_chart_df, x="Cadet Choice", y="Cadet Merit", color="Matched AFSC")
    fig.show()

    # Is it significantly correlated? Null: the samples are uncorrelated.
    # If p > .05, fail to reject Null - variables are NOT correlated
    # If p < .05, fail to accept Null - variables are correlated

    # These variables are NOT independent and the cadet choice values are NOT normally distributed so we can't use these tests.
    # (Spearman and two sample t-test)

    coef, p = spearmanr(merit_vs_choice_chart_df["Cadet Choice"], merit_vs_choice_chart_df["Cadet Merit"])
    print(coef)
    print(p)


def make_cadet_merit_box_plots(instance, Scoring_Build, afsc_matches):
    # Must initialize cadet ranking dictionary again because we eliminate values in the HR algorithm. This is used in the cadet choice dict.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        # print(i)
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)
            cadet_ranking[c] = afscs_from_cadets

    # Dict 1
    cadet_merit_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                c_merit = int(c[1:])
                cadet_merit = instance.parameters['merit'][c_merit]
                cadet_merit_dict[c] = cadet_merit

    # Dict 2
    cadet_to_matched_afsc_dict = {}
    for a in afsc_matches:
        for c in afsc_matches[a]:
            if a in cadet_ranking[c]:
                cadet_to_matched_afsc_dict[c] = a

    box_merit_plot_df = pd.DataFrame(
        {'Cadet Merit': pd.Series(cadet_merit_dict), 'Matched AFSC': pd.Series(cadet_to_matched_afsc_dict)})

    box_fig = px.box(box_merit_plot_df, y="Cadet Merit", x="Matched AFSC")
    box_fig.show()


def match_to_excel(M, file_name='Real 2023 MacDonald Results.xlsx', solution_name='HR Lower Quotas w/ Max Capacities'):
    """
    Script for appending sheet with data to an existing Excel workbook.
    """
    import time
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import collections

    new_M = {}
    for c in M.keys():
        i = int(c[1:])
        new_M[i] = M[c]  # getting the index of M

    file_name = 'Real 2023 MacDonald Results.xlsx'
    ordered_M = collections.OrderedDict(sorted(new_M.items()))
    path = os.getcwd()

    results_by_index = []
    for i in ordered_M.keys():
        results_by_index.append(ordered_M[i])
    print(results_by_index)
    # save current solutions
    # df_solutions = pd.read_excel(file_name, sheet_name='Solutions', engine='openpyxl', index_col=0)
    df_solutions = pd.read_excel(os.path.join(path, "afccp", "matching", file_name), sheet_name='Solutions',
                                 engine='openpyxl', index_col=0)

    # append new solution to current solutions
    solution_name = 'HR Lower Quotas w/ Max Capacities'
    if solution_name in df_solutions.columns:
        solution_name += '_' + datetime.now().strftime("%m%d_%H:%M:%S")

    matching_list = results_by_index
    df_new_solution = pd.DataFrame({solution_name: matching_list})
    print(df_new_solution)
    df_solutions = df_solutions.merge(df_new_solution, how='left', left_index=True, right_index=True)
    print(df_solutions)

    # df_solutions = df_solutions.merge(df_new_solution, how='left', left_on='HR Lower Quotas w/ Unlimited Capacities', right_on='HR Lower Quotas w/ Max Capacities')
    # The above line did append but it was the exact same column and went from 0 to 183,222 cadets.

    # save current and new solutions by deleting current sheet and creating a new one
    # ExcelWorkbook = load_workbook(file_name)
    ExcelWorkbook = load_workbook(os.path.join(path, "afccp", "matching", file_name))
    if 'Solutions' in ExcelWorkbook.sheetnames:
        ExcelWorkbook.remove(ExcelWorkbook['Solutions'])

    writer = pd.ExcelWriter(os.path.join(path, "afccp", "matching", file_name), mode='w',
                            engine='openpyxl')  # Had to manually give directory
    writer.book = ExcelWorkbook
    df_solutions.to_excel(writer, sheet_name='Solutions')
    # df_new_solution.to_excel(writer, sheet_name='Solutions')
    writer.save()
    writer.close()


# Added the four below to test exporting.

def test_function_1(instance):
    # Just to show you that you have all the problem instance data
    p = instance.parameters
    print("Hello! Your problem instance has " + str(p["N"]) + " cadets.")


def classic_hr_alg(instance):
    """
    This is the Classic Hospital/Residents Algorithm coded by Ian MacDonald
    :param instance:
    :return: Solution
    """

    # # I like to abbreviate this kind of thing
    # p = instance.parameters
    #
    # ########### THIS SECTION IS JUST TO SHOW YOU SOME EXAMPLES, DELETE IT LATER ####################
    # # If you have something like this:
    # afsc_solution = np.array(["15A", "63A", "14N"])  # For all N cadets..
    #
    # # Just do this  (np.where is a really nice function)
    # solution = np.array([np.where(p["afsc_vector"] == afsc)[0][0] for afsc in afsc_solution])
    #
    # # Just for example
    # print("Example Solution:", afsc_solution, "with indices", solution)
    #
    # # (Just to get an actual solution here that works for this example)
    # solution = instance.stable_matching(set_to_instance=False, printing=False, add_to_dict=False)
    # ########### THIS SECTION IS JUST TO SHOW YOU SOME EXAMPLES, DELETE IT LATER ####################

    # return solution
    start_time = time.time()
    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # My Score Calculations for CFMs
    import numpy as np
    np.random.seed(2)

    instance_creation_time = time.time()
    print("Time to create instance", instance_creation_time - start_time)

    # How AFSCs rank cadets within each AFSC:

    AFSC_scores = {}
    for a in range(instance.parameters['M']):
        afsc_scored_cadets = []
        for c in range(instance.parameters['N']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                afsc_scored_cadets.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(2 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(1 + instance.parameters['merit'][c] / 100)
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                afsc_scored_cadets.append(0.5 + instance.parameters['merit'][c] / 100)
            else:
                afsc_scored_cadets.append(0)
            # If AFSC score is below 0 then remove them from matching possibilities.

            AFSC_scores[a] = afsc_scored_cadets

    # Scores amended to break the ties for non-vol'd with a degree. My only concern is that cadet's with a higher merit are now more likely to get placed in a non-vol position than someone with lower merit who may end up in the "second matching".
    cadet_scores = {}
    for c in range(instance.parameters['N']):
        cadet_scored_afscs = []
        for a in range(instance.parameters['M']):
            if instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(5 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(4 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] > 0:
                cadet_scored_afscs.append(3 + instance.parameters['utility'][c][a] * instance.parameters['merit'][c])
            elif instance.parameters['mandatory'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(2 + np.random.random())
            elif instance.parameters['desired'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(1 + np.random.random())
            elif instance.parameters['permitted'][c][a] > 0 and instance.parameters['utility'][c][a] == 0:
                cadet_scored_afscs.append(0.5 + np.random.uniform(0, .09))
            else:
                cadet_scored_afscs.append(0)
            # If cadet score is below 0 then remove them from matching possibilities.

            cadet_scores[c] = cadet_scored_afscs

    # Try to make full list of all cadets in merit order for unmatched cadet list
    cadet_list = ['c' + str(c) for c in range(instance.parameters['N'])]
    cadet_merit_list = [m for m in instance.parameters['merit']]
    sorted_merit_list = sorted(cadet_merit_list, reverse=True)

    unmatched_ordered_cadets = [cadet_list for (cadet_merit_list, cadet_list) in
                                sorted(zip(cadet_merit_list, cadet_list), key=lambda x: x[0])]

    # This is the functional ranking structure. I put this in as a function under data data. This is the AFSCs ranking of cadets. AFSC_ranking relies on this sorted dictionary to build that dictionary in order of preference.
    ranking = {}
    for a in range(instance.parameters['M']):
        sorted_scores = sorted(((v, i) for i, v in enumerate(AFSC_scores[a])), reverse=True)
        result = []
        for i, (value, index) in enumerate(sorted_scores):
            result.append(index)
        ranking[a] = result

    # G2G - This is relied upon to  build the other dictionaries as well.
    # The below builds the score that a cadet assigns to each AFSC.
    cadet_score_from_AFSC = {}
    for c in range(instance.parameters['N']):
        cadet_of_AFSC_score_list = []
        for a in AFSC_scores.keys():
            s = cadet_scores[c][a]
            cadet_of_AFSC_score_list.append(s)
            cadet_score_from_AFSC[c] = cadet_of_AFSC_score_list

    # Try to eliminate cadets who preferred AFSCs they weren't eligible for this dictionary.

    for c in range(instance.parameters['N']):
        AFSC_list = np.where(instance.parameters['utility'][c] > 0)
        for a in AFSC_list[0]:
            if cadet_score_from_AFSC[c][a] >= 0 and AFSC_scores[a][c] < .01:
                cadet_score_from_AFSC[c][a] = -1
            else:
                cadet_score_from_AFSC[c][a] = instance.parameters['utility'][c][a] + 7

    cadet_scoring = {}
    for i in range(instance.parameters['N']):
        c = 'c' + str(i)
        cadet_scoring[c] = {}
        for j in range(instance.parameters['M']):
            a = instance.parameters['afsc_vector'][j]
            # print(a)
            score = cadet_score_from_AFSC[i][j]
            if score > 0:  # added
                cadet_scoring[c][a] = score  # added

    cadet_ranked_AFSC = {}
    for c in range(instance.parameters['N']):
        sorted_cadet_score_from_AFSC = sorted(((v, i) for i, v in enumerate(cadet_score_from_AFSC[c])), reverse=True)
        result2 = []
        for i, (value, index) in enumerate(sorted_cadet_score_from_AFSC):
            result2.append(index)
        cadet_ranked_AFSC[c] = result2

    # G2G. Removes ineligibles! Relies on the sorted and ordered cadet_ranked_AFSC dictionary above to work properly.
    cadet_ranking = {}
    for i in range(instance.parameters['N']):
        afscs_from_cadets = []
        c = 'c' + str(i)
        for j in range(instance.parameters['M']):
            AFSC = cadet_ranked_AFSC[i][j]
            # Sorted because of the cadet_ranked_AFSC dictionary. Put all values in place based on the indices.
            a = instance.parameters['afsc_vector'][AFSC]
            if a in cadet_scoring[c].keys():
                if cadet_scoring[c][a] > 0:
                    afscs_from_cadets.append(a)

            cadet_ranking[c] = afscs_from_cadets

    # G2G, Functional and removed ineligibles. Relies on the ranking dictionary near the very top of this code to create this dictionary.

    afsc_scoring = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        afsc_scoring[a] = {}
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            cadet_score_of_afsc = AFSC_scores[j][i]
            if cadet_score_of_afsc > 0:  # eliminate all ineligble cadets
                c = 'c' + str(i)
                afsc_scoring[a][c] = cadet_score_of_afsc

    # G2G. Removed ineligibles! Relies on the ranking dictionary near the very top of this code to create this dictionary in order of preferred cadets.

    afsc_ranking = {}
    for j in range(instance.parameters['M']):
        cadets_from_afscs = []
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            cadet = ranking[j][i]
            c = 'c' + str(cadet)
            if c in afsc_scoring[a].keys():
                if afsc_scoring[a][c] > 0:
                    cadets_from_afscs.append(c)

            afsc_ranking[a] = cadets_from_afscs

    # #Actual Max Capacities for AFSCs
    afscs = instance.parameters['afsc_vector']
    capacities = instance.parameters["quota_max"]
    afsc_capacities = dict(zip(afscs, capacities))

    # #Modified capacities with the lower PGL target set as the overall capacity for each AFSC
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [8, 19, 105, 7, 195, 35, 25, 181, 84, 29, 61, 29, 3, 7, 3, 3, 42, 3, 18, 84, 1, 13, 15, 12, 20, 63, 34, 12, 2, 69, 50, 34]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    # #Modified capacities with the increased max target for each AFSC to get remaining 25 cadets matched from the first run
    # afscs = instance.parameters['afsc_vector']
    # pgl_capacities = [14, 27, 210, 8, 218, 72, 50, 196, 92, 38, 68, 37, 6, 10, 6, 5, 60, 5, 23, 94, 2, 14, 30, 24, 40, 126, 48, 24, 4, 101, 77, 38]
    # afsc_capacities = dict(zip(afscs, pgl_capacities))
    # afsc_capacities

    afsc_matches = {}
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        for i in range(instance.parameters['N']):
            i = []
        afsc_matches[a] = i

    matching_data_structure_creation = time.time()
    print("Time to create matching data structures:", matching_data_structure_creation - instance_creation_time)

    # for c in cadet_ranking: # MODIFICATION
    #    cadet_ranking[c] = cadet_ranking[c][:6] # Looking at only the cadet's first 6 choices

    # initialize parameters
    unmatched_ordered_cadets = unmatched_ordered_cadets  # all cadets sorted decreasing by merit
    cannot_match = []  # cadets with empty ranking lists
    M = {}  # matches
    print_details = True
    use_test_cadets = False
    iter_num = 0
    iter_limit = 10000
    next_cadet = True  # tracks if it is a new cadet or the same one that got rejected
    # print(unmatched_ordered_cadets)
    test_cadets = ['c144', 'c82', 'c951', 'c1182', 'c460', 'c1183']
    if use_test_cadets:
        unmatched_ordered_cadets = test_cadets
    while len(unmatched_ordered_cadets) > 0 and iter_num <= iter_limit:  # while there are still cadets to match
        if print_details: print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
        if print_details: print('Length of perm unmatched cadets:', len(cannot_match))
        iter_num += 1
        if next_cadet:
            c = unmatched_ordered_cadets[0]  # next cadet to match
            if print_details: print('\nnext cadet', c)  # Can comment out
            unmatched_ordered_cadets.remove(c)  # take cadet out of unmatched list
            # print(unmatched_ordered_cadets) # Can comment out
        if len(cadet_ranking[c]) == 0:  # if no more AFSCs in ranking list
            if print_details: print('Cannot match', c)
            cannot_match.append(c)  # add them to cannot match list
            next_cadet = True
            continue  # go to beginning of while loop
        a = cadet_ranking[c][0]  # highest ranking AFSC for cadet c
        print('Trying to match', c, a)
        if c == 'c1464':
            print('\n\n\nXXXXX ', c,
                  'XXXXXXXX-----------------------------------------------------------------------------------------------------------------------------\n\n\n')
        if len(afsc_matches[a]) < afsc_capacities[a]:  # if not at capacity
            M[c] = a  # add match to M
            # if print_details: print("M", M)
            if print_details: print('Matched', c, 'and', a)
            afsc_matches[a].append(c)  # insert into AFSC a's matches
            # afsc_matches = {13H: ["c1", "c2", "c3",...]}
            # sort afsc_matches[a] list based on afsc_ranking list of cadets
            # afsc_scoring[a][c] = s
            # print(afsc_ranking[a])
            #      afsc_matches[a] = [afsc_ranking[b] for _, b in zip(afsc_matches[a], afsc_ranking[a])]
            # print('matches 1:', afsc_matches) # Can comment out
            next_cadet = True  # move to next cadet
            continue  # go to beginning of while loop
        else:  # if at capacity
            if print_details: print(a, "at capacity")
            c_ = afsc_matches[a][-1]  # set c_ as the lowest ranking cadet from AFSC list
            if print_details: print("AFSC_matches", afsc_matches[a])
            for c_hat in afsc_matches[a]:
                if afsc_scoring[a][c_hat] < afsc_scoring[a][c_]:
                    c_ = c_hat
            if afsc_scoring[a][c_] > afsc_scoring[a][c]:  # if c_ ranks higher than c
                if print_details: print(c_, 'higher than', c)
                # if print_details: print(cadet_ranking[c])
                if print_details: print("Line 61: Removing", a, 'from', c, 's list')
                cadet_ranking[c].remove(a)  # remove a from c's ranking list
                # if print_details: print(a, 'removed from', c,'s list')
                # M[c_] = a #added this define M[c_] so I could remove from it but c_ is already matched so this may be a problem.
                next_cadet = False  # keep trying to match this cadet
                continue  # go to beginning of while loop
            else:  # if c ranks higher than c_
                if print_details: print(c_, 'lower than', c)
                if print_details: print('Line 69: Removing', c_, 'from', a)
                afsc_matches[a].remove(c_)  # remove c_ from AFSC a's matches
                if c_ == 'c144': print(cadet_ranking[c_],
                                       '-----------------******************--------------------------------')
                if print_details: print("Removing", a, 'from', c_, 's list')
                cadet_ranking[c_].remove(a)  # Removing AFSC from cadet c_'s list
                # if print_details: print("Cadet AFSC List:", cadet_ranking[c_])
                M.pop(c_, print(c_))  # remove c_ from M - Gives error that can't use pop to remove a string.
                M[c] = a  # add match to M
                next_cadet = False
                afsc_matches[a].append(c)  # insert into AFSC a's matches
                c = c_
                if len(unmatched_ordered_cadets) == 0 and next_cadet == False:  # added to get the last cadet added as unmatched
                    cannot_match.append(c)  # added to get the last cadet added as unmatched
                continue  # go to beginning of while loop

    print('\nM', M)
    print("Total iterations:", iter_num)

    Classical_HR_algorithm_time = time.time()
    print("Time to run matching algorithm:", Classical_HR_algorithm_time - matching_data_structure_creation)

    print('Length of cadets still remaining to match:', len(unmatched_ordered_cadets))
    print('Length of perm unmatched cadets:', len(cannot_match))

    print("Total time:", time.time() - start_time)

    # Trying to loop through those who could not be matched to match them with the career field that wants them.
    print('# unmatched', len(cannot_match))

    afsc_matches_len = []
    for a in afsc_matches:
        a = ('Matched:', len(afsc_matches[a]))
        afsc_matches_len.append(a)

    quota_min = []
    for q in instance.parameters["quota_min"]:
        q = ('min:', q)
        quota_min.append(q)

    max_quotas = []
    for m in instance.parameters["quota_max"]:
        m = ('max:', m)
        max_quotas.append(m)

    quota_min_and_match_results = list(zip(afscs, afsc_matches_len, quota_min, max_quotas))
    print(quota_min_and_match_results)

    return M, afsc_matches, cadet_scoring


def classic_hr_alg_evaluate(M, afsc_matches, cadet_scoring, instance):
    """
    This function evaluates an AFSC cadet solution using the Classic HR algorithm evaluation metrics
    :param instance:
    :return:
    """
    # # I like to abbreviate this kind of thing
    # p = instance.parameters
    #
    # # Get your H/R metrics (can be called whatever you want inside this function)
    # m = np.zeros(p["M"])
    # for j in p["J"]:
    #     m[j] = np.random.normal(50, 10)
    # metrics = {"Random Stuff": m}
    #
    # meaningless_metric = np.mean(metrics["Random Stuff"])
    #
    # # Print statement
    # print("Evaluated H/R Algorithm with random metric of", meaningless_metric)

    # Griff's example above#-----------------------#-------------------------------#

    # AFSC Merit Dictionary
    afsc_avg_merit = {}
    list_of_cadet_merit_in_matched_afsc = []
    total_num_cadets = 0

    for a in afsc_matches:
        for c in afsc_matches[a]:
            total_num_cadets += 1
            c = int(c[1:])
            list_of_cadet_merit_in_matched_afsc.append(instance.parameters['merit'][c])
        afsc_summed_merit = sum(list_of_cadet_merit_in_matched_afsc)
        afsc_avg_cadet_merit = afsc_summed_merit / total_num_cadets
        afsc_avg_merit[a] = afsc_avg_cadet_merit

    # # and % of cadets who got one of their top 6 choices
    # Create dictionary of number/% of cadet who DID get top 6 (i.e. voluntary) and had a mandatory degree
    cadets_wMand_matched_with_top_6_afscs = {}
    percent_cadets_wMand_matched_with_top_6_afscs = {}
    for a in afsc_matches:
        num_cadets_wMand_matched_with_top_6_afscs = 0
        for c in afsc_matches[a]:
            if cadet_scoring[c][a] > 5:
                num_cadets_wMand_matched_with_top_6_afscs += 1
                cadets_wMand_matched_with_top_6_afscs[a] = num_cadets_wMand_matched_with_top_6_afscs
                percent_cadets_wMand_matched_with_top_6_afscs[a] = num_cadets_wMand_matched_with_top_6_afscs / len(
                    afsc_matches[a]) * 100

    # Create dictionary of number/% of cadet who DID NOT get top 6 (i.e. non-vol'd) and had a mandatory degree
    cadets_wMand_matched_pref_7_or_greater = {}
    percent_cadets_wMand_matched_pref_7_or_greater = {}

    for a in afsc_matches:
        num_cadets_wMand_matched_pref_7_or_greater = 0
        for c in afsc_matches[a]:
            if cadet_scoring[c][a] >= 2 and cadet_scoring[c][a] < 3:
                num_cadets_wMand_matched_pref_7_or_greater += 1
                cadets_wMand_matched_pref_7_or_greater[a] = num_cadets_wMand_matched_pref_7_or_greater
                percent_cadets_wMand_matched_pref_7_or_greater[a] = num_cadets_wMand_matched_pref_7_or_greater / len(
                    afsc_matches[a]) * 100

        # if cadet_scoring[c][a] >= 3:
        #     cadets_wMand_matched_pref_7_or_greater[a] = 0
        #     percent_cadets_wMand_matched_pref_7_or_greater[a] = 0

    # TODO # I would like to include all AFSCs, including those with 0 values. What should I pass in here to keep 62EXE from going to 0?

    # Create dictionary of number and percent of cadets who did not get their top 6.

    cadets_matched_pref_7_or_greater = {}
    percent_cadets_matched_pref_7_or_greater = {}

    for a in afsc_matches:
        num_cadets_matched_pref_7_or_greater = 0
        for c in afsc_matches[a]:
            if cadet_scoring[c][a] < 7:
                num_cadets_matched_pref_7_or_greater += 1
                cadets_matched_pref_7_or_greater[a] = num_cadets_matched_pref_7_or_greater
                percent_cadets_matched_pref_7_or_greater[a] = num_cadets_matched_pref_7_or_greater / len(
                    afsc_matches[a]) * 100
        if cadet_scoring[c][a] > 7:
            cadets_matched_pref_7_or_greater[a] = 0
            percent_cadets_matched_pref_7_or_greater[a] = 0

    # Create dictionary of the number and one of the percent of cadets who received one of their top 6 AFSCs.
    cadets_matched_with_top_6_afscs = {}
    percent_cadets_matched_with_top_6_afscs = {}
    for a in afsc_matches:
        num_cadets_matched_with_top_6 = 0
        for c in afsc_matches[a]:
            if cadet_scoring[c][a] > 7:
                num_cadets_matched_with_top_6 += 1
                cadets_matched_with_top_6_afscs[a] = num_cadets_matched_with_top_6
                percent_cadets_matched_with_top_6_afscs[a] = num_cadets_matched_with_top_6 / len(
                    afsc_matches[a]) * 100

    # PGL quota of USAFA cadets. What % over or under is each AFSC?
    desired_pgl_percent_cadets_USAFA_per_AFSC = {}

    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        num_USAFA_per_AFSC = 0

        for c in afsc_matches[a]:
            c = int(c[1:])

            if instance.parameters["usafa"][c] == 1:
                num_USAFA_per_AFSC += 1

            for j in range(instance.parameters['M']):
                desired_pgl_percent_cadets_USAFA_per_AFSC_list = []
                desired_pgl_percent_cadets_USAFA_per_AFSC_list.append(
                    (num_USAFA_per_AFSC / instance.parameters["usafa_quota"][j]) * 100)
                desired_pgl_percent_cadets_USAFA_per_AFSC[a] = desired_pgl_percent_cadets_USAFA_per_AFSC_list

    # % of AFSCs that are USAFA cadets
    usafa_percent_per_afsc = {}

    for a in afsc_matches:
        usafa_count = 0
        for c in afsc_matches[a]:
            if instance.parameters['usafa'][int(c[1:])] == 1:
                usafa_count += 1
                usafa_percent_per_afsc[a] = 100 * usafa_count / len(afsc_matches[a])

    # PGL quota of ROTC cadets. What % over or under is each AFSC?
    desired_pgl_percent_cadets_ROTC_per_AFSC = {}
    # num_ROTC_per_AFSC = 0
    for j in range(instance.parameters['M']):
        a = instance.parameters['afsc_vector'][j]
        # print(a)
        num_ROTC_per_AFSC = 0
        for c in afsc_matches[a]:
            c = int(c[1:])
            # print(afsc_matches[a])

            if instance.parameters["usafa"][c] == 0:
                num_ROTC_per_AFSC += 1
            for j in range(instance.parameters['M']):
                # a = instance.parameters['afsc_vector'][j]
                desired_pgl_percent_cadets_ROTC_per_AFSC_list = []
                desired_pgl_percent_cadets_ROTC_per_AFSC_list.append(
                    (num_ROTC_per_AFSC / instance.parameters["rotc_quota"][j]) * 100)
                desired_pgl_percent_cadets_ROTC_per_AFSC[a] = desired_pgl_percent_cadets_ROTC_per_AFSC_list
    # % of AFSCs that are ROTC cadets
    rotc_percent_per_AFSC = {}

    for a in afsc_matches:
        rotc_count = 0
        for c in afsc_matches[a]:
            if instance.parameters['usafa'][int(c[1:])] == 0:
                rotc_count += 1
        rotc_percent_per_AFSC[a] = 100 * rotc_count / len(afsc_matches[a])
    print('Percent of AFSC that is ROTC Cadets:', rotc_percent_per_AFSC)

    metrics = (afsc_avg_merit, cadets_wMand_matched_with_top_6_afscs, percent_cadets_wMand_matched_with_top_6_afscs,
               cadets_wMand_matched_pref_7_or_greater, percent_cadets_wMand_matched_pref_7_or_greater,
               cadets_matched_pref_7_or_greater,
               percent_cadets_matched_pref_7_or_greater, cadets_matched_with_top_6_afscs,
               percent_cadets_matched_with_top_6_afscs,
               desired_pgl_percent_cadets_USAFA_per_AFSC, usafa_percent_per_afsc,
               desired_pgl_percent_cadets_ROTC_per_AFSC, rotc_percent_per_AFSC)

    return metrics


def ian_export(instance, printing=True):
    """
    This procedures exports the specific problem instance/solution combination to excel
    :return: None.
    """

    if printing:
        print("Exporting to excel...")

    p = instance.parameters
    hr = instance.hr_metrics
    # hr = hr_metrics

    # Construct fixed parameter dataframes
    cadets_fixed, afscs_fixed = model_data_frame_from_fixed_parameters(p)

    # Get the filepath
    instance.full_name = instance.data_type + " " + instance.data_name + " " + instance.vp_name + " " + instance.solution_name
    filepath = paths_out['instances'] + instance.full_name + '.xlsx'

    # Your new dataframe to go on an excel sheet
    new_df = pd.DataFrame({'AFSC': p['afsc_vector'], "Random Metrics": hr["Random Stuff"]})

    # Your new dataframe for the solutions excel sheet
    df_solutions = pd.DataFrame({'Solution'})

    # Build the solution metrics dataframes if need be
    # if metrics is not None:
    #
    #     cadet_solution_df = pd.DataFrame({'Cadet': parameters['ID'], 'Matched': metrics['afsc_solution'],
    #                                       'Value': metrics['cadet_value'],
    #                                       'Weight': value_parameters['cadet_weight'],
    #                                       'Value Fail': metrics['cadet_constraint_fail']})
    #
    #     objective_measures = pd.DataFrame({'AFSC': parameters['afsc_vector']})
    #     objective_values = pd.DataFrame({'AFSC': parameters['afsc_vector']})
    #     afsc_constraints_df = pd.DataFrame({'AFSC': parameters['afsc_vector']})
    #     for k in range(value_parameters['O']):
    #         objective_measures[value_parameters['objectives'][k]] = metrics['objective_measure'][:, k]
    #         objective_values[value_parameters['objectives'][k]] = metrics['objective_value'][:, k]
    #         afsc_constraints_df[value_parameters['objectives'][k]] = metrics['objective_constraint_fail'][:, k]
    #
    #     objective_values['AFSC Value'] = metrics['afsc_value']
    #     afsc_constraints_df['AFSC Value Fail'] = metrics['afsc_constraint_fail']
    #
    #     metric_names = ['Z', 'Cadet Value', 'AFSC Value', 'Num Ineligible', 'Failed Constraints']
    #     metric_results = [metrics['z'], metrics['cadets_overall_value'], metrics['afscs_overall_value'],
    #                       metrics['num_ineligible'], metrics['total_failed_constraints']]
    #     for k, objective in enumerate(value_parameters['objectives']):
    #         metric_names.append(objective + ' Score')
    #         metric_results.append(metrics['objective_score'][k])
    #
    #     overall_solution = pd.DataFrame({'Solution Metric': metric_names, 'Result': metric_results})

    with pd.ExcelWriter(filepath) as writer:  # Export to excel
        cadets_fixed.to_excel(writer, sheet_name="Cadets Fixed", index=False)
        afscs_fixed.to_excel(writer, sheet_name="AFSCs Fixed", index=False)
        new_df.to_excel(writer, sheet_name="Example Metrics", index=False)
        df_solutions.to_excel(writer, sheet_name="Solutions", index=False)  # Added for solutions

        # if value_parameters is not None:
        #     overall_weights_df.to_excel(writer, sheet_name="Overall Weights", index=False)
        #     cadet_weights_df.to_excel(writer, sheet_name="Cadet Weights", index=False)
        #     afsc_weights_df.to_excel(writer, sheet_name="AFSC Weights", index=False)
        # if metrics is not None:
        #     cadet_solution_df.to_excel(writer, sheet_name="Cadet Solution Quality", index=False)
        #     objective_measures.to_excel(writer, sheet_name="AFSC Objective Measures", index=False)
        #     objective_values.to_excel(writer, sheet_name="AFSC Solution Quality", index=False)
        #     afsc_constraints_df.to_excel(writer, sheet_name="AFSC Constraint Fails", index=False)
        #     overall_solution.to_excel(writer, sheet_name="Overall Solution Quality", index=False)
